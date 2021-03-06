import os
import time
# import wmi,json
# import pythoncom
import MySQLdb.constants.ER
import pandas as pd

# # import clr
# # clr.AddReference("iMDS_AES")
# #
# # # clr.AddReference("iMDS_AES")
# # from iMDS_AES import *
# from Crypto.Cipher import DES
# import base64


from django.shortcuts import render
from django.http import HttpResponse,FileResponse
from django.urls.conf import re_path
from django.views import View
import json
import logging
import datetime
# from utils.import_common import test11
# from utils.test import opt
# # from utils import *
import shutil
import xlrd
import xlwt
from openpyxl.reader.excel import load_workbook


import openpyxl
from django.contrib.auth.hashers import make_password, check_password
from common.product_live import product_live
from common.db import DB
from page.Page import Page
from pathlib import Path
logger = logging.getLogger(__name__)
# logger.setLevel(logging.INFO)

# Create your views here.


class test2(View):

    # def get_live_data(self, product_id, finished_product_code):
    #     func = product_live()
    #     data_list = func.one_product_live(product_id, finished_product_code)
    #
    #     return data_list

    def get(self, request):
        try:
            db = DB()
            product_id = "co2"
            conn = db.get_connection(product_id)
            work_code = "Im_WorkStation_5"
            sql_search_type = """
               select work_id, work_type from work_station where work_code = '{0}'
               """
            sql_search_type = sql_search_type.format(work_code)
            drs = db.execute_sql(conn, sql_search_type)
            print("======>", drs)
            print("---->", len(drs))

            # sql_search_type = """
            #                select * from unqualified_product_Station_085
            #                """
            # # sql_search_type = sql_search_type.format(work_code)
            # drs = db.execute_sql(conn, sql_search_type)
            # print("======>", drs)
            # print("---->", len(drs))






            # if len(drs) > 0:
            #     if drs[0][1] == "??????":
            #         select_table = "product_transit_" + str(drs[0][0])
            #         sql_check = """
            #         select * from {0}
            #         """


            result = {"code": 0, "message": "????????????"}
            result = json.dumps(result)
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????"}
            result = json.dumps(result)
            logger.error("????????????")
            return HttpResponse(result)




class LogisticLogin(View):

    """
    ???????????????  ?????????????????????????????? ?????????????????????
    AEC ??????????????????
    """
    def get(self, request):
        try:

            db = DB()
            conn = db.get_connection_nodb()
            db_list = []
            sql = """
            show databases
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                db_list.append(dr[0])

            auth_user = request.GET.get("user")
            pwd = request.GET.get("password")

            product_id = request.GET.get("product_id")
            work_id = request.GET.get("work_id")
            # work_type = request.GET.get("work_type")
            request.session.set_expiry(0)
            if product_id in db_list:
                conn = db.get_connection(product_id)
                sql = """
                        select * from person where user_name = '{0}' and  user_password = '{1}'
                  """
                sql_main = sql.format(auth_user, pwd)
                drs = db.execute_sql(conn, sql_main)
                if drs:
                    sql_type = """
                            select work_type from work_station where work_id = '{0}'
                            """
                    sql_type = sql_type.format(work_id)
                    drs_type = db.execute_sql(conn, sql_type)
                    if drs_type:
                        work_type = drs_type[0][0]
                    else:
                        work_type = "??????????????????"

                    sql = """
                           select * from person where user_name = '{0}' and  user_password = '{1}'
                     """
                    sql_main = sql.format(auth_user, pwd)
                    dfs = db.execute_sql(conn, sql_main)

                    if dfs:
                        work_ids = dfs[0][4]
                        work_ids_list = str(work_ids).split(",")
                        if work_id in work_ids_list:
                            current_person_id = drs[0][0]
                            request.session['session_projectId'] = product_id
                            request.session['session_currentId'] = current_person_id
                            request.session['session_workId'] = work_id
                            request.session['session_workType'] = work_type
                            data = {"code": 0, "message": "????????????"}
                            data = json.dumps(data)
                            logger.info("????????????")
                            db.close_connection(conn)
                            return HttpResponse(data)
                        else:
                            data = {"code": 1, "message": "????????????????????????????????????"}
                            data = json.dumps(data)
                            logging.error("????????????????????????????????????")
                            return HttpResponse(data)
                else:
                    data = {"code": 1, "message": "????????????????????????????????????????????????????????????"}
                    data = json.dumps(data)
                    return HttpResponse(data)



            else:
                data = {"code": 1, "message": "????????????"}
                data = json.dumps(data)
                return HttpResponse(data)

        except Exception as e:

            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("??????????????????----%s" % e)
            print(e)
            return HttpResponse(data)


class ManageLogisticLogin(View):
    # def pad(self, text):
    #     """
    #     # ?????????????????????text??????8????????????????????????text?????????8?????????????????????????????????8?????????
    #     :param text:
    #     :return:
    #     """
    #     while len(text) % 8 != 0:
    #         text += ' '
    #     return text

    """
    ????????????????????? AES ??????????????????
    """
    def get(self, request):
        try:
            db = DB()
            auth_user = request.GET.get("user")
            pwd = request.GET.get("password")

            product_id = request.GET.get("product_id")
            request.session.set_expiry(0)
            if product_id:
                request.session.set_expiry(0)
                request.session['session_projectId'] = product_id
                conn = db.get_connection(product_id)
                sql = """
                select * from person where user_name = '{0}' and  user_password = '{1}'
                """
                sql = sql.format(auth_user, pwd)
                drs = db.execute_sql(db.get_connection("db_common"), sql)
                if drs:
                    for dr in drs:
                        current_person_id = dr[0]
                        request.session['session_projectId'] = product_id
                        request.session['session_currentId'] = current_person_id

                        data = {"code": 0, "message": "????????????"}
                        data = json.dumps(data)
                        logger.info("????????????")
                        db.close_connection(conn)
                        db.close_connection(db.get_connection("db_common"))
                        return HttpResponse(data)
                else:
                    data = {"code": 1, "message": "????????????????????????????????????????????????????????????"}
                    data = json.dumps(data)
                    logger.error("????????????????????????????????????????????????????????????")
                    return HttpResponse(data)
            else:
                db = DB()
                product_id = "db_common"
                conn = db.get_connection("db_common")
                sql = """
                select * from person where user_name = '{0}' and user_password = '{1}'
                """
                sql_main = sql.format(auth_user, pwd)
                drs = db.execute_sql(conn, sql_main)
                request.session.set_expiry(0)
                if drs:
                    for dr in drs:
                        if dr[7] == "???????????????":
                            current_person_id = dr[0]
                            request.session['session_projectId'] = "db_common"
                            request.session['session_currentId'] = current_person_id

                            data = {"code": 0, "message": "????????????"}
                            data = json.dumps(data)
                            logger.info("????????????")
                            db.close_connection(conn)
                            db.close_connection(db.get_connection("db_common"))
                            return HttpResponse(data)
                        else:
                            data = {"code": 1, "message": "??????????????????????????????"}
                            data = json.dumps(data)
                            return HttpResponse(data)
                else:
                    data = {"code": 1, "message": "??????????????????????????????"}
                    data = json.dumps(data)
                    logger.error("??????????????????????????????")
                    return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("?????????????????????????????????%s"% e)

            return HttpResponse(data)


class ModifyPassword(View):
    """
    ??????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id= request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            current_person_id = request.session['session_currentId']
            old_password = request.GET.get("old_password")
            new_password = request.GET.get("new_password")
            sql = """
            select user_password from person where user_code = '{0}'
            """
            sql = sql.format(current_person_id)
            drs = db.execute_sql(conn, sql)

            if old_password == drs[0][0]:
                sql_update = """
                     update person set 
                     user_password='{0}'
                     where user_code = '{1}'    
                """
                sql_update = sql_update.format(new_password, current_person_id)
                db.execute_sql(conn, sql_update)
                data = {"code": 0, "message": "????????????"}
                data = json.dumps(data)
                logger.info("??????????????????")
                db.close_connection(conn)

                return HttpResponse(data)
            else:
                data = {"code": 1, "message": "????????????????????????"}
                data = json.dumps(data)
                logger.error("????????????????????????")
                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????>>>>%s"% e)
            return HttpResponse(data)


class ShowDatabase(View):
    def get(self, request):
        try:
            data_list = []
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
            show tables
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_list.append(dr[0])
            data = {"code": 0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            logger.info("?????????????????????", data)
            return HttpResponse(data)

        except Exception as e:
            data = {"code": 1, "message": "????????????", "data": ""}
            data = json.dumps(data)
            print(e)
            logger.error("?????????????????????"%e)
            return HttpResponse(data)


class NewDatabase(View):
    """
    ?????????????????? ????????????????????????
    """
    def get(self, request):
        try:
            database_list = []
            db_base = "db_common"
            # db_base = request.GET.get("")
            db = DB()
            conn = db.get_connection_mysql()
            sql = """
            show databases
            """
            drs = db.execute_sql(conn, sql)

            for dr in drs:
                database_list.append(dr[0])
            if db_base in database_list:
                data = {"code": 1, "message": "?????????????????????"}
                logger.info("?????????????????????")
            else:
                sql = """
                CREATE DATABASE {0} DEFAULT CHARACTER SET utf8
                """
                sql_main = sql.format(db_base)
                drs = db.execute_sql(conn, sql_main)
                data = {"code": 0, "message": "?????????????????????"}
                logger.info("?????????????????????")

            # data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "?????????????????????"}
            logger.error("?????????????????????")
            data = json.dumps(data)
            return HttpResponse(data)


class PersonDeal(View):

    def get(self, request):
        """
        ???????????????????????? ???????????????????????????
        ????????????????????? ?????????db_common????????????
        """
        try:
            product_id = request.session.get("session_projectId")
            user_id = request.GET.get("user_id")
            user_name = request.GET.get("user_name")
            status = request.GET.get("status")
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            if user_id:
                user_id_sql = "and user_id = " + "'" + user_id + "'"
            else:
                user_id_sql = ""
            if user_name:
                user_name_sql = "and user_name = " + "'" + user_name + "'"
            else:
                user_name_sql = ""
            if status:
                status_sql = "and status = " + "'" + status + "'"
            else:
                status_sql = ""

            data_add_int = {}
            data = []
            data_list = []
            db = DB()
            conn = db.get_connection(product_id)
            sql = """
            select * from person where user_name != 'admin' {0} {1} {2}
            """
            sql = sql.format(user_id_sql, user_name_sql, status_sql)
            drs = db.execute_sql(conn, sql)
            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["user_code"] = dr[0]
                    dict_data["user_id"] = dr[1]
                    dict_data["user_name"] = dr[2]
                    dict_data["user_password"] = dr[3]
                    dict_data["user_authority"] = dr[4]
                    dict_data["status"] = dr[5]
                    dict_data["user_product_id"] = dr[6]
                    dict_data["user_role"] = dr[7]
                    data_list.append(dict_data)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                dfs = int(len(data_list))
                sql_num_int = dfs
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("????????????????????????:")
                db.close_connection(conn)
                return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("??????????????????????????????????????????:")
                return HttpResponse(result)
        except Exception as e:
            logger.error("??????person?????????????????? %s" % e)
            result = {"code": 1, "message": "????????????"}
            result = json.dumps(result)
            return HttpResponse(result)

    def post(self, request):

        """
        ??????????????????
        :param request:
        :return:
        """

        try:
            # aes = AES()
            product_id = request.session.get("session_projectId")
            json_data = request.body
            str_data = json.loads(json_data)
            user_id = str_data.get("user_id")
            user_name = str_data.get("user_name")
            user_password = str_data.get("user_password")

            user_authority = str_data.get("user_authority")
            status = str_data.get("status")
            user_product_id = str_data.get("user_product_id")
            user_role = str_data.get("user_role")

            if product_id == "db_common":
                # product_id = user_product_id
                db = DB()
                conn = db.get_connection(product_id)
            else:
                # product_id = request.session.get("session_projectId")
                db = DB()
                conn = db.get_connection(product_id)

            sql_check = """
            select * from person where user_name = '{0}' and user_id = '{1}'
            """
            sql_check = sql_check.format(user_name, user_id)

            dffs = db.execute_sql(conn, sql_check)

            if dffs:
                data = {"code": 1, "message": "??????????????????"}
                logger.error("????????????????????? ??????????????????")
                data = json.dumps(data)
                db.close_connection(conn)
                return HttpResponse(data)

            else:
                sql = """
                     select max(order_number) from person
                    """
                id_num = db.execute_sql(conn, sql)

                if id_num[0][0] == None:
                    id_num = 1
                else:
                    id_num = id_num[0][0] + 1
                user_code = str("Im_User_" + str(id_num))

                sql_insert = """
                    insert into person(user_code, user_id, user_name, user_password, user_authority, 
                    status, user_product_id, user_role, order_number)
                    values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}', '{7}', '{8}')
                    """

                sql_insert_format = sql_insert.format(user_code, user_id, user_name,
                                                      user_password, user_authority, status, user_product_id,
                                                      user_role, id_num)

                drs = db.execute_sql(conn, sql_insert_format)
                db.close_connection(conn)
                data = {"code": 0, "message": "??????????????????"}
                data = json.dumps(data)
                logger.info("??????????????????")
                return HttpResponse(data)

        except Exception as e:
            data = {"code": 1, "message": "??????????????????%s" % e}
            logger.error("??????????????????%s"% e)
            data = json.dumps(data)
            return HttpResponse(data)


class PutPerson(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            product_id = request.session.get("session_projectId")
            db = DB()
            conn = db.get_connection(product_id)

            user_code = request.GET.get("user_code")
            user_id = request.GET.get("user_id")
            user_name = request.GET.get("user_name")
            user_password = request.GET.get("user_password")
            user_authority = request.GET.get("user_authority")
            status = request.GET.get("status")
            user_product_id = request.GET.get("user_product_id")
            user_role = request.GET.get("user_role")
            sql = """
                      update person set user_id='{0}',
                      user_name='{1}',user_password='{2}',user_authority='{3}',status='{4}'
                      ,user_product_id = '{5}', user_role = '{6}' where user_code = '{7}'
                      """
            sql_format = sql.format(user_id, user_name, user_password, user_authority,
                                    status, user_product_id, user_role, user_code)
            db.execute_sql(conn, sql_format)

            data = {"code": 0, "message": "??????????????????"}
            data = json.dumps(data)
            logger.info("??????????????????")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            data = {"code": 1, "message": "??????????????????,---%s" % e}
            data = json.dumps(data)
            logger.error("??????????????????"%e)
            return HttpResponse(data)


class DeletePerson(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            product_id = request.session.get("session_projectId")
            db = DB()
            conn = db.get_connection(product_id)
            user_code = request.GET.get("user_code")
            sql = """
            delete FROM person where user_code = '{0}'       
            """
            sql_format = sql.format(user_code)
            db.execute_sql(conn, sql_format)
            data = {"code": 0, "message": "????????????"}
            data = json.dumps(data)
            logger.info("??????????????????")
            db.close_connection(conn)
            return HttpResponse(data)
        except Exception as e:
            data = {"code":1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("??????????????????%s"% e)
            return HttpResponse(data)


class PersonTable(View):
    """
    ?????????????????????  ????????????????????????????????? ?????????
    """
    def get(self, request):
        try:
            table_list = []
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
            create table person(user_code varchar(64) primary key not null,
            user_id varchar(64), user_name varchar(32),
            user_password varchar(64), user_authority varchar(128), status varchar(10), user_product_id varchar(128),
            user_role varchar(128), order_number int);
            """
            dr = db.execute_sql(conn, sql)
            sql_main = """
            show tables
            """
            drs = db.execute_sql(conn, sql_main)
            for pe in drs:
                table_list.append(pe[0])
            if "person" in table_list:
                data = {"code": 0, "message": "????????????"}
            else:
                data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)


class CreatePersonMatter(View):

    def get(self, request):
        try:
            table_list = []
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
            create table person_matter(matter_code varchar(64) primary key not null,
            matter_id varchar(64),
            matter_name varchar(64), matter_category varchar(64), rule varchar(64),
            matter_count int, product_time datetime, status varchar(128), order_number int);
            """
            dr = db.execute_sql(conn, sql)
            sql_main = """
            show tables
            """
            drs = db.execute_sql(conn, sql_main)
            for pe in drs:
                table_list.append(pe[0])
            if "person" in table_list:
                data = {"code": 0, "message": "????????????"}
                logger.info("????????????????????????????????????")
            else:
                data = {"code": 1, "message": "????????????"}
                logger.info("????????????????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            data = {"code": 0, "message": "????????????"}
            data = json.dumps(data)
            print(e)
            logger.error("----?????????????????????????????????%s"% e)
            return HttpResponse(data)


class PersonMatter(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            product_id = request.session.get("session_projectId")
            data_list = []
            data_add_int = {}
            matter_id = request.GET.get("matter_id")
            matter_name = request.GET.get("matter_name")
            matter_category = request.GET.get("matter_category")
            operate_user = request.GET.get("operate_user")

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            db = DB()
            conn = db.get_connection(product_id)

            if matter_id:
                matter_id_sql = " and matter_id =" + "'" + matter_id + "'"
            else:
                matter_id_sql = ""
            if matter_name:
                matter_name_sql = " and matter_name = " + "'" + matter_name + "'"
            else:
                matter_name_sql = ""
            if matter_category:
                matter_category_sql = " and matter_category = " + "'" + matter_category+"'"
            else:
                matter_category_sql = ""
            if operate_user:
                operate_user_sql = " and operate_user = " + "'" + operate_user+"'"
            else:
                operate_user_sql = ""

            sql = """
             select pm.matter_code, ml.matter_name, ml.rule, ml.matter_category, pm.matter_count,
              pm.product_time, pm.operate_user,ml.code_length from person_matter as pm left join matter_list as ml on pm.matter_code = ml.bom_matter_code
              where 2 > 1 {0} {1} {2} {3}
            """

            sql_main = sql.format(matter_id_sql, matter_name_sql, matter_category_sql, operate_user_sql)
            drs = db.execute_sql(conn, sql_main)
            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["matter_code"] = dr[0]
                    dict_data["matter_name"] = dr[1]
                    dict_data["rule"] = dr[2]
                    dict_data["matter_category"] = dr[3]
                    dict_data["matter_count"] = dr[4]
                    s = dr[5]
                    if s:
                        s1 = s.strftime("%Y-%m-%d %H:%M:%S ")
                    else:
                        s1 = ""
                    dict_data["product_time"] = s1
                    dict_data["operate_user"] = dr[6]
                    dict_data["code_length"] = dr[7]
                    data_list.append(dict_data)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                sql_num_int = int(len(data_list))
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int

                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("??????????????????")
                db.close_connection(conn)
                return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                result = json.dumps(result)
                db.close_connection(conn)
                logger.info("????????????????????????????????????")
                return HttpResponse(result)
        except Exception as e:
            logger.error("??????personmatter??????????????????%s"% e)
            result = {"code": 1, "message": "????????????", "data": ""}
            result = json.dumps(result)

            print(e)
            return HttpResponse(result)

    def post(self, request):
        """
        ???????????????
        :param request:
        :return:
        """

        try:
            db = DB()

            product_id = request.session.get("session_projectId")
            operate_user_code = request.session['session_currentId']
            conn = db.get_connection(product_id)
            conn_common = db.get_connection("db_common")

            sql_person = """
            select user_name from person where user_code = '{0}'
            """
            sql_person = sql_person.format(operate_user_code)
            dr_persons = db.execute_sql(conn_common, sql_person)
            if dr_persons:
                operate_user = dr_persons[0][0]
            else:
                operate_user = operate_user_code
            sql = """
             SELECT max(order_number) FROM person_matter;
            """
            matter_type = db.execute_sql(conn, sql)
            if matter_type[0][0] == None:
                id_num = 1
            else:
                id_num = matter_type[0][0] + 1
            json_data = request.body
            str_data = json.loads(json_data)

            matter_code = str_data.get("matter_code")

            matter_count = str_data.get("matter_count")
            product_time = str_data.get("product_time")
            product_plan_code = str_data.get("product_plan_code")

            order_number = id_num

            sql_insert = """
                       insert into person_matter(matter_code, matter_count, product_time, operate_user, 
                        order_number)
                       values('{0}', '{1}', '{2}', '{3}', '{4}')
                       """
            sql_insert_format = sql_insert.format(matter_code, matter_count, product_time,
                                                  operate_user, order_number)
            drs = db.execute_sql(conn, sql_insert_format)

            data = {"code": 0, "message": "???????????????????????????"}
            logger.info("???????????????????????????")

            data = json.dumps(data)

            db.close_connection(conn)
            db.close_connection(conn_common)

            return HttpResponse(data)

        except Exception as e:

            data = {"code": 1, "message": "???????????????????????????%s" % e}
            logger.error("???????????????????????????%s"% e)
            data = json.dumps(data)
            print(e)

            return HttpResponse(data)


class PutPersonMatter(View):
    """
    ??????????????????
    """
    def get(self, request):
        try:
            product_id = request.session.get("session_projectId")

            db = DB()
            conn = db.get_connection(product_id)
            matter_code = request.GET.get("matter_code")
            matter_count = request.GET.get("matter_count")
            product_time = request.GET.get("product_time")
            operate_user = request.GET.get("operate_user")
            product_plan_code = request.GET.get("product_plan_code")

            sql = """
             update person_matter set 
             matter_count ='{0}', product_time='{1}', operate_user = '{2}'
             where matter_code = '{3}'          
            """
            sql_main = sql.format(matter_count, product_time, operate_user, matter_code)
            drs = db.execute_sql(conn, sql_main)
            db.close_connection(conn)

            data = {"code": 0, "message": "????????????"}
            data = json.dumps(data)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message":"????????????"}
            logger.error("????????????---->%s",e)
            data = json.dumps(data)
            return HttpResponse(data)


class DeletePersonMatter(View):
    """
    ??????????????????

    """

    def get(self, request):
        try:
            product_id = request.session.get("session_projectId")
            db = DB()
            conn = db.get_connection(product_id)
            matter_code = request.GET.get("matter_code")
            sql = """
            delete from person_matter where matter_code = '{0}'
            """
            sql_main = sql.format(matter_code)
            des = db.execute_sql(conn, sql_main)

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message":"????????????"}
            data = json.dumps(data)
            return HttpResponse(data)


class CreateBOMProductList(View):
    """
    ???????????????????????????
    """

    def get(self, request):
        try:
            table_list = []
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
            create table ProductList(prodect_code varchar(128) primary key not null,
            product_name varchar(128), product_id varchar(128), rule varchar(128),
            product_status varchar(128), description varchar(128), order_number int);          
            """
            dr = db.execute_sql(conn, sql)
            sql_main = """
                        show tables                 
                      """
            drs = db.execute_sql(conn, sql_main)
            for pe in drs:
                table_list.append(pe[0])

            if "productlist" in table_list:
                data = {"code": 0, "message": "????????????"}
                logger.info("?????????????????????????????????")
            else:
                data = {"code": 1, "message": "????????????"}
                logger.info("???????????????????????????")
            data = json.dumps(data)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 0, "message": "????????????-%s"%e}
            data = json.dumps(data)
            logger.error("????????????????????????%s"% e)

            return HttpResponse(data)


class CreateBOMMatterList(View):
    """
    ?????????????????????
    """
    def get(self, request):
        try:
            table_list = []
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
                       create table matter_list(bom_matter_code varchar(64) primary key not null,
                    prodect_code varchar(64) not null,
                       matter_name varchar(64), rule varchar(64),matter_category varchar(64),
                       matter_usage int, order_number int)         
                       """
            dr = db.execute_sql(conn, sql)
            sql_main = """
                        show tables                 
                        """
            drs = db.execute_sql(conn, sql_main)
            for pe in drs:
                table_list.append(pe[0])

            if "matter_list" in table_list:
                data = {"code": 0, "message": "????????????"}
                logger.info("???????????????????????????????????????")
            else:
                data = {"code": 1, "message": "????????????"}
                logger.info("??????????????????????????????")
            data = json.dumps(data)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 0, "message": "????????????-%s" % e}
            data = json.dumps(data)
            logger.error("?????????????????????????????????????????????%s"% e)

            return HttpResponse(data)


class BOMProductList(View):
    """
    GET  ??????????????????
    post  ?????????????????? ?????? ?????????  ??????????????????
    ??????????????????db_common??????????????? ????????????????????????????????????????????????????????????matter_list???
    create ????????????
    """

    def get(self, request):
        try:
            data_list = []
            matter_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            product_name = request.GET.get("product_name")
            rule = request.GET.get("rule")
            product_status = request.GET.get("product_status")

            if product_id == "db_common":

                if product_id == "db_common":
                    s = ""
                else:
                    s = "and product_id =" + "'" + product_id + "'"

                if product_name:
                    product_name_sql = "and product_name = " + "'" + product_name + "'"
                else:
                    product_name_sql = ""
                if rule:
                    rule_sql = "and rule = " + "'" + rule + "'"
                else:
                    rule_sql = ""
                if product_status:
                    product_status_sql = "and product_status = " + "'" + product_status + "'"
                else:
                    product_status_sql = ""
                sql = """
                 select * from productlist where 2 > 1 {0} {1} {2} {3}
                """
                sql_main = sql.format(s, product_name_sql, rule_sql, product_status_sql)
                drs = db.execute_sql(db.get_connection("db_common"), sql_main)

                if drs:
                    for dr in drs:
                        dict_data = {}
                        dict_data["prodect_code"] = dr[0]
                        dict_data["product_name"] = dr[1]
                        dict_data["product_id"] = dr[2]
                        dict_data["rule"] = dr[3]
                        dict_data["product_status"] = dr[4]
                        dict_data["description"] = dr[5]
                        data_list.append(dict_data)
                    page_result = Page(page, page_size, data_list)
                    data = page_result.get_str_json()
                    sql_num_int = int(len(data_list))

                    data_add_int["data"] = data
                    data_add_int["total"] = sql_num_int

                    result = {"code": 0, "message": "????????????", "data": data_add_int}
                    result = json.dumps(result)

                    return HttpResponse(result)

                else:
                    data_add_int["data"] = []
                    data_add_int["total"] = 0
                    result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                    result = json.dumps(result)

                    return HttpResponse(result)

            if product_id != "db_common":
                if product_name:
                    product_name_sql = "and product_name = " + "'" + product_name + "'"
                else:
                    product_name_sql = ""
                if rule:
                    rule_sql = "and rule = " + "'" + rule + "'"
                else:
                    rule_sql = ""
                if product_status:
                    product_status_sql = "and product_status = " + "'" + product_status + "'"
                else:
                    product_status_sql = ""
                sql = """
                 select * from productlist where product_id = '{0}' {1} {2} {3} 
                """
                sql_main = sql.format(product_id, product_name_sql, rule_sql, product_status_sql)
                drs = db.execute_sql(db.get_connection("db_common"), sql_main)
                if drs:
                    for dr in drs:
                        dict_data = {}
                        dict_data["prodect_code"] = dr[0]
                        dict_data["product_name"] = dr[1]
                        dict_data["product_id"] = dr[2]
                        dict_data["rule"] = dr[3]
                        dict_data["product_status"] = dr[4]
                        dict_data["description"] = dr[5]
                        sql_matter = """
                        select * from matter_list where prodect_code = '{0}'
                        """
                        sql_matter_format = sql_matter.format(dr[0])
                        matter_dfs = db.execute_sql(conn, sql_matter_format)
                        if matter_dfs:
                            respose_list = []
                            for matter_df in matter_dfs:
                                response_data = {}
                                response_data["bom_matter_code"] = matter_df[0]
                                response_data["prodect_code"] = matter_df[1]
                                response_data["matter_name"] = matter_df[2]
                                response_data["rule"] = matter_df[3]
                                response_data["matter_category"] = matter_df[4]
                                response_data["matter_usage"] = matter_df[5]
                                response_data["code_length"] = matter_df[6]
                                respose_list.append(response_data)
                            dict_data["response_datas"] = respose_list
                            data_list.append(dict_data)
                        else:
                            dict_data["response_datas"] = []
                            data_list.append(dict_data)
                            # data_list.append(response_data)

                    page_result = Page(page, page_size, data_list)
                    data = page_result.get_str_json()
                    if data_list:
                        sql_num_int = int(len(data_list))
                    else:
                        sql_num_int = 0

                    data_add_int["data"] = data
                    data_add_int["total"] = sql_num_int

                    result = {"code": 0, "message": "????????????", "data": data_add_int}
                    result = json.dumps(result)
                    db.close_connection(conn)

                    return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                result = json.dumps(result)
                db.close_connection(conn)

                return HttpResponse(result)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "??????BOM??????????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????%s"% e)

            return HttpResponse(data)

    def post(self, request):
        """
        ????????????  ???????????????????????????????????????????????? ?????????????????????????????????????????????
        ?????????????????????????????????db_common????????????product_list??????,
        ??????????????????????????????????????????????????????????????????????????????????????????matter_list???

        :param request:
        :return:
        """

        try:
            product_id = request.session.get("session_projectId")
            json_data = request.body
            str_data = json.loads(json_data)
            product_db = str_data.get("product_id")
            db = DB()
            db.create_database(product_db)
            conn = db.get_connection(product_db)
            sql_person = """
                        CREATE TABLE IF NOT EXISTS person(user_code varchar(64) primary key not null,
                        user_id varchar(64), user_name varchar(32),
                        user_password varchar(64), user_authority varchar(128), status varchar(10), 
                        user_product_id varchar(128),
                        user_role varchar(128),order_number int);
                        """
            dr = db.execute_sql(conn, sql_person)

            sql_matter = """           
                       CREATE TABLE IF NOT EXISTS matter_list(bom_matter_code varchar(64) primary key not null,
                       prodect_code varchar(64) not null,
                       matter_name varchar(64), rule varchar(64),matter_category varchar(64),
                       matter_usage int, code_length int, order_number int)                                    
                     """
            dr = db.execute_sql(conn, sql_matter)

            sql_work = """
                            CREATE TABLE IF NOT EXISTS work_station(work_code varchar(128) primary key not null,
                            work_id varchar(128) not null,leader_work_id varchar(128),
                            work_name varchar(128), work_type varchar(128), order_number int)         
                            """
            dr = db.execute_sql(conn, sql_work)

            sql_process = """
                                          CREATE TABLE IF NOT EXISTS process_deal(production_code varchar(128) primary key not null,
                                          production_id varchar(128) not null,
                                          production_name varchar(128), work_id varchar(128),
                                          description varchar(128), order_number int)         
                                          """
            dr = db.execute_sql(conn, sql_process)

            sql_check = """
                            CREATE TABLE IF NOT EXISTS check_productdeal(check_code varchar(128) primary key not null,
                            prodect_code varchar(128),
                            work_code varchar(128) not null,
                            production_code varchar(128), check_method varchar(128), order_number int)         
                            """
            dr = db.execute_sql(conn, sql_check)

            sql_product_plan = """
                            CREATE TABLE IF NOT EXISTS product_plandeal(
                            product_plan_code varchar(128) primary key not null,
                            plan_name varchar(128), plan_count int, plan_start_day datetime, plan_end_day datetime, 
                            description varchar(128), plan_status varchar(128), order_number int)         
                            """
            dr = db.execute_sql(conn, sql_product_plan)

            sql_pick_matter = """
                                   CREATE TABLE IF NOT EXISTS product_pickmatter(
                                   materials_production_code varchar(128) primary key not null,
                                   materials_person varchar(128) not null,
                                   product_plan_code varchar(128), material_time datetime, description varchar(128),
                                    order_number int)         
                                   """
            dr = db.execute_sql(conn, sql_pick_matter)

            sql_transit = """
                            CREATE TABLE IF NOT EXISTS producttransitinfo(product_transit_code varchar(128) primary key not null,
                                                 matter_code varchar(128), user_code varchar(128),
                                                 work_code varchar(128),
                                                 test_result varchar(128), 
                                                 description varchar(128), 
                                                 enter_time datetime,
                                                 out_time datetime,
                                                 product_plan_code varchar(128),
                                                 end_product_code varchar(128),
                                                 product_code varchar(128),                                     
                                                 order_number int)         
                                                 """
            dr = db.execute_sql(conn, sql_transit)

            sql_martial = """
                       CREATE TABLE IF NOT EXISTS person_matter(matter_code varchar(128) primary key not null,
                       matter_count int, product_time datetime, operate_user varchar(128), 
                       order_number int);
                       """
            dr = db.execute_sql(conn, sql_martial)

            sql_product_parameter = """
                                                 CREATE TABLE IF NOT EXISTS 
                                                 product_parameter(test_code varchar(128) primary key not null,
                                                 check_code varchar(128), test_parameter varchar(128),
                                                 test_parameter_count varchar(128),
                                                 test_status varchar(128), order_number int)         
                                                 """
            dr = db.execute_sql(conn, sql_product_parameter)

            sql_matters = """
                        CREATE TABLE IF NOT EXISTS pick_matter(materials_code varchar(128) primary key not null, 
                        materials_production_code varchar(128) not null,
                        matter_code varchar(128), matter_count int, order_number int)
                        """
            df = db.execute_sql(conn, sql_matters)

            sql_pick = """
                        CREATE TABLE IF NOT EXISTS pick_box(pick_code varchar(128) primary key not null,
                        work_id varchar(128), 
                        pick_number int,
                        description varchar(128), order_number int)
                        """
            df = db.execute_sql(conn, sql_pick)

            sql_back = """
                        CREATE TABLE IF NOT EXISTS product_back_matter(materials_back_code varchar(128) primary key not null,
                        back_person varchar(128), 
                        product_plan_code varchar(128),
                        back_time datetime,
                        description varchar(128), order_number int)
                      """
            df = db.execute_sql(conn, sql_back)

            sql_deal_back = """
                        CREATE TABLE IF NOT EXISTS back_matter(deal_back_code varchar(128) primary key not null,
                        materials_back_code varchar(128), 
                        matter_code varchar(128),
                        matter_count int,
                        order_number int)
                      """
            df = db.execute_sql(conn, sql_deal_back)

            sql_enter_status = """           
            CREATE TABLE IF NOT EXISTS enter_storage_status(pack_id varchar(128) primary key not null,
                              product_id varchar(128),product_name varchar(128),
                              enter_user varchar(128), enter_time datetime, 
                              status varchar(128),product_plan_code varchar(128),order_number int)
            """
            db.execute_sql(conn, sql_enter_status)

            sql_enter = """
            CREATE TABLE IF NOT EXISTS enter_storage(enter_storage_code varchar(128) primary key not null, 
                              pack_id varchar(128),
                              finished_product_code varchar(128),
                              order_number int)
            """
            db.execute_sql(conn, sql_enter)

            sql_operate = """
            CREATE TABLE IF NOT EXISTS operate(ID int primary key not null, 
                     Package_Qty int,
                     Rv varchar(128),
                     Itemcode_C_Shipping varchar(128),
                     Supplier varchar(128),
                     No_Ship int,
                     CS_type varchar(128),
                     Shipping_SN_length int,
                     Product_length int,
                     order_number int);
            """
            db.execute_sql(conn, sql_operate)

            sql_function = """
                        CREATE TABLE IF NOT EXISTS function_list(function_list_code int AUTO_INCREMENT primary key not null, 
                                 function_name varchar(128));
                        """
            db.execute_sql(conn, sql_function)

            sql_work_tran = """
                        CREATE TABLE IF NOT EXISTS work_transit(table_code varchar(128) primary key not null,
                                              table_name varchar(128), work_code varchar(128),  
                                              description varchar(128),
                                              create_time datetime,
                                              order_number int)
            """
            db.execute_sql(conn, sql_work_tran)

            sql_file_list = """           
               CREATE TABLE IF NOT EXISTS field_list(field_code varchar(128) primary key not null, 
                                              table_name varchar(128),
                                              field_name varchar(128),  
                                              field_type varchar(128),
                                              field_PK varchar(128),
                                              field_NN varchar(128),
                                              field_AI varchar(128),order_number int)
            """
            db.execute_sql(conn, sql_file_list)

            sql_ = """
            CREATE TABLE IF NOT EXISTS process_matter_deal(process_matter_deal_code varchar(128) primary key not null,
                              work_id varchar(128) not null,
                              matter_code varchar(128),
                               install_number int, order_number int)
            """
            db.execute_sql(conn, sql_)

            sql_Work_Unqualified_Result = """
            CREATE TABLE IF NOT EXISTS work_unqualified_result(unqualified_result_code varchar(128) primary key not null,
                        work_id varchar(128), unqualified_result_name varchar(128),
                        content varchar(64), process_mode varchar(128), operate_user varchar(128), 
                        time DATETIME DEFAULT CURRENT_TIMESTAMP not null,order_number int);
            """
            db.execute_sql(conn, sql_Work_Unqualified_Result)

            sql_modify_station = """
            CREATE TABLE IF NOT EXISTS modify_in_workstation(modify_in_workstation_code varchar(128) primary key not null,
                        product_name varchar(128), finished_product_code varchar(128), work_id varchar(128), 
                        product_plan_code varchar(128),
                        operate_user varchar(128), 
                        enter_time DATETIME DEFAULT CURRENT_TIMESTAMP not null,process_method varchar(128),
                        out_time datetime, 
                        status varchar(128),
                        enter_work_code varchar(128),
                        order_number int);           
            """

            db.execute_sql(conn, sql_modify_station)

            sql_product_out = """
            CREATE TABLE IF NOT EXISTS finished_product_out(product_out_code varchar(128) primary key not null,
                       pack_id varchar(128), flex_part_number varchar(128) not null,
                       vendor_code varchar(128), part_qty varchar(128),package_sequence varchar(128),
                       part_revision varchar(128), manufacturing_date date,operate_user varchar(128), order_number int);
            """
            db.execute_sql(conn, sql_product_out)

            sql_matter = """
            SELECT max(order_number) FROM matter_list;
            """
            matter_type = db.execute_sql(conn, sql_matter)
            if matter_type[0][0] == None:
                matter_num = 0
            else:
                matter_num = matter_type[0][0]

            conn_product = db.get_connection("db_common")

            sql_common = """
                        SELECT max(order_number) FROM matter_list;
                        """
            common_matter = db.execute_sql(conn_product, sql_common)
            if common_matter[0][0] == None:
                common_num = 0
            else:
                common_num = common_matter[0][0]

            sql = """
                         SELECT max(order_number) FROM productlist;
                """
            product_type = db.execute_sql(conn_product, sql)
            if product_type[0][0] == None:
                id_num = 1
            else:
                id_num = product_type[0][0] + 1
            prodect_code = str("Im_Product_" + str(id_num))
            product_name = str_data.get("product_name")
            rule = str_data.get("rule")
            product_status = str_data.get("product_status")
            description = str_data.get("description")
            response_datas = str_data.get("response_datas")

            order_number = id_num
            sql_insert = """
                        insert into productlist(prodect_code, product_name, product_id, rule,
                        product_status, description, order_number)
                        values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}')
                        """
            sql_insert_format = sql_insert.format(prodect_code, product_name, product_db, rule, product_status,
                                                  description, order_number)

            drs = db.execute_sql(conn_product, sql_insert_format)

            if product_id == "db_common":
                data = {"code": 0, "message": "???????????????????????????"}
                logger.info("???????????????????????????")

                data = json.dumps(data)

                return HttpResponse(data)
            else:
                for response_data in response_datas:
                    # db = DB()
                    # conn = db.get_connection("db_common")
                    matter_name = response_data.get("matter_name")
                    rule = response_data.get("rule")
                    matter_category = response_data.get("matter_category")
                    matter_usage = response_data.get("matter_usage")
                    code_length = response_data.get("code_length")
                    matter_num = matter_num+1
                    common_num = common_num +1
                    bom_matter_code = str("Im_BOM_Matter_" + str(matter_num))
                    bom_matter_code2 = str("Im_BOM_Matter_" + str(common_num))
                    sql_response = """
                            insert into matter_list(bom_matter_code,prodect_code, matter_name, rule,
                            matter_category, matter_usage, code_length, order_number)
                            values('{0}', '{1}', '{2}', '{3}', '{4}','{5}', '{6}', '{7}')
                    """
                    sql_response_format = sql_response.format(bom_matter_code, prodect_code, matter_name,
                                                              rule, matter_category, matter_usage,
                                                              code_length, matter_num)

                    sql_common_format = sql_response.format(bom_matter_code2, prodect_code, matter_name,
                                                              rule, matter_category, matter_usage, code_length,
                                                            common_num)
                    matter_drs = db.execute_sql(conn, sql_response_format)
                    matter_dfs = db.execute_sql(conn_product, sql_common_format)

                data = {"code": 0, "message": "???????????????????????????"}
                logger.info("???????????????????????????")

                data = json.dumps(data)
                db.close_connection(conn)

                return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "???????????????????????????"}
            logger.error("???????????????????????????%s" %e)
            data = json.dumps(data)
            return HttpResponse(data)


class PutBOMProductList(View):

    """
    ??????????????????

    """
    def get(self, request):
        try:

            db = DB()
            conn = db.get_connection("db_common")

            prodect_code = request.GET.get("prodect_code")
            product_name = request.GET.get("product_name")
            rule = request.GET.get("rule")
            product_status = request.GET.get("product_status")
            description = request.GET.get("description")
            response_datas = request.GET.get("response_datas")
            product_id = request.session['session_projectId']

            conn_project = db.get_connection(product_id)
            if response_datas:
                response_datas = eval(response_datas)
            else:
                response_datas = response_datas
            sql = """
                update productlist set 
                   product_name='{0}',rule='{1}',product_status='{2}', description ='{3}'
                   where prodect_code = '{4}'
                   """
            sql_main = sql.format(product_name, rule, product_status, description, prodect_code)
            drs = db.execute_sql(conn, sql_main)

            if product_id != "db_common":
                for response_data in response_datas:
                    bom_matter_code = response_data.get("bom_matter_code")
                    if bom_matter_code:
                        sql_delete = """
                         delete FROM matter_list where bom_matter_code = '{0}'
                        """
                        sql_delete_format = sql_delete.format(bom_matter_code)
                        db.execute_sql(conn_project, sql_delete_format)

                sql_after = """
                            SELECT max(order_number) FROM matter_list
                            """
                num = db.execute_sql(conn_project, sql_after)
                if num[0][0]:
                    sql_num = num[0][0]
                else:
                    sql_num = 0

                for response_data in response_datas:
                    # db = DB()
                    # conn = db.get_connection("db_common")
                    matter_name = response_data.get("matter_name")
                    rule = response_data.get("rule")
                    matter_category = response_data.get("matter_category")
                    matter_usage = response_data.get("matter_usage")
                    code_length = response_data.get("code_length")
                    sql_num = sql_num + 1
                    bom_matter_code = str("Im_BOM_Matter_" + str(sql_num))
                    sql_response = """
                                        insert into matter_list(bom_matter_code,prodect_code, matter_name, rule,
                                        matter_category, matter_usage, code_length, order_number)
                                        values('{0}', '{1}', '{2}', '{3}', '{4}','{5}', '{6}', '{7}')
                                """
                    sql_response_format = sql_response.format(bom_matter_code, prodect_code, matter_name,
                                                              rule, matter_category, matter_usage, code_length, sql_num)
                    matter_drs = db.execute_sql(conn_project, sql_response_format)
                data = {"code": 0, "message": "????????????"}

                data = json.dumps(data)

                db.close_connection(conn)

                return HttpResponse(data)

            else:
                data = {"code": 0, "message": "????????????"}

                data = json.dumps(data)

                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            logger.error("????????????---->%s"% e)
            data = json.dumps(data)
            return HttpResponse(data)


class DeleteBOMProductList(View):
    """
    ???????????????????????????
    """
    def get(self, request):
        try:
            db = DB()
            db_base = request.GET.get("product_id")
            product_id = request.session['session_projectId']
            conn = db.get_connection("db_common")
            conn_project = db.get_connection(product_id)
            prodect_code = request.GET.get("prodect_code")

            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)

            if response_datas:
                for response_data in response_datas:
                    bom_matter_code = dict(response_data).get("bom_matter_code")
                    sql = """
                    delete from matter_list where bom_matter_code = '{0}'
                    """
                    sql_format = sql.format(bom_matter_code)
                    db.execute_sql(conn_project, sql_format)
            else:
                sql_database = """
                select product_id from productlist where prodect_code = '{0}'
                """
                sql_database = sql_database.format(prodect_code)
                drs_ids = db.execute_sql(conn, sql_database)
                drs_id = drs_ids[0][0]
                con_db = db.get_connection_nodb()
                sql_drop_db = """
                drop database {0}
                """
                sql_drop_db = sql_drop_db.format(drs_id)
                db.execute_sql(con_db, sql_drop_db)
                sql = """
                    delete from productlist where prodect_code = '{0}'
                    """
                sql_main = sql.format(prodect_code)
                des = db.execute_sql(conn, sql_main)

            data = {"code": 0, "message": "????????????"}
            logger.info("???????????????????????????????????????")

            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????,%s"% e)
            return HttpResponse(data)


class BOMMatterList(View):

    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            conn = db.get_connection("db_common")
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            matter_name = request.Get.get("matter_name")
            rule = request.Get.get("rule")
            if len(rule) == 0 and len(matter_name) == 0:
                sql = """
                    select * from matter_list
                    """
                drs = db.execute_sql(conn, sql)
                if drs:
                    for dr in drs:
                        dict_data = {}
                        dict_data["prodect_code"] = dr[0]
                        dict_data["matter_name"] = dr[1]
                        dict_data["rule"] = dr[2]
                        dict_data["matter_usage"] = dr[3]
                        dict_data["code_length"] = dr[4]
                        dict_data["order_number"] = dr[5]
                        data_list.append(dict_data)

                    page_result = Page(page, page_size, data_list)
                    data = page_result.get_str_json()
                    sql_num = """
                                    select count(*) from matter_list
                                    """
                    dfs = db.execute_sql(conn, sql_num)
                    sql_num_int = dfs[0][0]

                    data_add_int["data"] = data
                    data_add_int["total"] = sql_num_int

            else:
                sql = """
                select * from matter_list where matter_name = '{0}' and rule = '{1}'               
                """
                sql_main = sql.format(matter_name, rule)
                drs = db.execute_sql(conn, sql_main)
                if drs:
                    for dr in drs:
                        dict_data = {}
                        dict_data["prodect_code"] = dr[0]
                        dict_data["matter_name"] = dr[1]
                        dict_data["rule"] = dr[2]
                        dict_data["matter_usage"] = dr[3]
                        dict_data["code_length"] = dr[4]
                        dict_data["order_number"] = dr[5]
                        data_list.append(dict_data)

                    page_result = Page(page, page_size, data_list)
                    data = page_result.get_str_json()
                    sql_num = """
                        select count(*) from matter_list 
                        where matter_name = '{0}' and rule = '{1}'
                          """
                    sql_format = sql_num.format(matter_name, rule)
                    dfs = db.execute_sql(conn, sql_format)
                    sql_num_int = dfs[0][0]

                    data_add_int["data"] = data
                    data_add_int["total"] = sql_num_int

            data = {"code": 0, "message": "????????????", "data": data_add_int}
            data = json.dumps(data)
            logger.info("??????BOM??????????????????")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "??????BOM??????????????????"}
            data = json.dumps(data)
            logger.error("??????BOM??????????????????")

            return HttpResponse(data)

    def post(self, request):
        try:
            db = DB()
            conn = db.get_connection("db_common")

            sql = """
                         SELECT max(order_number) FROM matter_list;
                        """
            matter_type = db.execute_sql(conn, sql)
            if matter_type[0][0] == None:
                id_num = 1
            else:
                id_num = matter_type[0][0] + 1

            prodect_code = request.POST.get("prodect_code")
            matter_name = request.POST.get("matter_name")
            rule = request.POST.get("rule")
            matter_usage = request.POST.get("matter_usage")
            code_length = request.POST.get("code_length")
            order_number = id_num

            # prodect_code = "Im_Product_3"
            # matter_name = "??????"
            # rule = "??????"
            # matter_usage = 26
            # order_number = id_num

            # product_name = "??????"
            # rule = "10,23,22"
            # product_status = "??????"
            # description = "????????????"
            # order_number = id_num
            sql_insert = """
                        insert into matter_list(prodect_code, matter_name, rule,
                        matter_usage, code_length, order_number)
                        values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                        """
            sql_insert_format = sql_insert.format(prodect_code, matter_name, rule, matter_usage, code_length,
                                                  order_number)

            drs = db.execute_sql(conn, sql_insert_format)
            sql_after = """
                        SELECT max(order_number) FROM matter_list
                        """
            num = db.execute_sql(conn, sql_after)
            if num[0][0] == order_number:
                data = {"code": 0, "message": "???????????????????????????"}
                logger.info("???????????????????????????")
            else:
                data = {"code": 1, "message": "???????????????????????????"}
                logger.error("???????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "???????????????????????????"}
            logger.error("???????????????????????????")
            data = json.dumps(data)


class CreateWorkStation(View):
    def get(self, request):
        try:
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
                create table Work_Station(work_code varchar(64) primary key not null,
                work_id varchar(64) not null,
                work_name varchar(64), work_type varchar(128), leader_work_id varchar(128),order_number int)         
                """
            dr = db.execute_sql(conn, sql)

            data = {"code": 0, "message": "????????????"}
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

        data = json.dumps(data)

        return HttpResponse(data)


class WorkStation(View):
    """
    ????????????????????????--- ?????????????????????????????????
    """

    def get(self, request):
        try:
            data_list = []
            matter_list = []
            data_add_int = {}
            product_id = request.session.get("session_projectId")
            db = DB()
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            work_id = request.GET.get("work_id")
            work_name = request.GET.get("work_name")

            if work_id:
                work_id_sql = " and work_id =" + "'" + work_id + "'"
            else:
                work_id_sql = ""
            if work_name:
                work_name_sql = " and work_name =" + "'" + work_name + "'"
            else:
                work_name_sql = ""

            sql = """
                select * from  work_station where 2>1 {0} {1}
                """
            sql_main = sql.format(work_id_sql, work_name_sql)
            drs = db.execute_sql(conn, sql_main)

            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["work_code"] = dr[0]
                    dict_data["work_id"] = dr[1]
                    dict_data["work_name"] = dr[3]
                    dict_data["work_type"] = dr[4]
                    dict_data["leader_work_id"] = dr[2]
                    data_list.append(dict_data)
                data_list.sort(key=lambda x: (x['work_id']), reverse=False)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                sql_num = """
                          select count(*) from work_station
                           """
                dfs = db.execute_sql(conn, sql_num)
                sql_num_int = dfs[0][0]

                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            db.close_connection(conn)
            return HttpResponse(result)
        except Exception as e:
            print(e)
            result = {"code": 0, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("????????????????????????%s"% e)
            return HttpResponse(result)

    def post(self, request):
        """
        ??????????????????????????????

        ??????????????????????????? ??????????????????????????????????????????????????????

        :param request:
        :return:
        """
        try:
            respose_data_json = request.body
            respose_data_dict = json.loads(respose_data_json)
            work_id_str = respose_data_dict.get("work_id")

            work_id = str("product_transit_") + str(work_id_str)
            work_name = respose_data_dict.get("work_name")
            work_type = respose_data_dict.get("work_type")
            leader_work_id = respose_data_dict.get("leader_work_id")

            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            sql = """
             SELECT max(order_number) FROM work_station;
            """
            works = db.execute_sql(conn, sql)
            if works[0][0] == None:
                id_num = 1
            else:
                id_num = works[0][0] + 1
            work_code = str("Im_WorkStation_" + str(id_num))

            sql_insert = """
                               insert into work_station(work_code, work_id, work_name,
                               work_type, leader_work_id, order_number)
                               values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')   
                        """
            sql_insert_format = sql_insert.format(work_code, work_id_str, work_name, work_type, leader_work_id, id_num)
            drs = db.execute_sql(conn, sql_insert_format)

            if work_type == "??????":
                # work_id = "check_" + str(work_id_str)
                # select_table = "product_transit_" +
                sql_create_table = """
                    CREATE TABLE IF NOT EXISTS {0}(product_transit_code varchar(128) primary key not null,
                                         matter_code varchar(128),
                                         matter_id varchar(128),
                                         finished_product_code varchar(128),
                                         user_code varchar(128),
                                         work_code varchar(128),
                                         test_result varchar(128),
                                         description varchar(128),
                                         enter_time datetime,
                                         out_time datetime,
                                         product_plan_code varchar(128),
                                         end_product_code varchar(128),
                                         product_code varchar(128),
                                         order_number int) 
                                     """
                sql_create_table_main = sql_create_table.format(work_id)
                dr = db.execute_sql(conn, sql_create_table_main)

                sql_un_table = str("unqualified_product_") + str(work_id_str)
                sql_create_un_table = """
                CREATE TABLE IF NOT EXISTS {0}(unqualified_product_code varchar(128) primary key not null,
                                                          product_plan_code varchar(128),
                                                          finished_product_code varchar(128),
                                                          matter_code varchar(128),
                                                          matter_id varchar(128),
                                                          description varchar(128),
                                                          leader_work_id varchar(128),
                                                          later_work_id varchar(128),
                                                          solve_method varchar(128),
                                                          solve_result varchar(128),
                                                          order_number int)
                """
                sql_create_un_table = sql_create_un_table.format(sql_un_table)
                db.execute_sql(conn, sql_create_un_table)
                data = {"code": 0, "message": "?????????????????????????????????????????????"}

                data = json.dumps(data)
                logger.info("???????????????????????????????????????????????????%s")
                db.close_connection(conn)
                return HttpResponse(data)
            else:
                data = {"code": 0, "message": "???????????????????????????"}

                data = json.dumps(data)
                logger.info("?????????????????????????????????????????????%s")
                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)

            data = {"code":1, "message":"???????????????????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????%s" % e)

            return HttpResponse(data)


class PutWorkStation(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            work_code = request.GET.get("work_code")
            work_id = request.GET.get("work_id")
            work_name = request.GET.get("work_name")
            work_type = request.GET.get("work_type")
            leader_work_id = request.GET.get("leader_work_id")

            sql = """
                    update work_station set 
                    work_id='{0}',work_name='{1}', work_type ='{2}', leader_work_id = '{3}'
                    where work_code = '{4}'
                    """
            sql_main = sql.format(work_id, work_name, work_type, leader_work_id, work_code)
            drs = db.execute_sql(conn, sql_main)

            data = {"code": 0, "message": "????????????", "data": ""}
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 0, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("????????????????????????%s" % e)

            return HttpResponse(data)


class DeleteWorkStation(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            db = DB()
            table_list = []
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            work_code = request.GET.get("work_code")

            sql_delete = """
            select work_id from work_station where work_code = '{0}'
            """
            sql_delete = sql_delete.format(work_code)
            dres = db.execute_sql(conn, sql_delete)

            print("=>", dres)
            work_id = dres[0][0]
            drop_table = "product_transit_" + str(work_id)
            drop_table1 = "unqualified_product_" + str(work_id)

            sql_tabale = """
                      show tables like "product_transit_%" 
                      """
            dfs = db.execute_sql(conn, sql_tabale)
            if dfs[0][0]:
                for df in dfs:
                    table_list.append(df[0])
            else:
                pass
            if drop_table in table_list:

                sql_chek = """
                select * from {0}
                """
                sql_chek = sql_chek.format(drop_table)
                drss = db.execute_sql(conn, sql_chek)
                if len(drss) == 0:
                    sql = """
                               delete FROM work_station where work_code = "{0}"           
                               """
                    sql_num = sql.format(work_code)
                    drs = db.execute_sql(conn, sql_num)

                    sql_drop = """
                                drop table {0}
                                """
                    sql_drop = sql_drop.format(drop_table)
                    db.execute_sql(conn, sql_drop)

                    sql_drop_t = """
                    drop table {0}
                    """
                    sql_drop_t = sql_drop_t.format(drop_table1)
                    db.execute_sql(conn, sql_drop_t)

                    data = {"code": 0, "message": "????????????"}
                    data = json.dumps(data)
                    db.close_connection(conn)
                    return HttpResponse(data)
                else:
                    data = {"code": 1, "message": "????????????????????????"}
                    data = json.dumps(data)
                    # db.close_connection(conn)
                    return HttpResponse(data)

            else:
                data = {"code": 1, "message": "?????????????????????????????????"}
                data = json.dumps(data)
                # db.close_connection(conn)
                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("????????????????????????%s" % e)

            return HttpResponse(data)


class CreateProductPlanDeal(View):

    def get(self, request):
        try:
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
                create table Product_PlanDeal(prodect_code varchar(128) not null,
                product_plan_code varchar(128) primary key not null,
                plan_name varchar(128), plan_count int, plan_start_day datetime, plan_end_day datetime, 
                description varchar(128), order_number int)         
                """
            dr = db.execute_sql(conn, sql)

            data = {"code": 0, "message":"????????????"}
            # data = json.dumps(data)
        except Exception as e:
            print(e)
            data = {"code":1, "message":"????????????"}

        data = json.dumps(data)

        return HttpResponse(data)


class ProductPlanDeal(View):
    """
    ?????????????????????????????????   ??????????????????????????????????????????
    """

    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            product_id = request.session.get("session_projectId")
            db = DB()
            conn = db.get_connection(product_id)
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))
            product_plan_code = request.GET.get("product_plan_code")
            plan_name = request.GET.get("plan_name")
            plan_status = request.GET.get("plan_status")

            if plan_name:
                plan_name_sql = " and plan_name =" + "'" + plan_name + "'"
            else:
                plan_name_sql = ""
            if plan_status:
                plan_status_sql = " and plan_status =" + "'" + plan_status + "'"
            else:
                plan_status_sql = ""

            sql = """
                select * from product_plandeal where 2 >1 {0} {1}
                """
            sql = sql.format(plan_name_sql, plan_status_sql)
            drs = db.execute_sql(conn, sql)
            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["product_plan_code"] = dr[0]
                    dict_data["plan_name"] = dr[1]
                    dict_data["plan_count"] = dr[2]
                    s = dr[3]
                    s1 = s.strftime("%Y-%m-%d %H:%M:%S ")
                    dict_data["plan_start_day"] = s1
                    d = dr[4]
                    s2 = d.strftime("%Y-%m-%d %H:%M:%S ")
                    dict_data["plan_end_day"] = s2
                    dict_data["description"] = dr[5]
                    dict_data["plan_status"] = dr[6]
                    data_list.append(dict_data)

                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()

                if data_list:

                    sql_num_int = int(len(data_list))
                else:
                    sql_num_int = 0

                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                db.close_connection(conn)
                return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                result = json.dumps(result)
                db.close_connection(conn)
                return HttpResponse(result)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("?????????????????????????????????%s" % e)
            return HttpResponse(data)

    def post(self, request):
        """
        ??????????????????---  ????????????????????????????????????  ????????????????????????????????????

        :param request:
        :return:
        """
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            sql = """
             SELECT max(order_number) FROM product_plandeal;
            """
            works = db.execute_sql(conn, sql)
            if works[0][0] == None:
                id_num = 1
            else:
                id_num = works[0][0] + 1

            order_number = id_num

            response_data_json = request.body
            response_data = json.loads(response_data_json)
            product_plan_code = str("Im_ProductPlan_" + str(id_num))

            plan_name = response_data.get("plan_name")
            plan_count = response_data.get("plan_count")
            plan_start_day = response_data.get("plan_start_day")
            plan_end_day = response_data.get("plan_end_day")
            description = response_data.get("description")
            plan_status = response_data.get("plan_status")
            sql_status = """
            select * from product_plandeal where plan_status = '?????????'
            """
            # sql_status = sql_status.format(plan_status)
            df_status = db.execute_sql(conn, sql_status)
            if df_status:
                if plan_status == "?????????":
                    result = {"code": 1, "message": "??????????????????????????????????????????????????????"}

                    result = json.dumps(result)

                    return HttpResponse(result)
                else:
                    sql_insert = """
                                                   insert into product_plandeal(product_plan_code,
                                                   plan_name, plan_count, plan_start_day, plan_end_day, description, plan_status, order_number)
                                                   values('{0}', '{1}', '{2}', '{3}', '{4}','{5}', '{6}', '{7}')   
                                            """
                    sql_insert_format = sql_insert.format(product_plan_code, plan_name, plan_count,
                                                          plan_start_day, plan_end_day, description, plan_status,
                                                          order_number)
                    drs = db.execute_sql(conn, sql_insert_format)

                    data = {"code": 0, "message": "????????????????????????"}

                    data = json.dumps(data)
                    logger.info("????????????????????????")
                    db.close_connection(conn)
                    return HttpResponse(data)
            else:
                sql_insert = """
                                               insert into product_plandeal(product_plan_code,
                                               plan_name, plan_count, plan_start_day, plan_end_day, description, plan_status, order_number)
                                               values('{0}', '{1}', '{2}', '{3}', '{4}','{5}', '{6}', '{7}')   
                                        """
                sql_insert_format = sql_insert.format(product_plan_code, plan_name, plan_count,
                                                      plan_start_day, plan_end_day, description, plan_status,
                                                      order_number)
                drs = db.execute_sql(conn, sql_insert_format)

                data = {"code": 0, "message": "????????????????????????"}

                data = json.dumps(data)
                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)

            data = {"code": 1, "message": "????????????????????????"}
            data = json.dumps(data)
            logger.error("?????????????????????????????????%s" % e)

            return HttpResponse(data)


class PutProductPlanDeal(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            product_plan_code = request.GET.get("product_plan_code")
            plan_name = request.GET.get("plan_name")
            plan_count = request.GET.get("plan_count")
            plan_start_day = request.GET.get("plan_start_day")
            plan_end_day = request.GET.get("plan_end_day")
            description = request.GET.get("description")
            plan_status = request.GET.get("plan_status")

            sql_status = """
                        select * from product_plandeal where plan_status = '?????????'
                        """
            df_status = db.execute_sql(conn, sql_status)
            if df_status:
                if df_status[0][0] == product_plan_code:
                    sql = """
                        update product_plandeal set plan_name='{0}', plan_count ='{1}', 
                        plan_start_day='{2}',plan_end_day = '{3}', description = '{4}', 
                        plan_status = '{5}' where product_plan_code = '{6}'
                        """
                    sql_main = sql.format(plan_name, plan_count,
                                          plan_start_day, plan_end_day, description, plan_status, product_plan_code)
                    drs = db.execute_sql(conn, sql_main)
                    data = {"code": 0, "message": "????????????", "data": ""}
                    data = json.dumps(data)
                    logger.info("????????????????????????")
                    db.close_connection(conn)

                    return HttpResponse(data)
                else:
                    result = {"code": 1, "message": "???????????????????????????????????????,????????????,????????????????????????"}
                    result = json.dumps(result)
                    logger.error("???????????????????????????????????????,????????????,????????????????????????")
                    db.close_connection(conn)
                    return HttpResponse(result)

            else:
                sql = """
                            update product_plandeal set plan_name='{0}', plan_count ='{1}', 
                            plan_start_day='{2}',plan_end_day = '{3}', description = '{4}', 
                            plan_status = '{5}' where product_plan_code = '{6}'
                      """

                sql_main = sql.format(plan_name, plan_count,
                                      plan_start_day, plan_end_day, description, plan_status, product_plan_code)
                drs = db.execute_sql(conn, sql_main)

                data = {"code": 0, "message": "????????????", "data": ""}
                data = json.dumps(data)
                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("?????????????????????????????????%s" % e)

            return HttpResponse(data)


class DeleteProductPlanDeal(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            product_plan_code = request.GET.get("product_plan_code")
            sql = """
            delete FROM product_plandeal where product_plan_code = "{0}"           
            """
            sql_num = sql.format(product_plan_code)
            drs = db.execute_sql(conn, sql_num)

            data = {"code":0, "message": "????????????", "data":""}

            data = json.dumps(data)
            logger.info("????????????????????????")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code":1, "message": "????????????", "data":e}
            data = json.dumps(data)
            logger.error("?????????????????????????????????%s" % e)

            return HttpResponse(data)


class ProductCodeName(View):
    """
    ????????????????????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            db = DB()
            # product_id = request.session.get("session_projectId")
            # conn = db.get_connection(product_id)
            product_id = "db_common"
            conn = db.get_connection(product_id)
            sql = """
            select prodect_code,product_name from productlist
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_dict = {}
                data_dict["product_code"] = dr[0]
                data_dict["product_name"] = dr[1]
                data_list.append(data_dict)
            data = {"code": 0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("????????????????????????????????????%s" % e)
            return HttpResponse(data)


class CreateProductPickMatter(View):
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            sql = """
                       create table Product_PickMatter(materials_production_code varchar(128) primary key not null,
                       materials_person varchar(128) not null,
                       product_plan_code varchar(128), material_time datetime, description varchar(128), order_number int)         
                       """
            dr = db.execute_sql(conn, sql)

            sql_matter = """
            create table Pick_Matter(materials_code varchar(128) primary key not null, 
            materials_production_code varchar(128) not null,
            matter_code varchar(128), matter_count int, order_number int)
            """
            df = db.execute_sql(conn, sql_matter)

            data = {"code": 0, "message": "????????????"}
            # data = json.dumps(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

        data = json.dumps(data)

        return HttpResponse(data)


class ProductPickMatter(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            data_list = []
            matter_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            materials_person = request.GET.get("materials_person")
            product_plan_code = request.GET.get("product_plan_code")

            if materials_person:
                materials_person_sql = "and materials_person = " + "'" + materials_person + "'"
            else:
                materials_person_sql = ""
            if product_plan_code:
                product_plan_code_sql = "and product_plan_code = " + "'" + product_plan_code + "'"
            else:
                product_plan_code_sql = ""

            sql = """
            select * from product_pickmatter where 2 > 1 {0} {1} 
            """
            sql = sql.format(materials_person_sql, product_plan_code_sql)
            drs = db.execute_sql(conn, sql)

            for dr in drs:
                dict_data = {}
                dict_data["materials_production_code"] = dr[0]
                dict_data["materials_person"] = dr[1]
                dict_data["product_plan_code"] = dr[2]
                s = dr[3]
                s1 = s.strftime("%Y-%m-%d %H:%M:%S ")
                dict_data["material_time"] = s1

                dict_data["description"] = dr[4]
                sql_matter = """
                select * from pick_matter where materials_production_code = '{0}'
                """
                sql_matter_format = sql_matter.format(dr[0])
                matter_dfs = db.execute_sql(conn, sql_matter_format)
                if matter_dfs:
                    respose_list = []
                    for matter_df in matter_dfs:
                        response_data = {}
                        response_data["materials_code"] = matter_df[0]
                        response_data["materials_production_code"] = matter_df[1]
                        response_data["matter_code"] = matter_df[2]
                        response_data["matter_count"] = matter_df[3]
                        respose_list.append(response_data)
                    dict_data["response_datas"] = respose_list
                    data_list.append(dict_data)

            page_result = Page(page, page_size, data_list)
            data = page_result.get_str_json()

            sql_num_int = int(len(data_list))
            data_add_int["data"] = data
            data_add_int["total"] = sql_num_int
            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("????????????????????????")
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:

            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("????????????????????????%s" % e)
            return HttpResponse(result)

    def post(self, request):
        try:

            json_data = request.body
            str_data = json.loads(json_data)

            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            sql_matter = """
                   SELECT max(order_number) FROM pick_matter;
                   """
            matter_type = db.execute_sql(conn, sql_matter)
            if matter_type[0][0] == None:
                matter_num = 0
            else:
                matter_num = matter_type[0][0]

            sql = """
                  SELECT max(order_number) FROM product_pickmatter;
                  """
            product_type = db.execute_sql(conn, sql)
            if product_type[0][0] == None:
                id_num = 1
            else:
                id_num = product_type[0][0] + 1

            materials_production_code = str("Im_Product_Pick_Matter_" + str(id_num))
            materials_person = str_data.get("materials_person")
            product_plan_code = str_data.get("product_plan_code")
            material_time = str_data.get("material_time")
            description = str_data.get("description")
            response_datas = str_data.get("response_datas")

            order_number = id_num

            sql_insert = """
                               insert into product_pickmatter(materials_production_code, materials_person, 
                               product_plan_code,
                               material_time, description, order_number)
                               values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                               """
            sql_insert_format = sql_insert.format(materials_production_code, materials_person,
                                                  product_plan_code, material_time,
                                                  description, order_number)
            drs = db.execute_sql(conn, sql_insert_format)
            for response_data in response_datas:
                # db = DB()
                # conn = db.get_connection("db_common")
                matter_code = response_data.get("matter_code")
                matter_count = response_data.get("matter_count")
                matter_num = matter_num + 1
                materials_code = str("Im_Materials_Pick_" + str(matter_num))
                sql_response = """
                               insert into pick_matter(materials_code, materials_production_code,matter_code, matter_count,
                               order_number)
                               values('{0}', '{1}', '{2}', '{3}', '{4}')
                       """
                sql_response_format = sql_response.format(materials_code, materials_production_code,
                                                          matter_code, matter_count,
                                                          matter_num)
                matter_drs = db.execute_sql(conn, sql_response_format)
                # person_matter_count = ""
                sql_search_num = """
                select matter_count from person_matter
                """
                sql_search_nums = db.execute_sql(conn, sql_search_num)
                if len(sql_search_nums) > 0:
                    if sql_search_nums[0][0]:
                        sql_num = sql_search_nums[0][0]
                        person_matter_count = int(sql_num) - int(matter_count)
                        sql_update = """
                        update person_matter set matter_count = '{0}' where matter_code = '{1}'
                        """
                        sql_update = sql_update.format(person_matter_count, matter_code)
                        db.execute_sql(conn, sql_update)

                #     else:
                #         data = {"code": 1, "message": "?????????????????????????????????count????????????,???????????????????????????person_matter????????????????????????"}
                #         logger.info("?????????????????????????????????????????????????????????count??????????????????????????????????????????person_matter????????????????????????")
                #         data = json.dumps(data)
                #
                # else:
                #     data = {"code": 1, "message": "?????????????????????????????????count????????????"}
                #     logger.info("?????????????????????????????????????????????????????????count??????????????????????????????????????????person_matter????????????????????????")
                #     data = json.dumps(data)

            data = {"code": 0, "message": "?????????????????????,??????????????????????????????????????????"}
            logger.info("?????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "?????????????????????"}
            logger.error("?????????????????????")
            data = json.dumps(data)
            return HttpResponse(data)


class PutProductPickMatter(View):
    """
    ???????????????????????????
    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            materials_production_code = request.GET.get("materials_production_code")
            materials_person = request.GET.get("materials_person")
            product_plan_code = request.GET.get("product_plan_code")
            material_time = request.GET.get("material_time")
            description = request.GET.get("description")

            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            else:
                response_datas = response_datas

            sql = """
                      update product_pickmatter set 
                      materials_person='{0}',product_plan_code='{1}', material_time ='{2}',
                      description= '{3}' where materials_production_code = '{4}'
                      """
            sql_main = sql.format(materials_person, product_plan_code, material_time, description,
                                  materials_production_code)
            drs = db.execute_sql(conn, sql_main)

            sql_matter_num = """
             SELECT max(order_number) FROM pick_matter
            """
            dras = db.execute_sql(conn, sql_matter_num)

            if dras[0][0] != None:
                sql_num = dras[0][0]
            else:
                sql_num = 0

            for response_data in response_datas:
                # materials_production_code = response_data.get("materials_production_code")
                materials_code = response_data.get("materials_code")
                # materials_production_code = response_data.get("materials_production_code")
                matter_code = response_data.get("matter_code")
                matter_count = response_data.get("matter_count")
                if materials_code:
                    sql = """
                       update pick_matter set materials_production_code = '{0}', 
                       matter_code = '{1}', matter_count ='{2}' where materials_code = '{3}'
                    """
                    sql = sql.format(materials_production_code, matter_code, matter_count, materials_code)
                    db.execute_sql(conn, sql)

                    sql_search_num = """
                                    select matter_count from person_matter
                                    """
                    sql_search_nums = db.execute_sql(conn, sql_search_num)
                    if len(sql_search_nums) > 0:
                        if sql_search_nums[0][0]:
                            sql_num = sql_search_nums[0][0]
                            person_matter_count = int(sql_num) - int(matter_count)
                            sql_update = """
                                            update person_matter set matter_count = '{0}' where matter_code = '{1}'
                                            """
                            sql_update = sql_update.format(person_matter_count, matter_code)
                            db.execute_sql(conn, sql_update)


                else:
                    sql_num = sql_num + 1

                    materials_code = str("Im_Materials_Pick_" + str(sql_num))

                    sql_response = """
                                          insert into pick_matter(materials_code,materials_production_code, 
                                          matter_code, matter_count, order_number) values('{0}',
                                           '{1}', '{2}', '{3}', '{4}')
                                  """
                    sql_response_format = sql_response.format(materials_code, materials_production_code,
                                                              matter_code, matter_count, sql_num)
                    db.execute_sql(conn, sql_response_format)

                    sql_search_num = """
                                    select matter_count from person_matter
                                    """
                    sql_search_nums = db.execute_sql(conn, sql_search_num)
                    if len(sql_search_nums) > 0:
                        if sql_search_nums[0][0]:
                            sql_num = sql_search_nums[0][0]
                            person_matter_count = int(sql_num) - int(matter_count)
                            sql_update = """
                                            update person_matter set matter_count = '{0}' where matter_code = '{1}'
                                            """
                            sql_update = sql_update.format(person_matter_count, matter_code)
                            db.execute_sql(conn, sql_update)

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("???????????????????????????")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            logger.error("???????????????????????????---->%s"% e)
            data = json.dumps(data)
            return HttpResponse(data)


class DeleteProductPickMatter(View):
    """
    ???????????????????????????????????????

    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            materials_production_code = request.GET.get("materials_production_code")
            response_datas = request.GET.get("response_datas")
            response_datas = eval(response_datas)
            if response_datas:
                for response_data in response_datas:
                    materials_code = dict(response_data).get("materials_code")
                    sql = """
                       delete from pick_matter where materials_code = '{0}'
                       """
                    sql_format = sql.format(materials_code)
                    db.execute_sql(conn, sql_format)
            else:
                sql_matter = """
                                  SELECT * FROM product_pickmatter where materials_production_code = '{0}'
                                  """
                sql_matter_format = sql_matter.format(materials_production_code)
                matter_dfs = db.execute_sql(conn, sql_matter_format)
                for matter_df in matter_dfs:
                    materials_production_code = matter_df[0]
                    sql_safe = """
                    set sql_safe_updates=0;
                    """
                    db.execute_sql(conn, sql_safe)
                    sql_delete = """
                       delete FROM pick_matter where materials_production_code = '{0}'
                       """
                    sql_delete_format = sql_delete.format(materials_production_code)
                    db.execute_sql(conn, sql_delete_format)
                    sql_safe1 = """
                                       set sql_safe_updates=1;
                                       """
                    db.execute_sql(conn, sql_safe1)
                sql = """
                       delete from product_pickmatter where materials_production_code = '{0}'
                       """
                sql_main = sql.format(materials_production_code)
                des = db.execute_sql(conn, sql_main)

            data = {"code": 0, "message": "????????????"}
            logger.info("???????????????????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("???????????????????????????????????????,%s" % e)
            return HttpResponse(data)


class NoPagePersonMatter(View):
    """
    ????????????????????????,???????????????code name ?????? ??????
    """

    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            # conn = db.get_connection("db_common")
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            sql = """
            select pm.matter_code, ml.matter_name, ml.rule, ml.matter_category, 
            pm.matter_count,pm.product_time,ml.code_length from person_matter as pm
            left join matter_list as ml on pm.matter_code = ml.bom_matter_code
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                dict_data = {}
                dict_data["matter_code"] = dr[0]
                dict_data["matter_name"] = dr[1]
                dict_data["rule"] = dr[2]
                dict_data["matter_category"] = dr[3]
                dict_data["matter_count"] = dr[4]
                s = dr[5]
                if s:
                    s1 = s.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    s1 = ""
                dict_data["product_time"] = s1
                dict_data["code_length"] = dr[6]
                data_list.append(dict_data)
            if data_list:
                sql_num_int = int(len(data_list))
            else:
                sql_num_int = 0

            data_add_int["data"] = data_list
            data_add_int["total"] = sql_num_int
            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("????????????????????????????????????")
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("????????????????????????????????????")
            return HttpResponse(result)


class CreateProcessDeal(View):
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            sql = """
                              create table Process_Deal(production_code varchar(128) primary key not null,
                              production_id varchar(128) not null,
                              production_name varchar(128), prodect_code varchar(128), work_id varchar(128),
                               description varchar(128), order_number int)         
                              """
            dr = db.execute_sql(conn, sql)
            data = {"code": 0, "message": "????????????"}
            # data = json.dumps(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
        data = json.dumps(data)

        return HttpResponse(data)


class ProcessDeal(View):
    """
    ??????????????????
    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            current_person_id = request.session['session_currentId']
            conn = db.get_connection(product_id)
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))
            production_name = request.GET.get("production_name")
            work_id = request.GET.get("work_id")
            if production_name:
                production_name_sql = "and production_name =" + "'"+ production_name+"'"
            else:
                production_name_sql = ""

            if work_id:
                work_id_sql = "and work_id = " + "'" + work_id + "'"
            else:
                work_id_sql = ""

            sql = """
                       select * from  process_deal where 2>1 {0} {1}
                       """

            sql_main = sql.format(production_name_sql, work_id_sql)

            drs = db.execute_sql(conn, sql_main)
            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["production_code"] = dr[0]
                    dict_data["production_id"] = dr[1]
                    dict_data["production_name"] = dr[2]
                    dict_data["work_id"] = dr[3]
                    # dict_data["leader_work_id"] = dr[4]
                    dict_data["description"] = dr[4]
                    sql_matter = """
                            select * from process_matter_deal where work_id = '{0}'
                             """
                    sql_matter_format = sql_matter.format(dr[3])
                    matter_dfs = db.execute_sql(conn, sql_matter_format)
                    if matter_dfs:
                        respose_list = []
                        for matter_df in matter_dfs:
                            response_data = {}
                            response_data["process_matter_deal_code"] = matter_df[0]
                            response_data["work_id"] = matter_df[1]
                            response_data["matter_code"] = matter_df[2]
                            response_data["install_number"] = matter_df[3]
                            response_data["production_code"] = dr[0]
                            respose_list.append(response_data)
                        dict_data["response_datas"] = respose_list
                        data_list.append(dict_data)
                    else:
                        dict_data["response_datas"] = []
                        data_list.append(dict_data)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                sql_num_int = int(len(data_list))
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
                data = {"code": 0, "message": "????????????", "data": data_add_int}
                data = json.dumps(data)
                logger.info("??????????????????%s")
                db.close_connection(conn)
                return HttpResponse(data)

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                data = {"code": 0, "message": "??????????????????", "data": data_add_int}
                data = json.dumps(data)
                logger.info("??????????????????,????????????????????????")
                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("??????????????????%s"% e)
            return HttpResponse(data)

    def post(self, request):
        """
        ????????????---->  ????????????????????????????????????

        :param request:
        :return:
        """
        try:
            table_list = []
            db = DB()
            product_id = request.session.get("session_projectId")
            current_person_id = request.session['session_currentId']
            conn = db.get_connection(product_id)
            sql = """
                    SELECT max(order_number) FROM process_deal;
                   """
            works = db.execute_sql(conn, sql)
            if works[0][0] == None:
                id_num = 1
            else:
                id_num = works[0][0] + 1

            production_code = str("Im_Process_" + str(id_num))
            respose_data_json = request.body
            respose_data_dict = json.loads(respose_data_json)

            production_id = respose_data_dict.get("production_id")
            production_name = respose_data_dict.get("production_name")
            work_id = respose_data_dict.get("work_id")
            description = respose_data_dict.get("description")
            response_datas = respose_data_dict.get("response_datas")

            sql_check = """
            select * from process_deal where work_id = '{0}'
            """
            sql_check = sql_check.format(work_id)
            df_checks = db.execute_sql(conn, sql_check)
            if df_checks:
                data = {"code": 1, "message": "????????????????????????????????????????????????????????????????????????"}
                data = json.dumps(data)
                logger.error("????????????????????????????????????????????????????????????????????????")
                return HttpResponse(data)
            else:
                sql_insert = """          
                             insert into process_deal(production_code, production_id, production_name,
                              work_id, description, order_number)
                              values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')   
                            """
                sql_insert_format = sql_insert.format(production_code, production_id, production_name,
                                                      work_id, description, id_num)
                drs = db.execute_sql(conn, sql_insert_format)
                if response_datas:
                    for response_data in response_datas:
                        process_matter_deal_code = response_data.get("process_matter_deal_code")
                        matter_code = response_data.get("matter_code")
                        install_number = response_data.get("install_number")
                        if process_matter_deal_code:
                            sql_update = """
                            update process_matter_deal set matter_code = '{0}', 
                            install_number = '{1}',
                            work_id = '{2}',
                            where process_matter_deal_code = '{3}'
                            """
                            sql_update = sql_update.format(matter_code, install_number, work_id,
                                                           process_matter_deal_code)
                            drs = db.execute_sql(conn, sql_update)
                        else:

                            sql_matter_num = """
                                            SELECT max(order_number) FROM process_matter_deal
                                           """
                            dras = db.execute_sql(conn, sql_matter_num)
                            if dras[0][0] == None:
                                id_num = 1
                            else:
                                id_num = dras[0][0] + 1
                            process_matter_deal_code = str("Im_process_matter_deal_" + str(id_num))
                            sql_insert = """
                            insert into process_matter_deal(process_matter_deal_code, work_id, matter_code,
                                              install_number, order_number)
                                              values('{0}', '{1}', '{2}', '{3}', '{4}')   
                            """
                            sql_insert = sql_insert.format(process_matter_deal_code, work_id,
                                                           matter_code, install_number, id_num)
                            drs = db.execute_sql(conn, sql_insert)
                else:
                    pass

                data = {"code": 0, "message": "???????????????????????????"}

                data = json.dumps(data)
                logger.info("???????????????????????????%s" )
                db.close_connection(conn)

                return HttpResponse(data)
        except Exception as e:
            print(e)

            data = {"code": 1, "message": "???????????????????????????"}
            data = json.dumps(data)
            logger.error("???????????????????????????%s" % e)

            return HttpResponse(data)


class PutProcessDeal(View):
    """
    ????????????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            production_code = request.GET.get("production_code")
            production_id = request.GET.get("production_id")
            production_name = request.GET.get("production_name")
            prodect_code = request.GET.get("prodect_code")
            work_id = request.GET.get("work_id")
            description = request.GET.get("description")
            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            for response_data in response_datas:
                # response_data = dict(response_data)
                process_matter_deal_code = response_data.get("process_matter_deal_code")
                install_number = response_data.get("install_number")
                if process_matter_deal_code:
                    sql = """
                    update process_matter_deal set install_number = '{0}' where process_matter_deal_code = '{1}'
                    """
                    sql = sql.format(install_number, process_matter_deal_code)
                    db.execute_sql(conn, sql)
                else:
                    matter_code = response_data.get("matter_code")
                    sql_matter_num = """
                                        SELECT max(order_number) FROM process_matter_deal
                                       """
                    dras = db.execute_sql(conn, sql_matter_num)
                    if dras[0][0] == None:
                        id_num = 1
                    else:
                        id_num = dras[0][0] + 1
                    process_matter_deal_code = str("Im_process_matter_deal_" + str(id_num))
                    sql_insert = """
                                            insert into process_matter_deal(process_matter_deal_code, work_id, matter_code,
                                                              install_number, order_number)
                                                              values('{0}', '{1}', '{2}', '{3}', '{4}')   
                                            """
                    sql_insert = sql_insert.format(process_matter_deal_code, work_id,
                                                   matter_code, install_number, id_num)
                    drs = db.execute_sql(conn, sql_insert)
            sql = """
                            update process_deal set 
                            production_id='{0}', production_name ='{1}',
                            work_id = '{2}', description = '{3}' 
                            where production_code = '{4}'
                            """
            sql_main = sql.format(production_id, production_name, work_id,
                            description, production_code)
            drs = db.execute_sql(conn, sql_main)
            data = {"code": 0, "message": "????????????", "data": ""}
            data = json.dumps(data)
            logger.info("????????????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 0, "message": "????????????", "data":e}
            data = json.dumps(data)
            logger.error("????????????????????????%s"% e)
            return HttpResponse(data)


class DeleteProcessDeal(View):
    """
    ????????????????????????----> ???????????????????????????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            work_id = request.session.get("session_workId")
            conn = db.get_connection(product_id)

            production_code = request.GET.get("production_code")

            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            if response_datas:
                for response_data in response_datas:
                    process_matter_deal_code = dict(response_data).get("process_matter_deal_code")
                    sql = """
                       delete from process_matter_deal where process_matter_deal_code = '{0}'
                       """
                    sql_format = sql.format(process_matter_deal_code)
                    db.execute_sql(conn, sql_format)

            else:
                process_list = []
                sql_matter = """
                              SELECT work_id FROM process_deal where production_code = '{0}'
                              """
                sql_matter_format = sql_matter.format(production_code)
                matter_dfs = db.execute_sql(conn, sql_matter_format)
                if matter_dfs:
                    matter_df = matter_dfs[0][0]
                else:
                    matter_df = ""

                sql_delete = """
                                SELECT process_matter_deal_code FROM process_matter_deal where work_id = '{0}'
                               """
                sql_matter_format = sql_delete.format(matter_df)
                dfs = db.execute_sql(conn, sql_matter_format)
                for df in dfs:
                    sql_de = """
                                   delete from process_matter_deal where process_matter_deal_code = '{0}'
                                   """
                    sql_de = sql_de.format(df[0])
                    db.execute_sql(conn, sql_de)

                sql = """
                       delete from process_deal where production_code = '{0}'
                       """
                sql_format = sql.format(production_code)
                db.execute_sql(conn, sql_format)

            data = {"code": 0, "message": "????????????", "data": ""}
            data = json.dumps(data)
            logger.info("???????????????????????????????????????????????????%s" )
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message":"????????????", "data": e}
            data = json.dumps(data)
            logger.error("???????????????????????????????????????????????????%s"% e)
            return HttpResponse(data)


class ProductPlanCodeName(View):
    """
    ???????????????????????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            sql = """
            SELECT product_plan_code, plan_name FROM product_plandeal
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_dict = {}
                data_dict["product_plan_code"] = dr[0]
                data_dict["product_plan_name"] = dr[1]
                data_list.append(data_dict)
            data = {"code":0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            logger.info("?????????????????????????????????????????????%s" )
            db.close_connection(conn)
            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("?????????????????????????????????????????????%s" % e)
            return HttpResponse(data)


class PersonCodeName(View):
    """
    ???????????????code???name???????????????
    """
    def get(self, request):
        try:
            data_list = []
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            sql = """
                       SELECT user_code, user_name FROM person
                       """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_dict = {}
                data_dict["user_code"] = dr[0]
                data_dict["user_name"] = dr[1]
                data_list.append(data_dict)
            data = {"code": 0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            logger.info("????????????????????????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code":1, "message":"????????????", "data":e}
            data = json.dumps(data)
            logger.error("????????????????????????????????????%s" % e)

            return HttpResponse(data)


class WorkCodeName(View):
    """
    ??????????????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            db = DB()
            # project_id = request.GET.get("project_id")
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            # conn = db.get_connection(project_id)
            sql = """
                       SELECT work_code, work_id, work_name FROM work_station
                       """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_dict = {}
                data_dict["work_code"] = dr[0]
                data_dict["work_id"] = dr[1]
                data_dict["work_name"] = dr[2]
                # data_dict["leader_work_id"] = dr[3]
                # data_dict["work_type"] = dr[4]
                data_list.append(data_dict)
            data_list.sort(key=lambda x: (x['work_id']), reverse=False)
            data = {"code": 0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            logger.info("????????????????????????????????????%s")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data":e}
            data = json.dumps(data)
            logger.error("????????????????????????????????????%s"% e)

            return HttpResponse(data)


class CreateCheckProductDeal(View):
    def get(self, request):
        try:
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
                                     create table Check_ProductDeal(check_code varchar(128) primary key not null,
                                     prodect_code varchar(128),
                                     work_code varchar(128) not null,
                                     production_code varchar(128), check_method varchar(128), order_number int)         
                                     """
            dr = db.execute_sql(conn, sql)
            data = {"code": 0, "message": "????????????"}

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
        data = json.dumps(data)

        return HttpResponse(data)


class CreateProductParameter(View):
    def get(self, request):
        try:
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
                                     create table Product_Parameter(test_code varchar(128) primary key not null,
                                     check_code varchar(128), test_parameter varchar(128),
                                     test_parameter_count varchar(128),
                                     test_status varchar(128), order_number int)         
                                     """
            dr = db.execute_sql(conn, sql)
            data = {"code": 0, "message": "????????????"}

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
        data = json.dumps(data)

        return HttpResponse(data)


class CheckProductDeal(View):

    """
    c????????????????????????????????????  ----??? ????????????????????????

    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            prodect_code = request.GET.get("prodect_code")
            work_code = request.GET.get("work_code")
            production_code = request.GET.get("production_code")
            check_method = request.GET.get("check_method")

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            if prodect_code:
                prodect_code_sql = " and prodect_code =" + "'" + prodect_code + "'"
            else:
                prodect_code_sql = ""
            if work_code:
                work_code_sql = " and work_code = " + "'" + work_code + "'"
            else:
                work_code_sql = ""
            if production_code:
                production_code_sql = " and production_code = " + "'" + production_code+"'"
            else:
                production_code_sql = ""
            if check_method:
                check_method_sql = " and check_method = " + "'"+check_method+"'"
            else:
                check_method_sql = ""

            sql = """
             select * from check_productdeal where 2 > 1 {0} {1} {2} {3}
            """
            sql_main = sql.format(prodect_code_sql, work_code_sql, production_code, check_method_sql)

            # sql_main = sql.format(matter_id_sql, matter_name_sql, matter_category_sql, status_sql)
            # print(sql_main)
            drs = db.execute_sql(conn, sql_main)
            for dr in drs:
                dict_data = {}
                dict_data["check_code"] = dr[0]
                dict_data["prodect_code"] = dr[1]
                dict_data["work_code"] = dr[2]
                dict_data["production_code"] = dr[3]
                dict_data["check_method"] = dr[4]
                # data_list.append(dict_data)

                sql_matter = """
                            select * from product_parameter where check_code = '{0}'
                            """
                sql_matter_format = sql_matter.format(dr[0])
                matter_dfs = db.execute_sql(conn, sql_matter_format)
                if matter_dfs:
                    respose_list = []
                    for matter_df in matter_dfs:
                        response_data = {}
                        response_data["test_code"] = matter_df[0]
                        response_data["check_code"] = matter_df[1]
                        response_data["test_parameter"] = matter_df[2]
                        response_data["test_parameter_count"] = matter_df[3]
                        response_data["test_status"] = matter_df[4]
                        respose_list.append(response_data)
                    dict_data["response_datas"] = respose_list
                    data_list.append(dict_data)
                else:
                    dict_data["response_datas"] = []
                    data_list.append(dict_data)

            page_result = Page(page, page_size, data_list)
            data = page_result.get_str_json()

            sql_num = """
                                        select count(*) from check_productdeal 
                                        where 2 > 1 {0} {1} {2} {3}
                                          """
            sql_format = sql_num.format(prodect_code_sql, work_code_sql, production_code, check_method_sql)
            dfs = db.execute_sql(conn, sql_format)
            sql_num_int = dfs[0][0]

            data_add_int["data"] = data
            data_add_int["total"] = sql_num_int

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("c??????????????????????????????????????????%s")
            db.close_connection(conn)

            return HttpResponse(result)

        except Exception as e:
            logger.error("??????check_productdeal??????????????????%s"% e)
            result = {"code": 1, "message": "????????????", "data": ""}
            result = json.dumps(result)
            print(e)
            return HttpResponse(result)

    def post(self, request):

        """
        ?????????????????????????????????   ???????????????
        :param request:
        :return:
        """

        db = DB()
        product_id = request.session.get("session_projectId")
        conn = db.get_connection(product_id)

        try:
            json_data = request.body
            str_data = json.loads(json_data)
            sql_matter = """
                          SELECT max(order_number) FROM check_productdeal;
                          """
            checkproduct = db.execute_sql(conn, sql_matter)
            if checkproduct[0][0] == None:
                matter_num = 1
            else:
                matter_num = checkproduct[0][0] + 1
            sql = """
                 SELECT max(order_number) FROM product_parameter;
                 """
            product_parameter = db.execute_sql(conn, sql)
            if product_parameter[0][0] == None:
                id_num = 0
            else:
                id_num = product_parameter[0][0]

            check_code = str("Im_Check_Product_" + str(matter_num))
            prodect_code = str_data.get("prodect_code")
            work_code = str_data.get("work_code")
            production_code = str_data.get("production_code")
            check_method = str_data.get("check_method")
            response_datas = str_data.get("response_datas")

            sql_insert = """
                                      insert into check_productdeal(check_code, prodect_code, 
                                      work_code,
                                      production_code, check_method, order_number)
                                      values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                                      """
            sql_insert_format = sql_insert.format(check_code, prodect_code,
                                                  work_code, production_code,
                                                  check_method, matter_num)
            drs = db.execute_sql(conn, sql_insert_format)
            for response_data in response_datas:
                # db = DB()
                # conn = db.get_connection("db_common")
                test_parameter = response_data.get("test_parameter")
                test_parameter_count = response_data.get("test_parameter_count")
                test_status = response_data.get("test_status")
                id_num = id_num + 1
                test_code = str("Im_Test_" + str(id_num))
                sql_response = """
                                      insert into product_parameter(test_code, check_code, 
                                      test_parameter,test_parameter_count, test_status,
                                      order_number)
                                      values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                              """
                sql_response_format = sql_response.format(test_code, check_code, test_parameter,
                                                          test_parameter_count, test_status,
                                                          id_num)
                matter_drs = db.execute_sql(conn, sql_response_format)

            data = {"code": 0, "message": "???????????????????????????"}
            logger.info("???????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "???????????????????????????"}
            logger.error("???????????????????????????%s" %e)
            data = json.dumps(data)
            return HttpResponse(data)


class PutCheckProductDeal(View):
    def get(self, request):
        """
        ???????????????????????????????????? ???????????????
        :param request:
        :return:
        """
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            check_code = request.GET.get("check_code")
            prodect_code = request.GET.get("prodect_code")
            work_code = request.GET.get("work_code")
            production_code = request.GET.get("production_code")
            check_method = request.GET.get("check_method")

            response_datas = request.GET.get("response_datas")

            if response_datas:
                response_datas = eval(response_datas)
            else:
                response_datas = response_datas

            sql = """
                   update check_productdeal set 
                   prodect_code='{0}',work_code='{1}', production_code ='{2}'
                   ,check_method = '{3}' where check_code = '{4}'
                   """
            sql_main = sql.format(prodect_code, work_code, production_code, check_method, check_code)
            drs = db.execute_sql(conn, sql_main)

            for response_data in response_datas:
                test_code = response_data.get("test_code")
                if test_code:
                    sql_delete = """
                     delete FROM product_parameter where test_code = '{0}'
                    """
                    sql_delete_format = sql_delete.format(test_code)
                    db.execute_sql(conn, sql_delete_format)

            sql_after = """
                        SELECT max(order_number) FROM product_parameter
                        """
            num = db.execute_sql(conn, sql_after)
            if num[0][0]:
                sql_num = num[0][0]
            else:
                sql_num = 0

            for response_data in response_datas:
                # db = DB()
                # conn = db.get_connection("db_common")
                # check_code = response_data.get("check_code")
                test_parameter = response_data.get("test_parameter")
                test_parameter_count = response_data.get("test_parameter_count")
                test_status = response_data.get("test_status")
                sql_num = sql_num + 1
                test_code = str("Im_Test_" + str(sql_num))
                sql_response = """
                                    insert into product_parameter(test_code,check_code, test_parameter, test_parameter_count,
                                    test_status,order_number)
                                    values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                            """
                sql_response_format = sql_response.format(test_code, check_code, test_parameter, test_parameter_count,
                                                          test_status, sql_num)
                matter_drs = db.execute_sql(conn, sql_response_format)
            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)

            db.close_connection(conn)
            logger.info("???????????????????????????????????????%s")

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            logger.error("????????????---->%s"% e)
            data = json.dumps(data)
            return HttpResponse(data)


class DeleteCheckProductDeal(View):
    """
    ???????????????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            check_code = request.GET.get("check_code")
            response_datas = request.GET.get("response_datas")
            response_datas = eval(response_datas)
            # materials_production_code = "Im_Product_Pick_Matter_1"
            # response_datas = [{"materials_code": "Im_Materials_Pick_1"}]
            if response_datas:
                for response_data in response_datas:
                    test_code = dict(response_data).get("test_code")
                    sql = """
                              delete from product_parameter where test_code = '{0}'
                              """
                    sql_format = sql.format(test_code)
                    db.execute_sql(conn, sql_format)
            else:
                sql_matter = """
                                         SELECT * FROM check_productdeal where check_code = '{0}'
                                         """
                sql_matter_format = sql_matter.format(check_code)
                matter_dfs = db.execute_sql(conn, sql_matter_format)
                for matter_df in matter_dfs:
                    test_code = matter_df[0]
                    sql_delete = """
                              delete FROM product_parameter where test_code = '{0}'
                              """
                    sql_delete_format = sql_delete.format(test_code)
                    db.execute_sql(conn, sql_delete_format)
                sql = """
                              delete from check_productdeal where check_code = '{0}'
                              """
                sql_main = sql.format(check_code)
                des = db.execute_sql(conn, sql_main)

            data = {"code": 0, "message": "????????????"}
            logger.info("?????????????????????????????????")

            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("?????????????????????????????????,%s" % e)
            return HttpResponse(data)


class CreateProductTransitInfo(View):

    def get(self, request):
        try:
            db = DB()
            conn = db.get_connection("db_common")
            sql = """
                                     create table ProductTransitInfo(product_transit_code varchar(128) primary key not null,
                                     matter_code varchar(128), user_code varchar(128),
                                     work_code varchar(128),
                                     test_result varchar(128), 
                                     enter_time datetime,
                                     out_time datetime,
                                     product_plan_code varchar(128),
                                     end_product_code varchar(128),
                                     product_code varchar(128),                                     
                                     order_number int)         
                                     """
            dr = db.execute_sql(conn, sql)
            data = {"code": 0, "message": "????????????"}

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
        data = json.dumps(data)

        return HttpResponse(data)


class ProductTransitInfo(View):
    """
    ????????????????????????

    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            session_work_id = request.session.get("session_workId")
            conn = db.get_connection(product_id)
            sql_work_code = """
                       SELECT work_code FROM work_station where work_id = '{0}'
                       """
            sql_work_code_main = sql_work_code.format(session_work_id)
            drss = db.execute_sql(conn, sql_work_code_main)
            work_code = drss[0][0]

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            matter_code = request.GET.get("matter_code")
            finished_product_code = request.GET.get("finished_product_code")
            matter_id = request.GET.get("matter_id")
            user_code = request.GET.get("user_code")
            description = request.GET.get("description")
            # work_code = request.GET.get("work_code")
            test_result = request.GET.get("test_result")
            enter_time = request.GET.get("enter_time")
            out_time = request.GET.get("out_time")
            product_plan_code = request.GET.get("product_plan_code")
            product_code = request.GET.get("product_code")


            if matter_code:
                matter_code_sql = " and matter_code =" + "'" + matter_code + "'"
            else:
                matter_code_sql = ""
            if user_code:
                user_code_sql = " and user_code = " + "'" + user_code + "'"
            else:
                user_code_sql = ""
            if work_code:
                work_code_sql = " and work_code = " + "'" + work_code+"'"
            else:
                work_code_sql = ""
            if test_result:
                test_result_sql = " and test_result = " + "'"+test_result+"'"
            else:
                test_result_sql = ""
            # if enter_time:
            #     enter_time_sql = " and enter_time = " + "'"+enter_time+"'"
            # else:
            #     enter_time_sql = ""
            if out_time:
                out_time_sql = " and out_time = " + "'"+out_time+"'"
            else:
                out_time_sql = ""
            if product_plan_code:
                product_plan_code_sql = " and product_plan_code = " + "'"+product_plan_code+"'"
            else:
                product_plan_code_sql = ""
            if product_code:
                product_code_sql = " and product_code = " + "'" + product_code + "'"
            else:
                product_code_sql = ""

            if finished_product_code:
                finished_product_code_sql = " and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""
            select_table = str("product_transit_") + str(session_work_id)

            sql = """
             select * from {0} where 2 > 1 {1} {2} {3} {4} {5} {6} {7} {8} 
            """

            sql_main = sql.format(select_table, matter_code_sql, finished_product_code_sql, user_code_sql,
                                  work_code_sql, test_result_sql,
                                  out_time_sql,
                                  product_plan_code_sql, product_code_sql)
            drs = db.execute_sql(conn, sql_main)

            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["product_transit_code"] = dr[0]
                    dict_data["matter_code"] = dr[1]
                    dict_data["matter_id"] = dr[2]
                    dict_data["finished_product_code"] = dr[3]
                    dict_data["user_code"] = dr[4]
                    dict_data["work_code"] = dr[5]
                    dict_data["test_result"] = dr[6]
                    dict_data["description"] = dr[7]
                    if dr[8]:
                        dp = dr[8]
                        dp = dp.strftime("%Y-%m-%d %H:%M:%S ")
                        dict_data["enter_time"] = dp
                    else:
                        dict_data["enter_time"] = ""
                    ds = dr[9]
                    sq = ds.strftime("%Y-%m-%d %H:%M:%S ")
                    dict_data["out_time"] = sq
                    dict_data["product_plan_code"] = dr[10]
                    dict_data["end_product_code"] = dr[11]
                    dict_data["product_code"] = dr[12]
                    data_list.append(dict_data)

                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                # sql_num = """
                #            select count(*) from {0} where 2 > 1 {1} {2} {3} {4} {5} {6} {7}
                #            """
                # sql_num_format = sql_num.format(select_table, matter_code_sql, user_code_sql, work_code_sql,
                #                                 test_result_sql,
                #                                  out_time_sql, product_plan_code_sql,
                #                                 product_code_sql)
                # dfs = db.execute_sql(conn, sql_num_format)
                dfs = int(len(data_list))
                sql_num_int = dfs

                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int

                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("??????????????????????????????%s")
                db.close_connection(conn)
                return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "?????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("????????????????????????????????????????????????%s")
                db.close_connection(conn)
                return HttpResponse(result)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("??????????????????????????????%s"% e)

            return HttpResponse(data)

    # def get_first_workstation(self):
    #     try:
    #         pass
    #     except Exception as e:
    #         print(e)
    #         data = "??????????????????????????????"
    #
    #         return data

    def post(self, request):
        """
        ??????????????????????????????   ?????????????????????????????????????????????????????????product_transit_(work_id) ???????????????,  ???????????????????????????
        unqualified_product_???work_id?????????
        :param request:
        :return:
        """

        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            work_type = request.session.get('session_workType')

            str_data = request.body
            json_data = json.loads(str_data)
            session_work_id = request.session.get("session_workId")
            user_code = request.session.get("session_currentId")
            conn = db.get_connection(product_id)
            response_datas = json_data.get("response_datas")
            matter_code_list = []
            matter_id_list = []
            for response_data in response_datas:
                matter_code_list.append(dict(response_data).get("matter_code"))
                matter_id_list.append(dict(response_data).get("matter_id"))
            matter_code = matter_code_list
            matter_id = matter_id_list
            finished_product_code = json_data.get("finished_product_code")
            test_result = json_data.get("test_result")
            description = json_data.get("description")

            out_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ")
            product_plan_code = json_data.get("product_plan_code")
            product_code = json_data.get("product_code")


            # #?????????????????????
            # sql_work_one = """
            # SELECT * FROM work_station where work_type = "??????";
            # """
            # # sql_work_one =sql_work_one.format()
            # dr_works = db.execute_sql(conn, sql_work_one)
            # if len(dr_works) > 0:
            #     for dr_work in dr_works:
            #         if dr_work[2]:
            #             pass
            #         else:
            #             first_work_id = dr_work[1]
            # else:
            #     data = {"code": 1, "message": "????????????,?????????????????????,????????????"}
            #     data = json.dumps(data)
            #     return HttpResponse(data)
            #
            # # ????????????????????????????????????????????????????????????????????????????????????????????????????????????????????????
            #
            # sql_plan_count = """
            # select plan_count, plan_start_day from product_plandeal
            # """
            # plan_counts = db.execute_sql(conn, sql_plan_count)
            # if len(plan_counts) > 0:
            #     if plan_counts[0][0]:
            #         plan_count = plan_counts[0][0]
            #         plan_start_day = plan_counts[0][1]
            #     else:
            #         data = {"code": 1, "message": "????????????,?????????????????????????????????????????????"}
            #         data = json.dumps(data)
            #         return HttpResponse(data)
            # else:
            #     data = {"code": 1, "message": "????????????,?????????????????????????????????????????????"}
            #     data = json.dumps(data)
            #     return HttpResponse(data)
            #
            # select_table = "product_transit_" + str(first_work_id)
            # sql_relity_count = """
            # select * from {0} where 2 > 1 and product_plan_code = '{1}' group by finished_product_code
            # """
            # sql_relity_count = sql_relity_count.format(select_table, product_plan_code)
            # relity_counts = db.execute_sql(conn, sql_relity_count)
            # if len(relity_counts) > 0:
            #     relity_count = int(len(relity_counts))
            #     # if relity_counts[0][0]:
            #     #     relity_count = relity_counts[0][0]
            #     #     # plan_start_day = plan_counts[0][1]
            #     # else:
            #     #     relity_count = 0
            # else:
            #     relity_count = 0
            #
            # if relity_count <= plan_count:
            #     pass
            # else:
            #     # data = "??????????????????????????????????????????"
            #     data = {"code": 1, "message": "????????????,??????????????????????????????????????????"}
            #     data = json.dumps(data)
            #     return HttpResponse(data)


            if test_result == "FAIL":
                tables_list = []

                un_pro_table = str("unqualified_product_") + str(session_work_id)

                sql_unqualified_product = """
                                         SELECT max(order_number) FROM {0};
                                         """
                sql_unqualified_product = sql_unqualified_product.format(un_pro_table)
                unqualified_product_nums = db.execute_sql(conn, sql_unqualified_product)
                if unqualified_product_nums[0][0]:
                    matter_num = unqualified_product_nums[0][0]
                else:
                    matter_num = 0

                s = 0

                for matter_code_one in matter_code:
                    matter_num = matter_num + 1
                    matter_id_one = matter_id_list[s]
                    unqualified_product_code = str("Im_Unqualified_Product_" + str(matter_num))
                    sql_insert = """
                        insert into {0}(unqualified_product_code, product_plan_code, matter_code, matter_id, 
                        finished_product_code,
                        description, order_number)
                                        values('{1}', '{2}', '{3}', '{4}', '{5}', '{6}', '{7}')
                    """
                    sql_insert_main = sql_insert.format(un_pro_table, unqualified_product_code,
                                                        product_plan_code, matter_code_one, matter_id_one,
                                                        finished_product_code, description, matter_num)
                    db.execute_sql(conn, sql_insert_main)
                    s += 1
            sql_work_code = """
            SELECT work_code FROM work_station where work_id = '{0}'
            """
            sql_work_code_main = sql_work_code.format(session_work_id)
            drss = db.execute_sql(conn, sql_work_code_main)
            work_code = drss[0][0]

            select_table = str("product_transit_") + str(session_work_id)

            dater = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sql_producttransit = """
                                      SELECT max(order_number) FROM {0};
                                      """
            sql_producttransit_main = sql_producttransit.format(select_table)
            producttransit = db.execute_sql(conn, sql_producttransit_main)
            if producttransit[0][0] == None:
                matter_num = 0
            else:
                matter_num = producttransit[0][0]
            # end_product_code = product_id + "_" + dater + "_" + str(matter_num + 1)
            s = 0
            reponse_data_list = []
            for matter_code_one in matter_code:
                dict_data = {}

                matter_num = matter_num + 1
                matter_id_one = matter_id_list[s]
                product_transit_code = str("Im_Product_Transit_" + str(matter_num))

                sql_insert = """
                                            insert into {0}(product_transit_code, matter_code, matter_id,
                                             finished_product_code, user_code,
                                              work_code, test_result, description, out_time, product_plan_code, 
                                              end_product_code, product_code,
                                              order_number)
                                              values('{1}', '{2}', '{3}', '{4}', '{5}', '{6}','{7}', 
                                              '{8}', '{9}', '{10}', '{11}', '{12}','{13}')
                                            """
                sql_insert_main = sql_insert.format(select_table, product_transit_code, matter_code_one, matter_id_one,
                                                    finished_product_code, user_code, work_code,
                                                    test_result, description, out_time,
                                                    product_plan_code, matter_id_one, product_code, matter_num)
                drs = db.execute_sql(conn, sql_insert_main)
                s +=1

            cn = db.close_connection(conn)

            data = {"code": 0, "message": "????????????", "data": reponse_data_list}
            data = json.dumps(data)
            logger.info("??????????????????????????????%s")

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("????????????????????????%s"% e)
            return HttpResponse(data)


class MatterNameCode(View):
    """
    ???????????????????????????????????????
    """
    def get(self, request):
        try:
            data_list = []

            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            sql = """
             SELECT pm.matter_code, ml.matter_name FROM person_matter as pm left join matter_list as ml
              on pm.matter_code = ml.bom_matter_code
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_dict = {}
                data_dict["matter_code"] = dr[0]
                data_dict["matter_name"] = dr[1]
                data_list.append(data_dict)

            data = {"code": 0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            logger.info("?????????????????????????????????????????????%s")
            db.close_connection(conn)
            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("?????????????????????????????????????????????%s" % e)

            return HttpResponse(data)


class ProcessNameCode(View):
    """
    ??????????????????????????????

    """
    def get(self, request):
        try:
            data_list = []
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            sql = """
             SELECT production_code, production_name FROM process_deal
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_dict = {}
                data_dict["production_code"] = dr[0]
                data_dict["production_name"] = dr[1]
                data_list.append(data_dict)

            data = {"code": 0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            logger.info("????????????????????????????????????%s" )
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("????????????????????????????????????%s" % e)

            return HttpResponse(data)


class CurrentUserInfo(View):
    """
    ???????????????????????????????????????work_id, ??????, ?????????
    """
    def get(self, request):
        try:
            dict_data = {}
            product_id = request.session.get("session_projectId")

            work_id = request.session.get("session_workId")
            user_code = request.session.get("session_currentId")
            work_type = request.session.get("session_workType")
            db = DB()

            if product_id:
                conn_num = db.get_connection(product_id)
                sql_num = """
                 SELECT pick_number FROM pick_box where work_id = '{0}'
                """
                sql_num = sql_num.format(work_id)
                drs = db.execute_sql(conn_num, sql_num)

                if drs:
                    pick_number = drs[0][0]
                else:
                    pick_number = 0
                sql = """
                       select * from person where user_code = '{0}'
                       """
                sql_main = sql.format(user_code)
                drs = db.execute_sql(conn_num, sql_main)


                if drs:
                    for dr in drs:
                        dict_data["user_code"] = dr[0]
                        dict_data["user_id"] = dr[1]
                        dict_data["user_name"] = dr[2]
                        dict_data["user_password"] = dr[3]
                        dict_data["user_authority"] = dr[4]
                        dict_data["status"] = dr[5]
                        dict_data["user_product_id"] = dr[6]
                        dict_data["user_role"] = dr[7]
                        dict_data["product_id"] = product_id
                        dict_data["work_type"] = work_type
                        dict_data["pick_number"] = pick_number
                        dict_data["work_id"] = work_id
                    data = {"code": 0, "message": "????????????", "data": dict_data}
                    data = json.dumps(data)
                    logger.info("????????????????????????????????????%s")
                    return HttpResponse(data)
                else:

                    conn_conm = db.get_connection("db_common")
                    sql = """
                    select * from person where user_code = '{0}'
                    """
                    sql = sql.format(user_code)
                    drs = db.execute_sql(conn_conm, sql)
                    for dr in drs:
                        dict_data["user_code"] = dr[0]
                        dict_data["user_id"] = dr[1]
                        dict_data["user_name"] = dr[2]
                        dict_data["user_password"] = dr[3]
                        dict_data["user_authority"] = dr[4]
                        dict_data["status"] = dr[5]
                        dict_data["user_product_id"] = dr[6]
                        dict_data["user_role"] = dr[7]
                        dict_data["product_id"] = product_id
                        dict_data["work_type"] = ""
                        dict_data["pick_number"] = pick_number
                        dict_data["work_id"] = ""
                    data = {"code": 0, "message": "????????????", "data": dict_data}
                    data = json.dumps(data)
                    logger.info("????????????????????????????????????%s")
                    # db.close_connection(conn)
                    return HttpResponse(data)

            else:
                product_id = "db_common"
                conn_num = db.get_connection(product_id)
                sql_num = """
                        SELECT pick_number FROM pick_box where work_id = '{0}'
                       """
                sql_num = sql_num.format(work_id)
                drs = db.execute_sql(conn_num, sql_num)

                if drs:
                    pick_number = drs[0][0]
                else:
                    pick_number = 0
                sql = """
                      select * from person where user_code = '{0}'
                      """
                sql_main = sql.format(user_code)
                drs = db.execute_sql(conn_num, sql_main)
                for dr in drs:
                    dict_data["user_code"] = dr[0]
                    dict_data["user_id"] = dr[1]
                    dict_data["user_name"] = dr[2]
                    dict_data["user_password"] = dr[3]
                    dict_data["user_authority"] = dr[4]
                    dict_data["status"] = dr[5]
                    dict_data["user_product_id"] = dr[6]
                    dict_data["user_role"] = dr[7]
                    dict_data["product_id"] = product_id
                    dict_data["work_type"] = ""
                    dict_data["pick_number"] = pick_number
                    dict_data["work_id"] = ""
                data = {"code": 0, "message": "????????????", "data": dict_data}
                data = json.dumps(data)
                logger.info("????????????????????????????????????%s")
                db.close_connection(conn_num)
                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data":e}
            data = json.dumps(data)
            logger.info("????????????????????????????????????%s" % e)

            return HttpResponse(data)


class ProjectSearchWorkcode(View):
    """
    ???????????????????????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            product_id = request.GET.get("product_id")
            db = DB()
            conn = db.get_connection(product_id)
            sql = """
            select work_id, work_name from work_station
            """
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                data_dict = {}
                data_dict["work_id"] = dr[0]
                data_dict["work_name"] = dr[1]
                data_list.append(data_dict)
            data_list.sort(key=lambda x: (x['work_id']), reverse=False)
            data = {"code": 0, "message": "????????????", "data": data_list}
            data = json.dumps(data)
            logger.info("?????????????????????????????????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("?????????????????????????????????????????????%s" % e)
            return HttpResponse(data)


class GetProductId(View):
    """
    ??????????????????????????????
    """

    def get(self, request):
        try:
            product_id_list = []

            db = DB()
            conn = db.get_connection("db_common")
            sql_product = """
                       select product_id, product_name from productlist
                       """
            dfs = db.execute_sql(conn, sql_product)
            for df in dfs:
                product_id_dict = {}
                product_id_dict["product_id"] = df[0]
                product_id_dict["product_name"] = df[1]
                product_id_list.append(product_id_dict)

            data = {"code": 0, "message": "????????????", "data": product_id_list}
            data = json.dumps(data)
            logger.info("????????????????????????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("????????????????????????????????????%s" % e)

            return HttpResponse(data)


class UnqualifiedProduct(View):
    """
    ?????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            product_id = request.session.get("session_projectId")
            session_work_id = request.session.get("session_workId")
            user_code = request.session.get("session_currentId")
            conn = db.get_connection(product_id)
            product_plan_code = request.GET.get("product_plan_code")
            matter_code = request.GET.get("matter_code")
            solve_method = request.GET.get("solve_method")
            matter_id = request.GET.get("matter_id")
            finished_product_code = request.GET.get("finished_product_code")

            un_pro_table = str("unqualified_product_") + str(session_work_id)

            if finished_product_code:
                finished_product_code_sql = " and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""
            if product_plan_code:
                product_plan_code_sql = " and product_plan_code = " + "'" + product_plan_code + "'"
            else:
                product_plan_code_sql = ""
            if matter_code:
                matter_code_sql = " and matter_code = " + "'" + matter_code + "'"
            else:
                matter_code_sql = ""
            if solve_method:
                solve_method_sql = " and solve_method = " + "'" + solve_method + "'"
            else:
                solve_method_sql = ""
            sql = """
                select * from {0} where 2 > 1 {1} {2} {3} {4} group by finished_product_code
               """
            sql_main = sql.format(un_pro_table, product_plan_code_sql, matter_code_sql,
                                  solve_method_sql, finished_product_code_sql)
            drs = db.execute_sql(conn, sql_main)

            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["unqualified_product_code"] = dr[0]
                    dict_data["product_plan_code"] = dr[1]
                    dict_data["finished_product_code"] = dr[2]
                    dict_data["matter_code"] = []
                    dict_data["matter_id"] = []
                    dict_data["description"] = dr[5]
                    dict_data["leader_work_id"] = dr[6]
                    dict_data["later_work_id"] = dr[7]
                    dict_data["solve_method"] = dr[8]
                    dict_data["solve_result"] = dr[9]
                    data_list.append(dict_data)

                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                sql_num = """
                        select count(*) from {0} where 2 > 1 {1} {2} {3} {4} group by finished_product_code
                        """
                sql_num_format = sql_num.format(un_pro_table, product_plan_code_sql, matter_code_sql,
                                                solve_method_sql, finished_product_code_sql)
                dfs = db.execute_sql(conn, sql_num_format)
                sql_num_int = dfs[0][0]

                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int

                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("???????????????????????????%s")
                db.close_connection(conn)

                return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "?????????", "data": data_add_int}
                result = json.dumps(result)
                logger.error("????????????????????????????????????????????????%s" % result)
                db.close_connection(conn)
                return HttpResponse(result)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("??????????????????????????????%s" % e)
            return HttpResponse(data)


class PutUnqualifiedProduct(View):
    """
    ?????????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            session_work_id = request.session.get("session_workId")
            user_code = request.session.get("session_currentId")
            conn = db.get_connection(product_id)
            unqualified_product_code = request.GET.get("unqualified_product_code")
            product_plan_code = request.GET.get("product_plan_code")
            matter_code = request.GET.get("matter_code")
            solve_method = request.GET.get("solve_method")
            leader_work_id = request.GET.get("leader_work_id")
            later_work_id = request.GET.get("later_work_id")
            solve_result = request.GET.get("solve_result")
            matter_id = request.GET.get("matter_id")
            description = request.GET.get("description")
            un_pro_table = str("unqualified_product_") + str(session_work_id)
            sql_update_unqualified = """
            update {0} set product_plan_code = '{1}',
                       matter_code='{2}',matter_id= '{3}',description='{4}',solve_method='{5}', leader_work_id ='{6}'
                       ,later_work_id = '{7}', solve_result = '{8}' where unqualified_product_code = '{9}'
            """
            sql_update_unqualified_main = sql_update_unqualified.format(un_pro_table, product_plan_code, matter_code,
                                                                        matter_id, description, solve_method,
                                                                        leader_work_id, later_work_id,
                                                                        solve_result, unqualified_product_code)
            drs = db.execute_sql(conn, sql_update_unqualified_main)
            data = {"code": 0, "message": "????????????", "data": []}
            data = json.dumps(data)
            logger.info("??????????????????????????????%s" )
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("???????????????????????????%s"% e)


            return HttpResponse(data)


class DeleteUnqualifiedProduct(View):
    """
    ?????????????????????
    """
    def get(self, request):

        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            session_work_id = request.session.get("session_workId")
            user_code = request.session.get("session_currentId")
            conn = db.get_connection(product_id)
            un_pro_table = str("unqualified_product_") + str(session_work_id)
            unqualified_product_code = request.GET.get("unqualified_product_code")

            sql = """
                      delete from {0} where unqualified_product_code = '{1}'
                      """
            sql_main = sql.format(un_pro_table, unqualified_product_code)
            des = db.execute_sql(conn, sql_main)

            data = {"code": 0, "message": "????????????", "data": ""}
            data = json.dumps(data)
            logger.info("???????????????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("???????????????????????????%s"% e)

            return HttpResponse(data)


class QualifiedMatterCode(View):
    #
    #     ??????????????? ???????????????????????????????????????????????????  ???????????????????????????
    #     ??????????????????????????? ?????????????????????  ?????????????????????
    #
    #     ?????????????????????????????????????????????????????????
    #
    #

    def get_live_data(self, product_id, finished_product_code):
        func = product_live()
        data_list = func.one_product_live(product_id, finished_product_code)

        return data_list


    def get(self, request):
        try:


            product_id = request.session.get("session_projectId")
            session_work_id = request.session.get("session_workId")
            finished_product_code = request.GET.get("finished_product_code")

            db = DB()
            conn = db.get_connection(product_id)
            sql_check = """
                        select * from {0} where work_id = '{1}'
                        """
            sql_check = sql_check.format('work_station', session_work_id)
            drs_checks = db.execute_sql(conn, sql_check)
            work_type = drs_checks[0][4]

            table_se = "product_transit_" + str(session_work_id)

            sql_check = """
                                   select * from {0} where finished_product_code = '{1}'
                                   """
            sql_check = sql_check.format(table_se, finished_product_code)

            drs_checks = db.execute_sql(conn, sql_check)

            if drs_checks:
                data = {"code": 1, "message": "?????????????????????????????????,???????????????"}
                data = json.dumps(data)
                logger.error("?????????????????????????????????,???????????????")
                db.close_connection(conn)
                return HttpResponse(data)
            else:
                da_list = self.get_live_data(product_id, finished_product_code)
                da_list.sort(key=lambda x: (x['time']), reverse=False)

                if da_list:
                    enter_work_code = da_list[-1]["work_code"]
                    test_result = da_list[-1]["test_result"]

                    sql_work_id = """               
                                                                select work_id from work_station where work_code = '{0}'
                                                                """
                    sql_work_id = sql_work_id.format(enter_work_code)

                    des = db.execute_sql(conn, sql_work_id)
                    leader_workid = ''
                    if des[0][0]:
                        leader_workid = des[0][0]

                    sql_work_id = """               
                                            select work_id, leader_work_id from work_station where work_id = '{0}'
                                            """
                    sql_work_id = sql_work_id.format(session_work_id)
                    des = db.execute_sql(conn, sql_work_id)
                    db.close_connection(conn)


                    if des:
                        cureent_leader_id = des[0][1]

                        if cureent_leader_id == None:
                            data = {"code": 0, "message": "?????????????????????????????????"}
                            data = json.dumps(data)
                            logger.info("?????????????????????????????????")
                            return HttpResponse(data)

                        elif cureent_leader_id != leader_workid :
                            data = {"code": 1, "message": "???????????????????????????????????????"}
                            data = json.dumps(data)
                            logger.info("???????????????????????????????????????")
                            return HttpResponse(data)

                        else:
                                if test_result=='PASS':
                                    data = {"code": 0, "message": "????????????", "data": test_result}
                                    data = json.dumps(data)
                                    logger.info("????????????,?????????????????????")
                                    return HttpResponse(data)
                                else:
                                    data = {"code": 1, "message": "???????????????????????????????????????"}
                                    data = json.dumps(data)
                                    logger.info("???????????????????????????????????????")
                                    return HttpResponse(data)


                else:
                    data = {"code": 0, "message": "??????????????????????????????????????????"}
                    data = json.dumps(data)
                    logger.info("?????????????????????????????????")
                    return HttpResponse(data)



        except Exception as e:
            print(e)
            data = {"code": 0, "message": "???????????????????????????????????????", "data": e}
            data = json.dumps(data)
            logger.error("???????????????????????????????????????%s"% e)
            return HttpResponse(data)




class BomMatterCodeName(View):
    """
    ????????????????????? ?????? ?????? ??????
    """
    def get(self, request):
        try:
            data_list = []
            db = DB()
            product_id = request.session.get("session_projectId")
            session_work_id = request.session.get("session_workId")
            conn = db.get_connection(product_id)

            sql = """
                SELECT bom_matter_code, matter_name, rule, code_length FROM matter_list;
                """
            sql = sql.format(session_work_id)
            drs = db.execute_sql(conn, sql)
            if drs:
                for df in drs:
                    matter_code_dict = {}
                    matter_code_dict["matter_code"] = df[0]
                    matter_code_dict["matter_name"] = df[1]
                    matter_code_dict["rule"] = df[2]
                    matter_code_dict["code_length"] = df[3]
                    data_list.append(matter_code_dict)

                data = {"code": 0, "message": "????????????", "data": data_list}
                data = json.dumps(data)
                logger.info("????????????????????? ?????? ?????? ????????????%s")
                db.close_connection(conn)
            else:
                data = {"code": 0, "message": "??????????????????", "data": []}
                data = json.dumps(data)
                logger.info("????????????????????? ?????? ?????? ???????????? ????????????????????????%s")
                db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "??????matter_list???????????????", "data": e}
            data = json.dumps(data)
            logger.error("????????????????????? ?????? ?????? ????????????%s" % data)
            return HttpResponse(data)


class WorkStationGetData(View):
    """
    ??????????????????????????? ????????????
    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            dict_data = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            work_id = request.session.get("session_workId")
            current_person_id = request.session['session_currentId']
            # work_id = request.session['session_workId']
            conn = db.get_connection(product_id)
            # page = int(request.GET.get("page"))
            # page_size = int(request.GET.get("page_size"))
            sql_matter = """
                        select * from process_matter_deal where work_id = '{0}'
                         """
            sql_matter_format = sql_matter.format(work_id)
            matter_dfs = db.execute_sql(conn, sql_matter_format)
            if matter_dfs:
                respose_list = []
                for matter_df in matter_dfs:
                    response_data = {}
                    response_data["process_matter_deal_code"] = matter_df[0]
                    response_data["work_id"] = matter_df[1]
                    response_data["matter_code"] = matter_df[2]
                    response_data["install_number"] = matter_df[3]
                    respose_list.append(response_data)
                dict_data["response_datas"] = respose_list
                data_list.append(dict_data)

            else:
                dict_data["response_datas"] = []
                data_list.append(dict_data)

            data = {"code": 0, "message": "????????????", "data": data_list}

            data = json.dumps(data)
            logger.info("??????????????????????????? ??????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("??????????????????????????? ??????????????????%s"% e)
            return HttpResponse(data)


class FinishedCodeGetMatterCodeID(View):
    """
    ????????????????????? ?????????????????????????????????????????????????????????(???????????????????????????)

    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")

            finished_product_code = request.GET.get("finished_product_code")
            conn = db.get_connection(product_id)

            if finished_product_code:
                finished_product_code_sql = "and finished_product_code = " + "'" + finished_product_code + "'"
                current_person_id = request.session['session_currentId']
                sql_tabale = """
                            show tables like "product_transit_%" 
                            """
                table_list = db.execute_sql(conn, sql_tabale)

                respose_list = []
                for table_type in table_list:
                    table_one = table_type[0]

                    sql_matter = """
                                           select finished_product_code, matter_code, matter_id, 
                                           product_transit_code from {0} where 2 > 1 {1}
                                            """
                    sql_matter_format = sql_matter.format(table_one, finished_product_code_sql)
                    matter_dfs = db.execute_sql(conn, sql_matter_format)
                    if matter_dfs:
                        for matter_df in matter_dfs:
                            response_data = {}
                            response_data["finished_product_code"] = matter_df[0]
                            response_data["matter_code"] = matter_df[1]
                            response_data["matter_id"] = matter_df[2]
                            response_data["work_id"] = table_one[16:]
                            response_data["product_transit_code"] = matter_df[3]
                            respose_list.append(response_data)
                    else:
                        pass
            else:
                respose_list = []
            data = {"code": 0, "message": "????????????", "data": respose_list}
            data = json.dumps(data)
            logger.info("??????????????????????????????????????????????????????%s")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????", "data": e}
            data = json.dumps(data)
            logger.error("??????????????????????????????????????????????????????%s"% e)

            return HttpResponse(data)


class PutFinishedCodeGetMatterCodeID(View):
    """
    ??????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            response_datas = request.GET.get("response_datas")

            if response_datas:
                response_datas = eval(response_datas)
            finished_product_code = request.GET.get("finished_product_code")
            return_work_id = request.GET.get("return_work_id")
            conn = db.get_connection(product_id)
            sql_tabale = """
                      show tables like "product_transit_%" 
                      """
            table_list = db.execute_sql(conn, sql_tabale)

            for table_type in table_list:
                for response_data in response_datas:

                    product_transit_code = response_data.get("product_transit_code")
                    matter_id = response_data.get("matter_id")
                    table_one = table_type[0]
                    sql_matter = """
                                      update {0} set matter_id = '{1}' 
                                      where product_transit_code = '{2}'
                                       """
                    sql_matter_format = sql_matter.format(table_one, matter_id, product_transit_code)
                    matter_dfs = db.execute_sql(conn, sql_matter_format)

            sql_select_table = "unqualified_product_" + str(return_work_id)
            sql_is_create_table = """
               CREATE TABLE IF NOT EXISTS {0}(unqualified_product_code varchar(128) primary key not null,
                              product_plan_code varchar(128), finished_product_code varchar(128), 
                              matter_code varchar(128), matter_id varchar(128),description varchar(128),
                              leader_work_id varchar(128), later_work_id varchar(128), solve_method varchar(128),
                              solve_result varchar(128), order_number int) 
            """
            sql_is_create_table = sql_is_create_table.format(sql_select_table)
            drs = db.execute_sql(conn, sql_is_create_table)

            sql_unqualified_product = """
                                     SELECT max(order_number) FROM {0};
                                     """
            sql_unqualified_product = sql_unqualified_product.format(sql_select_table)
            unqualified_product_nums = db.execute_sql(conn, sql_unqualified_product)
            if unqualified_product_nums[0][0]:
                matter_num = unqualified_product_nums[0][0] + 1
            else:
                matter_num = 1

            unqualified_product_code = "Im_Unqualified_Product_" + str(matter_num)

            sql_disqualification = """
            insert into {0}(unqualified_product_code, finished_product_code, leader_work_id, order_number)
                                              values('{1}', '{2}', '{3}','{4}')
            """
            sql_disqualification = sql_disqualification.format(sql_select_table, unqualified_product_code,
                                                               finished_product_code, return_work_id, matter_num)
            db.execute_sql(conn, sql_disqualification)

            print("-----------------------")

            modify_in_workstation_code = request.GET.get("modify_in_workstation_code")

            if modify_in_workstation_code:
                sql_modify_station = """
                select * from modify_in_workstation where modify_in_workstation_code = '{0}'            
                """
                sql_modify_station = sql_modify_station.format(modify_in_workstation_code)
                drs_modifys = db.execute_sql(conn, sql_modify_station)
                out_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                process_method = request.GET.get("process_method")
                status = request.GET.get("status")
                enter_work_code = request.GET.get("enter_work_code")


                if drs_modifys:
                    sql_update_modify_sta_table = """
                                   update modify_in_workstation set out_time = '{0}', process_method = '{1}', status = '{2}'
                                                     where modify_in_workstation_code = '{3}'                   
                                   """
                    sql_update_modify_sta_table = sql_update_modify_sta_table.format(out_time, process_method,
                                                                                     status, modify_in_workstation_code)
                    db.execute_sql(conn, sql_update_modify_sta_table)
                else:
                    pass
                time.sleep(1)

                if process_method != "????????????":
                    sql_a = """
                    select work_id from work_station where work_code = '{0}'
                    """
                    sql = sql_a.format(enter_work_code)


                    enters_ids = db.execute_sql(conn, sql)
                    if enters_ids:
                        select_table_str = enters_ids[0][0]
                        select_table = "product_transit_" + str(select_table_str)


                        sql_in = """
                                         select max(order_number) from {0}
                                        """
                        sql_in =sql_in.format(select_table)
                        id_num = db.execute_sql(conn, sql_in)
                        if id_num[0][0] == None:
                            id_num = 0
                        else:
                            id_num = id_num[0][0]
                        sql_one = """
                        select * from {0} where finished_product_code = '{1}'
                        """
                        sql_one = sql_one.format(select_table, finished_product_code)
                        drfs = db.execute_sql(conn, sql_one)


                        if drfs:
                            in_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            for drf in drfs:
                                id_num += 1
                                product_transit_code = "Im_Product_Transit_" + str(id_num)
                                # dr_dict = {}

                                # dr_dict["matter_id"] =drf[2]
                                # dr_dict["finished_product_code"] =drf[3]
                                # dr_dict["user_code"] =drf[4]
                                # dr_dict["work_code"] =drf[5]
                                # dr_dict["test_result"] = "??????"
                                # dr_dict["description"] = drf[7]
                                # dr_dict["out_time"] = drf[9]
                                # dr_dict["product_plan_code"] = drf[10]
                                # dr_dict["end_product_code"] = drf[11]
                                # dr_dict["product_code"] = drf[12]

                                sql_insert = """
                                insert into {0}(product_transit_code, matter_code, matter_id, finished_product_code,user_code,
                                work_code,test_result, description, out_time, product_plan_code, end_product_code,product_code,
                                order_number)
                                values('{1}', '{2}', '{3}', '{4}', '{5}', '{6}', '{7}', 
                                '{8}', '{9}', '{10}', '{11}', '{12}', '{13}')                            
                                """
                                sql_insert = sql_insert.format(select_table, product_transit_code,
                                                               drf[1],drf[2], drf[3], drf[4], drf[5], "PASS",
                                                               drf[7], in_time, drf[10], drf[11], drf[12], id_num)
                                db.execute_sql(conn, sql_insert)

                        else:
                            pass

                    else:
                        data = {"code": 1, "message": "??????????????????????????????????????????????????????"}
                        data = json.dumps(data)
                        logger.info("??????????????????????????????????????????????????????")
                        return HttpResponse(data)


            data = {"code": 0, "message": "????????????"}
            data = json.dumps(data)
            logger.info("????????????????????????")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 0, "message": e }

            data = json.dumps(data)
            logger.error("????????????????????????")
            return HttpResponse(data)


class DeleteFinishedCodeGetMatterCodeID(View):
    """
    ???????????????????????????
    """
    def get(self, request):
        try:
            db = DB()

            product_id = request.session.get("session_projectId")

            conn = db.get_connection(product_id)

            finished_product_code = request.GET.get("finished_product_code")
            return_work_id = request.GET.get("work_id")

            finished_product_code_sql = "and finished_product_code = " + "'" + finished_product_code + "'"

            sql_tabale = """
                       show tables like "product_transit_%" 
                       """
            table_list = db.execute_sql(conn, sql_tabale)

            for table_type in table_list:
                table_one = table_type[0]
                sql_key = """
                SET SQL_SAFE_UPDATES = 0
                """
                db.execute_sql(conn, sql_key)
                sql_matter = """
                           delete from {0} where 2 > 1 {1}
                            """
                sql_matter_format = sql_matter.format(table_one, finished_product_code_sql)
                matter_dfs = db.execute_sql(conn, sql_matter_format)

            sql_delete = """
            SELECT table_name FROM work_station left join work_transit
             on work_station.work_code = work_transit.work_code
            """
            dr_tables = db.execute_sql(conn, sql_delete)
            print("-----===>", dr_tables)
            if dr_tables[0][0]:
                for dr_tbale in dr_tables:
                    if dr_tbale[0]:
                        select_table = dr_tbale[0]
                        sql_key = """
                        SET SQL_SAFE_UPDATES = 0
                        """
                        db.execute_sql(conn, sql_key)
                        sql_delet = """
                        delete from {0} where 2 > 1 {1}
                        """
                        sql_delet = sql_delet.format(select_table, finished_product_code_sql)
                        db.execute_sql(conn, sql_delet)

            else:
                pass

            sql_modify_check = """
            select * FROM modify_in_workstation where 2 > 1 {0}
            """
            sql_modify_check = sql_modify_check.format(finished_product_code_sql)

            drfs = db.execute_sql(conn, sql_modify_check)
            print("===>drfs", drfs)
            if drfs[0][0]:
                sql_key = """
                SET SQL_SAFE_UPDATES = 0
                """
                db.execute_sql(conn, sql_key)
                sql_dele_modify = """
                delete from modify_in_workstation where 2 > 1 {0}
                """
                sql_dele_modify = sql_dele_modify.format(finished_product_code_sql)
                db.execute_sql(conn, sql_dele_modify)
            else:
                pass

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("?????????????????????????????????????????????%s")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 0, "message": e}

            data = json.dumps(data)
            logger.error("?????????????????????????????????????????????%s"% e)
            return HttpResponse(data)


class Pick(View):
    """
    ????????????????????????

    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            work_id = request.GET.get("work_id")

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            if work_id:
                work_id_sql = "and work_id = " + "'" + work_id + "'"
            else:
                work_id_sql = ""

            sql = """
                        select * from pick_box where 2 > 1 {0}
                       """
            sql_main = sql.format(work_id_sql)

            drs = db.execute_sql(conn, sql_main)

            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["pick_code"] = dr[0]
                    dict_data["work_id"] = dr[1]
                    dict_data["pick_number"] = dr[2]
                    dict_data["description"] = dr[3]
                    data_list.append(dict_data)

                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()

                sql_num = """
                            select count(*) from pick_box where 2 > 1 {0}
                            """
                sql_num_main = sql_num.format(work_id_sql)

                dfs = db.execute_sql(conn, sql_num_main)
                if dfs:
                    sql_num_int = dfs[0][0]
                else:
                    sql_num_int = 0
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("????????????????????????%s")
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("????????????????????????%s"% e)
            return HttpResponse(result)

    def post(self, request):
        """
        ???????????????
        :param request:
        :return:
        """
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            json_data = request.body
            str_data = json.loads(json_data)
            work_id = str_data.get("work_id")
            pick_number = str_data.get("pick_number")
            description = str_data.get("description")

            sql = """
                 select max(order_number) from pick_box
                """
            id_num = db.execute_sql(conn, sql)
            if id_num[0][0] == None:
                id_num = 1
            else:
                id_num = id_num[0][0] + 1

            pick_code = str(product_id) + "pickbox" + str(id_num)

            sql_insert = """
                        insert into pick_box(pick_code, work_id, pick_number, description,order_number)
                        values('{0}', '{1}', '{2}', '{3}', '{4}')
                        """

            sql_insert_format = sql_insert.format(pick_code, work_id, pick_number,
                                                  description, id_num)

            drs = db.execute_sql(conn, sql_insert_format)
            data = {"code": 0, "message": "?????????????????????"}

            data = json.dumps(data)
            logger.info("??????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "?????????????????????"}

            data = json.dumps(data)
            logger.error("??????????????????")

            return HttpResponse(data)


class PutPick(View):
    """
    ?????????????????????

    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            pick_code = request.GET.get("pick_code")
            work_id = request.GET.get("work_id")
            pick_number = request.GET.get("pick_number")
            description = request.GET.get("description")

            sql_update = """
                       update pick_box set work_id='{0}',
                        pick_number='{1}',description='{2}' where pick_code = '{3}'
                       """

            sql_insert_format = sql_update.format(work_id, pick_number, description,
                                                  pick_code)

            drs = db.execute_sql(conn, sql_insert_format)
            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("???????????????????????????")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

            data = json.dumps(data)
            logger.error("???????????????????????????")

            return HttpResponse(data)


class DeletePick(View):
    """
    ????????????

    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            pick_code = request.GET.get("pick_code")
            sql_delete = """
                       delete from pick_box where pick_code = '{0}'
                       """
            sql_insert_format = sql_delete.format(pick_code)

            drs = db.execute_sql(conn, sql_insert_format)
            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("??????????????????")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

            data = json.dumps(data)
            logger.error("??????????????????")
            return HttpResponse(data)


class FinishedProductStorage(View):
    """
    ??????????????????????????????????????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            enter_time = request.GET.get("enter_time")
            enter_user = request.GET.get("enter_user")
            product_plan_code = request.GET.get("product_plan_code")

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            start_time = request.GET.get("start_time")
            end_time = request.GET.get("end_time")

            conn = db.get_connection(product_id)

            if start_time:
                start_time_sql = "and enter_time >= " + "'" + start_time + "'"
                num_start_time_sql = " and b.enter_time >=" + "'" + start_time + "'"
            else:
                start_time_sql = ""
                num_start_time_sql = ""

            if end_time:
                end_time_sql = "and enter_time < " + "'" + end_time + "'"
                num_end_time_sql = " and b.enter_time < " + "'" + end_time + "'"
            else:
                end_time_sql = ""
                num_end_time_sql = ""
            if enter_user:
                enter_user_sql = "and enter_user = " + "'" + enter_user + "'"
            else:
                enter_user_sql = ""

            if product_plan_code:
                product_plan_code_sql = " and product_plan_code = " + "'" + product_plan_code + "'"
            else:
                product_plan_code_sql = ""
            sql = """
                    select * from enter_storage_status where 2 > 1 {0} {1} {2} {3}
                   """
            sql = sql.format(start_time_sql, end_time_sql, enter_user_sql, product_plan_code_sql)
            drs = db.execute_sql(conn, sql)

            sql_num = """           
            SELECT count(*) FROM enter_storage as a left join enter_storage_status as b on a.pack_id = b.pack_id 
            where 2 > 1 {0} {1} {2}
            """
            sql_num = sql_num.format(num_start_time_sql, num_end_time_sql, product_plan_code_sql)
            product_nums = db.execute_sql(conn, sql_num)
            print("==>", product_nums)
            print("==>", len(product_nums))
            if product_nums:
                product_num = product_nums[0][0]
            else:
                product_num = 0

            if drs:
                for dr in drs:
                    dict_data = {}
                    pick_id = dr[0]
                    dict_data["pack_id"] = dr[0]
                    dict_data["product_id"] = dr[1]
                    dict_data["product_name"] = dr[2]
                    dict_data["enter_user"] = dr[3]
                    s = dr[4]
                    s = s.strftime("%Y-%m-%d %H:%M:%S ")
                    dict_data["enter_time"] = s
                    dict_data["status"] = dr[5]
                    dict_data["product_plan_code"] = dr[6]

                    sql_status = """
                    select * from enter_storage where pack_id = '{0}'
                    """
                    sql_status = sql_status.format(pick_id)
                    dfs = db.execute_sql(conn, sql_status)

                    response_datas = []

                    for df in dfs:
                        datas_dict = {}
                        datas_dict["enter_storage_code"] = df[0]
                        datas_dict["pack_id"] = df[1]
                        datas_dict["finished_product_code"] = df[2]
                        response_datas.append(datas_dict)
                        dict_data["response_datas"] = response_datas
                    data_list.append(dict_data)
                data_list.sort(key=lambda x: (x['enter_time']), reverse=True)

                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()

                dfs = int(len(data_list))
                if dfs:
                    sql_num_int = dfs
                else:
                    sql_num_int = 0
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
                data_add_int["product_num"] = product_num

                result = {"code": 0, "message": "????????????", "data": data_add_int}

                result = json.dumps(result)
                logger.info("??????????????????????????????????????????%s")
                db.close_connection(conn)
                return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                data_add_int["product_num"] = 0

                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                db.close_connection(conn)

                result = json.dumps(result)
                return HttpResponse(result)
        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("??????????????????????????????????????????%s"% e)
            return HttpResponse(result)

    def post(self, request):
        """
        ????????????????????????
        :param request:
        :return:
        """
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            work_id = request.session.get("session_workId")
            enter_user = request.session.get("session_currentId")
            json_data = request.body
            str_data = json.loads(json_data)
            pack_id = str_data.get("pack_id")
            product_plan_code = str_data.get("product_plan_code")
            # product_name = str_data.get("product_name")
            status = "PASS"
            response_datas = str_data.get("response_datas")
            conn_product_name = db.get_connection("db_common")
            sql_product_name = """
            SELECT product_name FROM productlist where product_id = '{0}'
            """
            sql_product_name = sql_product_name.format(product_id)
            drs = db.execute_sql(conn_product_name, sql_product_name)
            if drs:
                product_name = drs[0][0]
            else:
                product_name = ""

            conn = db.get_connection(product_id)
            sql1 = """
            CREATE TABLE IF NOT EXISTS enter_storage(enter_storage_code varchar(128) primary key not null, 
                              pack_id varchar(128),
                              finished_product_code varchar(128),
                              order_number int) 
            """
            sql2 = """
            CREATE TABLE IF NOT EXISTS enter_storage_status(pack_id varchar(128) primary key not null,
                              product_id varchar(128),product_name varchar(128),
                              enter_user varchar(128), enter_time datetime, 
                              status varchar(128),product_plan_code varchar(128), order_number int)           
            """
            drs = db.execute_sql(conn, sql1)
            dfs = db.execute_sql(conn, sql2)

            sql_num1 = """
                             select max(order_number) from enter_storage
                            """
            id_num = db.execute_sql(conn, sql_num1)
            if id_num[0][0] == None:
                id_num = 0
            else:
                id_num = id_num[0][0]

            sql_num2 = """
                         select max(order_number) from enter_storage_status
                        """
            id_num2 = db.execute_sql(conn, sql_num2)
            if id_num2[0][0] == None:
                id_num2 = 1
            else:
                id_num2 = id_num2[0][0] + 1
            for response_data in response_datas:
                id_num += 1
                enter_storage_code = str("Im_enter_storage_") + str(id_num)
                pack_id = response_data.get("pack_id")
                finished_product_code = response_data.get("finished_product_code")
                # product_name = response_data.get("product_name")
                # enter_user = response_data.get("enter_user")

                sql_insert = """
                insert into enter_storage(enter_storage_code, pack_id, finished_product_code, order_number)
                values('{0}', '{1}','{2}','{3}')
                """
                sql_insert = sql_insert.format(enter_storage_code, pack_id, finished_product_code,
                                               id_num)

                drs = db.execute_sql(conn, sql_insert)
            enter_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            insert_sql = """
            insert into enter_storage_status(pack_id, product_id, 
            product_name, enter_user, enter_time, status, product_plan_code, order_number) 
            values('{0}', '{1}','{2}', '{3}', '{4}', '{5}', '{6}', '{7}')
            """
            insert_sql = insert_sql.format(pack_id, product_id, product_name, enter_user,
                                           enter_time, status, product_plan_code, id_num2)
            dfs = db.execute_sql(conn, insert_sql)

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("????????????i??????")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

            data = json.dumps(data)
            logger.error("??????????????????")


            return HttpResponse(data)


class PutFinishedProductStorage(View):
    """
    ??????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            work_id = request.session.get("session_workId")
            enter_user = request.session.get("session_currentId")

            pack_id = request.GET.get("pack_id")
            # product_id = request.GET.get("product_id")
            product_name = request.GET.get("product_name")
            enter_user = request.GET.get("enter_user")
            enter_time = request.GET.get("enter_time")
            status = request.GET.get("status")
            product_plan_code = request.GET.get("product_plan_code")
            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)

            conn = db.get_connection(product_id)
            for response_data in response_datas:
                enter_storage_code = response_data.get("enter_storage_code")
                pack_id = response_data.get("pack_id")
                finished_product_code = response_data.get("finished_product_code")
                sql_update = """
                 update enter_storage set pack_id='{0}',
                        finished_product_code='{1}' where enter_storage_code = '{2}' 
                """
                sql_update = sql_update.format(pack_id, finished_product_code, enter_storage_code)
                db.execute_sql(conn, sql_update)
            sql_update_status = """
            update enter_storage_status set 
                        product_id='{0}', product_name = '{1}', enter_user = '{2}',enter_time = '{3}',
                        status = '{4}' , product_plan_code = '{5}' 
                        where pack_id = '{6}' 
            """
            sql_update_status = sql_update_status.format(product_id, product_name,
                                                         enter_user, enter_time, status, product_plan_code, pack_id)
            db.execute_sql(conn, sql_update_status)

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("????????????????????????")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

            data = json.dumps(data)
            logger.error("????????????????????????")
            return HttpResponse(data)


class DeleteFinishedProductStorage(View):
    """
    ??????????????????????????????????????????

    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            work_id = request.session.get("session_workId")
            enter_user = request.session.get("session_currentId")
            pack_id = request.GET.get("pack_id")
            # response_datas = request.GET.get("response_datas")
            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            else:
                response_datas = response_datas
            # product_id = "pen"
            # pack_id = "pen_20210803_1"
            # response_datas = []
            conn = db.get_connection(product_id)
            if response_datas:
                for response_data in response_datas:
                    storage_id = response_data.get("enter_storage_code")
                    sql_delete = """
                    delete FROM enter_storage where enter_storage_code = "{0}"
                    """
                    sql_delete = sql_delete.format(storage_id)
                    db.execute_sql(conn, sql_delete)
            else:
                sql_delete_status = """
                delete FROM enter_storage_status where pack_id = "{0}"
                """
                sql_delete_status = sql_delete_status.format(pack_id)
                db.execute_sql(conn, sql_delete_status)
                sql_storage = """
                SELECT enter_storage_code FROM enter_storage where pack_id = "{0}";
                """
                sql_storage = sql_storage.format(pack_id)
                drs = db.execute_sql(conn, sql_storage)
                for dr in drs:
                    # enter_list.append(dr[0])
                    enter_id = dr[0]
                    sql_delete = """
                    delete FROM enter_storage where enter_storage_code = "{0}"
                    """
                    sql_delete = sql_delete.format(enter_id)
                    db.execute_sql(conn, sql_delete)

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("????????????????????????????????????????????????%s")
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

            data = json.dumps(data)
            logger.error("????????????????????????????????????????????????%s"% e)

            return HttpResponse(data)


class stockInStockOut(View):

    """
    ???????????????(????????????????????????????????? ?????????????????????)

    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            work_id = request.session.get("session_workId")
            enter_user = request.session.get("session_currentId")
            enter_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ")

            # product_id = "pen"
            # status = "??????"
            # response_datas = [{"pack_id":"pen_20210803_5"}, {"pack_id":"pen_20210803_6"}]
            conn = db.get_connection(product_id)
            status = request.GET.get("status")
            response_datas = request.GET.get("response_datas")

            if response_datas:
                response_datas = eval(response_datas)
                for response_data in response_datas:
                    pack_id = response_data.get("pack_id")
                    sql_update = """
                    update enter_storage_status set 
                        status ='{0}', enter_time = '{1}' 
                        where pack_id = '{2}' 
                    """
                    sql_update = sql_update.format(status, enter_time, pack_id)
                    db.execute_sql(conn, sql_update)
                    data = {"code": 0, "message": "????????????"}
                    data = json.dumps(data)
                    logger.info("?????????????????????")
                    db.close_connection(conn)
                    return HttpResponse(data)
            else:
                data = {"code": 1, "message": "response_datas??????????????????"}
                data = json.dumps(data)
                db.close_connection(conn)
                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}

            data = json.dumps(data)
            logger.error("???????????????")

            return HttpResponse(data)


class BackMatter(View):
    """
    ??????????????????
    """
    def get(self, request):
        try:
            data_list = []
            matter_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            back_person = request.GET.get("back_person")
            product_plan_code = request.GET.get("product_plan_code")

            if back_person:
                back_person_sql = "and back_person = " + "'" + back_person + "'"
            else:
                back_person_sql = ""
            if product_plan_code:
                product_plan_code_sql = "and product_plan_code = " + "'" + product_plan_code + "'"
            else:
                product_plan_code_sql = ""

            sql = """
            select * from product_back_matter where 2 > 1 {0} {1} 
            """
            sql = sql.format(back_person_sql, product_plan_code_sql)
            drs = db.execute_sql(conn, sql)
            if len(drs) > 0 :
                if drs[0][0]:
                    for dr in drs:
                        dict_data = {}
                        dict_data["materials_back_code"] = dr[0]
                        dict_data["back_person"] = dr[1]
                        dict_data["product_plan_code"] = dr[2]
                        s = dr[3]
                        if s:
                            s = s.strftime("%Y-%m-%d %H:%M:%S ")
                        else:
                            s = ""
                        dict_data["back_time"] = s
                        dict_data["description"] = dr[4]
                        sql_back = """
                        select * from back_matter where materials_back_code = '{0}'
                        """
                        sql_back = sql_back.format(dr[0])
                        back_dfs = db.execute_sql(conn, sql_back)
                        if len(back_dfs) > 0:
                            if back_dfs[0][0]:
                                respose_list = []
                                for matter_df in back_dfs:
                                    response_data = {}
                                    response_data["deal_back_code"] = matter_df[0]
                                    response_data["materials_back_code"] = matter_df[1]
                                    response_data["matter_code"] = matter_df[2]
                                    response_data["matter_count"] = matter_df[3]
                                    respose_list.append(response_data)
                                dict_data["response_datas"] = respose_list
                                data_list.append(dict_data)
                            else:
                                dict_data["response_datas"] = []
                                data_list.append(dict_data)
                        else:
                            dict_data["response_datas"] = []
                            data_list.append(dict_data)
                    page_result = Page(page, page_size, data_list)
                    data = page_result.get_str_json()

                    sql_num_int = int(len(data_list))

                    data_add_int["data"] = data
                    data_add_int["total"] = sql_num_int
                    result = {"code": 0, "message":"????????????", "data": data_add_int}
                    result = json.dumps(result)
                    logger.info("????????????????????????%s"% result)
                    db.close_connection(conn)
                    return HttpResponse(result)
                else:
                    data_add_int["data"] = []
                    data_add_int["total"] = 0
                    result = {"code": 0, "message": "????????????,??????????????????", "data": data_add_int}
                    result = json.dumps(result)
                    db.close_connection(conn)
                    return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "????????????,??????????????????", "data": data_add_int}
                result = json.dumps(result)
                db.close_connection(conn)
                return HttpResponse(result)

        except Exception as e:

            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("????????????????????????%s"% e)

            return HttpResponse(result)

    def post(self, request):
        """
        ??????????????????
        :param request:
        :return:
        """
        try:

            json_data = request.body
            str_data = json.loads(json_data)

            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            sql_matter = """
                   SELECT max(order_number) FROM back_matter;
                   """
            matter_type = db.execute_sql(conn, sql_matter)
            if matter_type[0][0] == None:
                matter_num = 0
            else:
                matter_num = matter_type[0][0]

            sql = """
                  SELECT max(order_number) FROM product_back_matter;
                  """
            product_type = db.execute_sql(conn, sql)
            if product_type[0][0] == None:
                id_num = 1
            else:
                id_num = product_type[0][0] + 1

            materials_back_code = str("Im_Back_Matter_" + str(id_num))
            back_person = str_data.get("back_person")
            product_plan_code = str_data.get("product_plan_code")
            back_time = str_data.get("back_time")
            description = str_data.get("description")
            response_datas = str_data.get("response_datas")

            order_number = id_num
            sql_insert = """
                               insert into product_back_matter(materials_back_code, back_person,
                               product_plan_code,
                               back_time, description, order_number)
                               values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                               """

            sql_insert_format = sql_insert.format(materials_back_code, back_person,
                                                  product_plan_code, back_time,
                                                  description, order_number)

            drs = db.execute_sql(conn, sql_insert_format)
            if len(response_datas) > 0:
                for response_data in response_datas:
                    # db = DB()
                    # conn = db.get_connection("db_common")
                    matter_code = response_data.get("matter_code")
                    matter_count = response_data.get("matter_count")
                    matter_num = matter_num + 1
                    deal_back_code = str("Im_Deal_Back_" + str(matter_num))
                    sql_response = """
                                   insert into back_matter(deal_back_code, materials_back_code,matter_code, matter_count,
                                   order_number)
                                   values('{0}', '{1}', '{2}', '{3}', '{4}')
                           """
                    sql_response_format = sql_response.format(deal_back_code, materials_back_code,
                                                              matter_code, matter_count,
                                                              matter_num)
                    matter_drs = db.execute_sql(conn, sql_response_format)

                    sql_search_num = """
                                   select matter_count from person_matter
                                   """
                    sql_search_nums = db.execute_sql(conn, sql_search_num)
                    if len(sql_search_nums) > 0:
                        if sql_search_nums[0][0]:
                            sql_num = sql_search_nums[0][0]
                            person_matter_count = int(sql_num) + int(matter_count)
                            sql_update = """
                                           update person_matter set matter_count = '{0}' where matter_code = '{1}'
                                           """
                            sql_update = sql_update.format(person_matter_count, matter_code)
                            db.execute_sql(conn, sql_update)

                data = {"code": 0, "message": "??????????????????,??????????????????????????????????????????"}
                logger.info("BackMatter??????????????????")
                data = json.dumps(data)
                db.close_connection(conn)
                return HttpResponse(data)
            else:
                data = {"code": 0, "message": "????????????"}
                logger.info("BackMatter??????????????????")
                data = json.dumps(data)
                db.close_connection(conn)
                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????%s"%e}
            logger.error("BackMatter??????????????????")
            data = json.dumps(data)
            return HttpResponse(data)


class PutBackMatter(View):
    """
    ??????????????????

    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            materials_back_code = request.GET.get("materials_back_code")
            back_person = request.GET.get("back_person")
            product_plan_code = request.GET.get("product_plan_code")
            back_time = request.GET.get("back_time")
            description = request.GET.get("description")

            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            else:
                response_datas = response_datas

            sql = """
                      update product_back_matter set 
                      back_person='{0}',product_plan_code='{1}', back_time ='{2}',
                      description= '{3}' where materials_back_code = '{4}'
                      """
            sql_main = sql.format(back_person, product_plan_code, back_time, description,
                                  materials_back_code)
            drs = db.execute_sql(conn, sql_main)

            sql_matter_num = """
             SELECT max(order_number) FROM back_matter
            """
            dras = db.execute_sql(conn, sql_matter_num)

            if len(dras)>0:
                if dras[0][0]:
                    sql_num = dras[0][0]
                else:
                    sql_num = 0
            else:
                sql_num = 0
            if len(response_datas)>0:
                for response_data in response_datas:
                    # materials_production_code = response_data.get("materials_production_code")
                    deal_back_code = response_data.get("deal_back_code")
                    # materials_production_code = response_data.get("materials_production_code")
                    matter_code = response_data.get("matter_code")
                    matter_count = response_data.get("matter_count")
                    if deal_back_code:
                        sql = """
                           update back_matter set materials_back_code = '{0}', 
                           matter_code = '{1}', matter_count ='{2}' where deal_back_code = '{3}'
                        """
                        sql = sql.format(materials_back_code, matter_code, matter_count, deal_back_code)
                        db.execute_sql(conn, sql)


                        sql_search_num = """
                                       select matter_count from person_matter
                                       """
                        sql_search_nums = db.execute_sql(conn, sql_search_num)
                        if len(sql_search_nums) > 0:
                            if sql_search_nums[0][0]:
                                sql_num = sql_search_nums[0][0]
                                person_matter_count = int(sql_num) + int(matter_count)
                                sql_update = """
                                                                   update person_matter set matter_count = '{0}' where matter_code = '{1}'
                                                                   """
                                sql_update = sql_update.format(person_matter_count, matter_code)
                                db.execute_sql(conn, sql_update)

                    else:
                        sql_num = sql_num + 1

                        deal_back_code = str("Im_Deal_Back_" + str(sql_num))

                        sql_response = """
                                          insert into back_matter(deal_back_code,materials_back_code, 
                                          matter_code, matter_count, order_number) values('{0}',
                                           '{1}', '{2}', '{3}', '{4}')
                                      """
                        sql_response_format = sql_response.format(deal_back_code, materials_back_code,
                                                                  matter_code, matter_count, sql_num)
                        db.execute_sql(conn, sql_response_format)

                        sql_search_num = """
                                       select matter_count from person_matter
                                       """
                        sql_search_nums = db.execute_sql(conn, sql_search_num)
                        if len(sql_search_nums) > 0:
                            if sql_search_nums[0][0]:
                                sql_num = sql_search_nums[0][0]
                                person_matter_count = int(sql_num) + int(matter_count)
                                sql_update = """
                                           update person_matter set matter_count = '{0}' where matter_code = '{1}'
                                           """
                                sql_update = sql_update.format(person_matter_count, matter_code)
                                db.execute_sql(conn, sql_update)

                data = {"code": 0, "message": "????????????"}

                data = json.dumps(data)
                logger.info("????????????????????????%s")

                db.close_connection(conn)

                return HttpResponse(data)
            else:
                data = {"code": 0, "message": "????????????"}

                data = json.dumps(data)
                logger.info("????????????????????????%s")

                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            logger.error("????????????????????????---->%s"% e)
            data = json.dumps(data)
            return HttpResponse(data)


class DeleteBackMatter(View):
    """
    ??????????????????

    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            materials_back_code = request.GET.get("materials_back_code")

            response_datas = request.GET.get("response_datas")
            response_datas = eval(response_datas)
            # materials_production_code = "Im_Product_Pick_Matter_1"
            # response_datas = [{"materials_code": "Im_Materials_Pick_1"}]
            if len(response_datas) != 0:

                for response_data in response_datas:
                    deal_back_code = dict(response_data).get("deal_back_code")
                    sql = """
                       delete from back_matter where deal_back_code = '{0}'
                       """
                    sql_format = sql.format(deal_back_code)
                    db.execute_sql(conn, sql_format)
            else:
                sql_matter = """
                              SELECT * FROM back_matter where materials_back_code = '{0}'
                              """
                sql_matter_format = sql_matter.format(materials_back_code)
                matter_dfs = db.execute_sql(conn, sql_matter_format)
                if len(matter_dfs) > 0:
                    if matter_dfs[0][0]:
                        for matter_df in matter_dfs:
                            deal_back_code = matter_df[0]
                            sql_delete = """
                               delete FROM back_matter where deal_back_code = '{0}'
                               """
                            sql_delete_format = sql_delete.format(deal_back_code)
                            db.execute_sql(conn, sql_delete_format)
                        sql = """
                               delete from product_back_matter where materials_back_code = '{0}'
                               """
                        sql_main = sql.format(materials_back_code)
                        des = db.execute_sql(conn, sql_main)
            data = {"code": 0, "message": "????????????"}
            logger.info("???????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("???????????????????????????,%s" % e)
            return HttpResponse(data)


class ModifyFinishProductStatus(View):
    """

    ??????????????????????????????PASS/FAIL???
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            work_type = request.session.get("session_workType")
            work_id = request.session.get("session_workId")

            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)

            for response_data in response_datas:
                finished_product_code = response_data.get("finished_product_code")
                status = response_data.get("status")
                conn = db.get_connection(product_id)
                table_select = "check_" + str(work_id)

                sql_transit = """
                             SELECT max(order_number) FROM {0};
                             """
                sql_transit = sql_transit.format(table_select)
                unqualified_product_nums = db.execute_sql(conn, sql_transit)
                if unqualified_product_nums[0][0]:
                    matter_num = unqualified_product_nums[0][0] + 1
                else:
                    matter_num = 1

                product_transit_code = "Im_Product_Transit_" + str(matter_num)

                sql_insert = """
                            insert into {0}(product_transit_code, finished_product_code, test_result, order_number)
                            values('{1}', '{2}', '{3}', '{4}')
                            """
                sql_insert = sql_insert.format(table_select, product_transit_code,
                                               finished_product_code, status, matter_num)
                db.execute_sql(conn, sql_insert)

            conn = db.get_connection(product_id)
            sql_tabale = """
                        show tables like "product_transit_%" 
                        """
            table_list = db.execute_sql(conn, sql_tabale)
            for table_type in table_list:
                table_one = table_type[0]

                for response_data in response_datas:
                    finished_product_code = response_data.get("finished_product_code")
                    status = response_data.get("status")
                    finished_product_code_sql = "and finished_product_code = " + "'" + finished_product_code + "'"

                    sql_matter = """
                                  select product_transit_code, finished_product_code, test_result from {0} where 2 > 1 {1}
                                   """
                    sql_matter_format = sql_matter.format(table_one, finished_product_code_sql)
                    matter_dfs = db.execute_sql(conn, sql_matter_format)
                    if matter_dfs:
                        transit_list = []
                        for matter_df in matter_dfs:
                            sql_update = """
                                       update {0} set test_result='{1}' where product_transit_code = '{2}'
                                       """
                            sql_update = sql_update.format(table_one, status, matter_df[0])
                            db.execute_sql(conn, sql_update)
                        if status == "FAIL":
                            sql_un_table = str("unqualified_product_") + str(table_one[16:])
                            sql = """
                            CREATE TABLE IF NOT EXISTS {0}(unqualified_product_code varchar(128) primary key not null, 
                                          product_plan_code varchar(128),
                                          finished_product_code varchar(128),
                                          matter_code varchar(128),
                                          matter_id varchar(128),
                                          description varchar(128),
                                          leader_work_id varchar(128),
                                          later_work_id varchar(128),
                                          solve_method varchar(128),
                                          solve_result varchar(128),                             
                                          order_number int) 
                            """
                            sql = sql.format(sql_un_table)
                            db.execute_sql(conn, sql)

                            sql_unqualified_product = """
                                                     SELECT max(order_number) FROM {0};
                                                     """
                            sql_unqualified_product = sql_unqualified_product.format(sql_un_table)
                            unqualified_product_nums = db.execute_sql(conn, sql_unqualified_product)
                            if unqualified_product_nums[0][0]:
                                matter_num = unqualified_product_nums[0][0] + 1
                            else:
                                matter_num = 1

                            unqualified_product_code = "Im_Unqualified_Product_" + str(matter_num)

                            sql_insert = """
                            insert into {0}(unqualified_product_code, finished_product_code, description, order_number)
                            values('{1}', '{2}', '{3}', '{4}')
                            """

                            sql_insert = sql_insert.format(sql_un_table,
                                                           unqualified_product_code,
                                                           finished_product_code, "????????????", matter_num)
                            db.execute_sql(conn, sql_insert)
                        else:
                            sql_un_status = "unqualified_product_" + str(table_one[16:])
                            sql_code = """
                            SELECT unqualified_product_code FROM {0} where finished_product_code = '{1}'
                            """
                            sql_code = sql_code.format(sql_un_status, finished_product_code)
                            drs = db.execute_sql(conn, sql_code)
                            if drs:
                                for dr in drs:
                                    sql_delete = """
                                    delete from {0} where unqualified_product_code='{1}'
                                    """
                                    sql_delete = sql_delete.format(sql_un_status, dr[0])
                                    db.execute_sql(conn, sql_delete)
                            else:
                                pass
                    else:
                        pass

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("?????????????????????/?????????????????? ??????")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????,%s" % e)
            return HttpResponse(data)


class MatterSearch(View):
    """

    ????????????
    """
    def get(self, request):
        try:
            data_add_int = {}
            db = DB()


            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            finished_product_code = request.GET.get("finished_product_code")
            matter_id = request.GET.get("matter_id")
            work_id = request.GET.get("work_id")
            user_code = request.GET.get("user_code")
            out_start_time = request.GET.get("out_start_time")
            out_end_time = request.GET.get("out_end_time")


            sql_table = """
                        show tables like "product_transit_%" 
                        """
            table_list = db.execute_sql(conn, sql_table)
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            if finished_product_code:
                finished_product_code_sql = "and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""
            if matter_id:
                matter_id_sql = "and matter_id = " + "'" + matter_id + "'"
            else:
                matter_id_sql = ""
            if work_id:
                work_id_sql = "and work_id = " + "'" + work_id + "'"
            else:
                work_id_sql = ""
            if user_code:
                user_code_sql = "and user_code = " + "'" + user_code + "'"
            else:
                user_code_sql = ""
            if out_start_time:
                out_start_time_sql = "and out_time >= " + "'" + out_start_time + "'"
            else:
                out_start_time_sql = ""
            if out_end_time:
                out_end_time_sql = "and out_time <=" + "'" + out_end_time + "'"
            else:
                out_end_time_sql = ""
            data_list = []
            for table_type in table_list:
                table_one = table_type[0]

                sql = """
                SELECT b.matter_name,a.matter_id,b.rule,
                b.matter_category,c.product_time,
                d.work_name,a.user_code,a.out_time,
                a.finished_product_code, a.product_plan_code FROM {0} as a left join 
                matter_list as b on a.matter_code = b.bom_matter_code 
                left join person_matter as c on c.matter_code = a.matter_code 
                left join work_station as d on d.work_code = a.work_code where 2 > 1 {1} {2} {3} {4} {5} {6}
                order by a.out_time desc;
                """
                sql = sql.format(table_one, finished_product_code_sql, matter_id_sql,
                                 work_id_sql, user_code_sql, out_start_time_sql, out_end_time_sql)
                drs = db.execute_sql(conn, sql)
                if drs:
                    for dr in drs:
                        dict_data = {}
                        dict_data["matter_name"] = dr[0]
                        dict_data["matter_id"] = dr[1]
                        dict_data["rule"] = dr[2]
                        dict_data["matter_category"] = dr[3]
                        s2 = dr[4]
                        if s2:
                            s2 = s2.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            s2 = ""
                        dict_data["product_time"] = s2
                        dict_data["work_name"] = dr[5]
                        dict_data["user_code"] = dr[6]
                        s = dr[7]
                        if s:
                            s = s.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            s= ""
                        dict_data["out_time"] = s
                        dict_data["finished_product_code"] = dr[8]
                        dict_data["product_plan_code"] = dr[9]
                        data_list.append(dict_data)
                    data_list.sort(key=lambda x: (x['out_time']), reverse=True)
                    page_result = Page(page, page_size, data_list)
                    data = page_result.get_str_json()

                    sql_num = int(len(data_list))
                    data_add_int["data"] = data
                    data_add_int["total"] = sql_num

                else:
                    data_add_int["data"] = []
                    data_add_int["total"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("??????????????????%s")
            db.close_connection(conn)

            return HttpResponse(result)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????,%s" % e)
            return HttpResponse(data)


class ProductInfoSearch(View):
    """
    ???????????????

    """
    def get(self, request):
        try:
            data_add_int = {}
            db = DB()
            data_list = []
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))
            product_id = request.session.get("session_projectId")
            finished_product_code = request.GET.get("finished_product_code")
            matter_id = request.GET.get("matter_id")

            conn = db.get_connection(product_id)
            if finished_product_code:
                finished_product_code_sql = "and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""
            if matter_id:
                matter_get_product_code = ""
                sql_table = """
                            show tables like "product_transit_%" 
                            """
                table_list = db.execute_sql(conn, sql_table)
                for table_type in table_list:
                    table_one = table_type[0]
                    sql = """
                    select finished_product_code from {0} where matter_id = '{1}'
                    """
                    sql = sql.format(table_one, matter_id)
                    finisheds = db.execute_sql(conn, sql)
                    if finisheds:
                        matter_get_product_code = finisheds[0][0]
                    else:
                        pass
                matter_id_sql = "and finished_product_code = " + "'" + matter_get_product_code + "'"

            else:
                matter_id_sql = ""

            conn_db = db.get_connection("db_common")

            sql_table1 = """
            select rule from productlist where product_id = '{0}'
            """
            sql_table1 = sql_table1.format(product_id)
            drs = db.execute_sql(conn_db, sql_table1)
            rule = drs[0][0]
            sql_enter_storage = """
            SELECT a.product_name, b.finished_product_code,a.status,a.pack_id,a.enter_time
            FROM enter_storage_status as a left join enter_storage as b on a.pack_id = b.pack_id
            where 2>1 {0} {1} order by a.enter_time desc;
            """
            sql_enter_storage = sql_enter_storage.format(finished_product_code_sql, matter_id_sql)
            dfs = db.execute_sql(conn, sql_enter_storage)
            if dfs:
                for df in dfs:
                    data_dict = {}
                    data_dict["product_name"] = df[0]
                    data_dict["finished_product_code"] = df[1]
                    data_dict["rule"] = rule
                    data_dict["status"] = df[2]
                    data_dict["pack_id"] = df[3]
                    s = df[4]
                    if s:
                        s = s.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        s = ""
                    data_dict["enter_time"] = s
                    sql_table = """
                                show tables like "product_transit_%" 
                                """
                    table_list = db.execute_sql(conn, sql_table)

                    finish_matter_list = []
                    for table_type in table_list:
                        table_one = table_type[0]
                        sql = """
                        select matter_id from {0} where finished_product_code = '{1}'
                        """
                        sql = sql.format(table_one, df[1])
                        matters = db.execute_sql(conn, sql)
                        if matters:
                            for matter in matters:
                                finish_matter_list.append(matter[0])
                        else:
                            pass
                    data_dict["response_datas"] = finish_matter_list

                    data_list.append(data_dict)
                data_list.sort(key=lambda x: (x['enter_time']), reverse=True)

                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                sql_num = int(len(data_list))
                data_add_int["data"] = data
                data_add_int["total"] = sql_num
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("??????????????????%s")
            db.close_connection(conn)
            return HttpResponse(result)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("???????????????????????????,%s" % e)
            return HttpResponse(data)


class NoPageMatterSearch(View):
    """
    ???i?????????????????????

    """
    def get(self, request):
        try:
            data_add_int = {}
            db = DB()

            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            finished_product_code = request.GET.get("finished_product_code")
            matter_id = request.GET.get("matter_id")

            sql_table = """
                        show tables like "product_transit_%" 
                        """
            table_list = db.execute_sql(conn, sql_table)

            if finished_product_code:
                finished_product_code_sql = "and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""
            if matter_id:
                matter_id_sql = "and matter_id = " + "'" + matter_id + "'"
            else:
                matter_id_sql = ""

            data_list = []
            for table_type in table_list:
                table_one = table_type[0]
                sql = """
                SELECT b.matter_name,a.matter_id,b.rule,
                b.matter_category,c.product_time,
                d.work_name,a.user_code,a.out_time,
                a.finished_product_code FROM {0} as a left join 
                matter_list as b on a.matter_code = b.bom_matter_code 
                left join person_matter as c on c.matter_code = a.matter_code 
                left join work_station as d on d.work_code = a.work_code where 2 > 1 {1} {2}
                """
                sql = sql.format(table_one, finished_product_code_sql, matter_id_sql)
                drs = db.execute_sql(conn, sql)
                if drs:
                    for dr in drs:
                        dict_data = {}
                        dict_data["matter_name"] = dr[0]
                        dict_data["matter_id"] = dr[1]
                        dict_data["rule"] = dr[2]
                        dict_data["matter_category"] = dr[3]
                        s2 = dr[4]
                        if s2:
                            s2 = s2.strftime("%Y-%m-%d %H:%M:%S ")
                        dict_data["product_time"] = s2
                        dict_data["work_name"] = dr[5]
                        dict_data["user_code"] = dr[6]
                        s = dr[7]
                        if s:
                            s = s.strftime("%Y-%m-%d %H:%M:%S ")
                        dict_data["out_time"] = s
                        dict_data["finished_product_code"] = dr[8]
                        data_list.append(dict_data)
                data = data_list
                sql_num = int(len(data_list))
                data_add_int["data"] = data
                data_add_int["total"] = sql_num

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("??????????????????????????????")
            db.close_connection(conn)
            return HttpResponse(result)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????????????????,%s" % e)
            return HttpResponse(data)


class NoPageProductInfoSearch(View):
    """
    ????????????????????????
    """
    def get(self, request):
        try:
            data_add_int = {}
            db = DB()
            data_list = []
            product_id = request.session.get("session_projectId")
            finished_product_code = request.GET.get("finished_product_code")
            matter_id = request.GET.get("matter_id")

            conn = db.get_connection(product_id)
            if finished_product_code:
                finished_product_code_sql = "and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""
            if matter_id:
                matter_get_product_code = ""
                sql_table = """
                            show tables like "product_transit_%" 
                            """
                table_list = db.execute_sql(conn, sql_table)
                for table_type in table_list:
                    table_one = table_type[0]
                    sql = """
                    select finished_product_code from {0} where matter_id = '{1}'
                    """
                    sql = sql.format(table_one, matter_id)
                    finisheds = db.execute_sql(conn, sql)
                    if finisheds:
                        matter_get_product_code = finisheds[0][0]
                    else:
                        pass
                matter_id_sql = "and finished_product_code = " + "'" + matter_get_product_code + "'"

            else:
                matter_id_sql = ""

            conn_db = db.get_connection("db_common")

            sql_table1 = """
            select rule from productlist where product_id = '{0}'
            """
            sql_table1 = sql_table1.format(product_id)
            drs = db.execute_sql(conn_db, sql_table1)
            rule = drs[0][0]
            sql_enter_storage = """
            SELECT a.product_name, b.finished_product_code,a.status,a.pack_id,a.enter_time
            FROM enter_storage_status as a left join enter_storage as b on a.pack_id = b.pack_id
            where 2>1 {0} {1} order by a.enter_time desc;
            """
            sql_enter_storage = sql_enter_storage.format(finished_product_code_sql, matter_id_sql)
            dfs = db.execute_sql(conn, sql_enter_storage)
            if dfs:
                for df in dfs:
                    data_dict = {}
                    data_dict["product_name"] = df[0]
                    data_dict["finished_product_code"] = df[1]
                    data_dict["rule"] = rule
                    data_dict["status"] = df[2]
                    data_dict["pack_id"] = df[3]
                    s = df[4]
                    if s:
                        s = s.strftime("%Y-%m-%d %H:%M:%S ")
                    else:
                        s = ""
                    data_dict["enter_time"] = s
                    sql_table = """
                                show tables like "product_transit_%" 
                                """
                    table_list = db.execute_sql(conn, sql_table)

                    finish_matter_list = []
                    for table_type in table_list:
                        table_one = table_type[0]
                        sql = """
                        select matter_id from {0} where finished_product_code = '{1}'
                        """
                        sql = sql.format(table_one, df[1])
                        matters = db.execute_sql(conn, sql)
                        if matters:
                            for matter in matters:
                                finish_matter_list.append(matter[0])
                        else:
                            pass
                    data_dict["response_datas"] = finish_matter_list

                    data_list.append(data_dict)
                data_list.sort(key=lambda x: (x['enter_time']), reverse=True)

                sql_num = int(len(data_list))
                data_add_int["data"] = data_list
                data_add_int["total"] = sql_num

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("??????????????????????????????")
            db.close_connection(conn)
            return HttpResponse(result)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("???????????????????????????,%s" % e)
            return HttpResponse(data)


class OperationSystem(View):
    """
    ????????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            sql = """
                    select * from operate where 2 > 1
                   """
            drs = db.execute_sql(conn, sql)

            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["ID"] = dr[0]
                    dict_data["Package_Qty"] = dr[1]
                    dict_data["Rv"] = dr[2]
                    dict_data["Itemcode_C_Shipping"] = dr[3]
                    dict_data["Supplier"] = dr[4]
                    dict_data["No_Ship"] = dr[5]
                    dict_data["CS_type"] = dr[6]
                    dict_data["Shipping_SN_length"] = dr[7]
                    dict_data["Product_length"] = dr[8]
                    data_list.append(dict_data)

                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                sql_num = """
                            select count(*) from operate where 2 > 1
                            """

                dfs = db.execute_sql(conn, sql_num)
                if dfs:
                    sql_num_int = dfs[0][0]
                else:
                    sql_num_int = 0
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("OperationSystem>>>>????????????%s")
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("OperationSyste????????????%s"% e)
            return HttpResponse(result)

    def post(self, request):
        """

        ???????????????????????????
        :param request:
        :return:
        """
        try:
            json_data = request.body
            str_data = json.loads(json_data)

            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            sql_matter = """
                   SELECT max(order_number) FROM operate;
                   """
            matter_type = db.execute_sql(conn, sql_matter)
            if matter_type[0][0] == None:
                matter_num = 1
            else:
                matter_num = matter_type[0][0] + 1
            # id = str_data.get("ID")
            Package_Qty = str_data.get("Package_Qty")
            Rv = str_data.get("Rv")
            Itemcode_C_Shipping = str_data.get("Itemcode_C_Shipping")
            Supplier = str_data.get("Supplier")
            No_Ship = str_data.get("No_Ship")
            CS_type = str_data.get("CS_type")
            Shipping_SN_length = str_data.get("Shipping_SN_length")
            Product_length = str_data.get("Product_length")

            sql_insert = """
                               insert into operate(ID, Package_Qty,
                               Rv,
                               Itemcode_C_Shipping, Supplier, No_Ship, CS_type, Shipping_SN_length, Product_length, order_number)
                               values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}', '{7}', '{8}', '{9}')
                               """
            sql_insert_format = sql_insert.format(matter_num, Package_Qty,
                                                  Rv, Itemcode_C_Shipping,
                                                  Supplier, No_Ship, CS_type, Shipping_SN_length,
                                                  Product_length, matter_num)
            drs = db.execute_sql(conn, sql_insert_format)

            data = {"code": 0, "message": "????????????"}
            logger.info("OperationSystem????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            logger.error("OperationSystem????????????????????????")
            data = json.dumps(data)
            return HttpResponse(data)


class PutOperationSystem(View):
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            ID = request.GET.get("ID")
            Package_Qty = request.GET.get("Package_Qty")
            Rv = request.GET.get("Rv")
            Itemcode_C_Shipping = request.GET.get("Itemcode_C_Shipping")
            Supplier = request.GET.get("Supplier")
            No_Ship = request.GET.get("No_Ship")
            CS_type = request.GET.get("CS_type")
            Shipping_SN_length = request.GET.get("Shipping_SN_length")
            Product_length = request.GET.get("Product_length")
            sql = """
                   update operate set
                  Package_Qty='{0}',Rv='{1}',Itemcode_C_Shipping='{2}',Supplier='{3}',
                  No_Ship='{4}',CS_type='{5}', Shipping_SN_length='{6}', Product_length='{7}' where ID = '{8}'
                   """
            sql = sql.format(Package_Qty, Rv, Itemcode_C_Shipping, Supplier, No_Ship,
                             CS_type, Shipping_SN_length, Product_length, ID)
            drs = db.execute_sql(conn, sql)
            result = {"code": 0, "message": "????????????"}
            result = json.dumps(result)
            logger.info("OperationSystem>>>>??????????????????%s")
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("??????OperationSystem ??????%s"% e)
            return HttpResponse(result)


class DeleteOperationSystem(View):
    """
    ??????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            ID = request.GET.get("ID")
            sql = """
                   delete from operate where ID = '{0}'
                   """
            sql = sql.format(ID)
            drs = db.execute_sql(conn, sql)
            result = {"code": 0, "message": "????????????"}
            result = json.dumps(result)
            logger.info("??????DeleteOperationSystem?????????")
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("??????DeleteOperationSystem?????????%s"% e)
            return HttpResponse(result)


class AddModel(View):
    def post(self, request):
        """
        ?????????????????????

        :param request:
        :return:
        """
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            type_name = request.POST.get("fileType")
            f = request.FILES.get("file")  # ?????????????????????
            BASE_DIR = Path(__file__).resolve().parent.parent
            path_workstation = os.path.join(BASE_DIR, "media")
            excit_psth = os.path.join(path_workstation, product_id)
            if product_id:
                if "txt" in str(f).split("."):
                    sql_create = """
                    CREATE TABLE IF NOT EXISTS project_model(id int AUTO_INCREMENT primary key not null, 
                                              product_id varchar(128),
                                              model_name varchar(128),
                                              type_name varchar(128))
                    """
                    db.execute_sql(conn, sql_create)

                    sql_check = """
                    select * from project_model where product_id = "{0}" and type_name = "{1}"
                    """
                    sql_check = sql_check.format(product_id, type_name)
                    drs = db.execute_sql(conn, sql_check)

                    if drs:
                        result = {"code": 1, "message": "?????????txt??????????????????????????????????????????"}
                        result = json.dumps(result)

                    else:
                        sql_insert = """
                        insert into project_model(product_id, model_name,
                                       type_name)
                                       values('{0}', '{1}', '{2}')
                        """

                        sql_insert = sql_insert.format(product_id, f.name, type_name)

                        db.execute_sql(conn, sql_insert)
                        if os.path.exists(excit_psth):  # ?????????????????????????????????????????????????????????
                            with open(os.path.join(excit_psth, f.name), "wb+") as k:
                                for chunk in f.chunks():
                                    k.write(chunk)
                        else:
                            os.makedirs(excit_psth)   # ????????????????????????????????????????????????????????????????????????

                            with open(os.path.join(excit_psth, f.name), "wb+") as k:
                                for chunk in f.chunks():
                                    k.write(chunk)

                        result = {"code": 0, "message": "????????????"}
                        result = json.dumps(result)
                        logger.info("??????????????????i??????%s")

                else:
                    if os.path.exists(excit_psth):  # ?????????????????????????????????????????????????????????
                        with open(os.path.join(excit_psth, f.name), "wb+") as k:
                            for chunk in f.chunks():
                                k.write(chunk)
                    else:
                        os.makedirs(excit_psth)  # ????????????????????????????????????????????????????????????????????????

                        with open(os.path.join(excit_psth, f.name), "wb+") as k:
                            for chunk in f.chunks():
                                k.write(chunk)
                        print(excit_psth, "????????????,??????????????????")
                    result = {"code": 0, "message": "????????????"}
                    result = json.dumps(result)
                    logger.info("chengggong")


            else:
                result = {"code": 1, "message": "session???????????????????????????????????????"}
                result = json.dumps(result)
            db.close_connection(conn)

            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("????????????")
            return HttpResponse(result)


class test(View):
    def get(self, request):
        try:
            BASE_DIR = Path(__file__).resolve().parent.parent
            path_file = os.path.join(BASE_DIR, r"utils\import_common.py")

            result = {"code": 0, "message": "????????????"}
            result = json.dumps(result)

            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            return HttpResponse(result)


class PrintApai(View):
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")

            conn = db.get_connection(product_id)
            type_name = request.GET.get("fileType")

            work_id = request.session.get('session_workId')
            BASE_DIR = Path(__file__).resolve().parent.parent
            path_workstation = os.path.join(BASE_DIR, "media")
            store_work = str(product_id) + "_" + str(work_id)
            print_path = os.path.join(path_workstation, store_work)

            if os.path.exists(print_path):
                pass
            else:
                os.makedirs(print_path)

            if type_name == "box_code":
                stem1 = request.GET.get("stem1")
                sql = """
                            select model_name from project_model where product_id = "{0}" and type_name = "{1}"        
                            """
                sql = sql.format(product_id, type_name)

                drs = db.execute_sql(conn, sql)
                if drs[0][0]:
                    model_name = drs[0][0]
                    file_path1 = os.path.join(path_workstation, product_id)
                    file_path = os.path.join(file_path1, model_name)
                    with open(file_path, "r", encoding="utf-8") as f:
                        str1 = f.read()
                        str2 = str1.replace("{0}", stem1)
                    print_file_path = os.path.join(print_path, model_name)
                    with open(print_file_path, "w+", encoding="utf-8") as k:
                        k.write(str2)
                    data = "print_file_path"
                    result = {"code": 0, "message": "????????????", "data": data}
                    result = json.dumps(result)
                    return HttpResponse(result)
                else:
                    result = {"code": 1, "message": "????????????txt????????????"}
                    result = json.dumps(result)
                    return HttpResponse(result)

            elif type_name == "matter_code":
                stem1 = request.GET.get("stem1")
                stem2 = request.GET.get("stem2")
                stem3 = request.GET.get("stem3")
                stem4 = request.GET.get("stem4")
                stem5 = request.GET.get("stem5")
                stem6 = request.GET.get("stem6")
                sql = """
                    select model_name from project_model where product_id = "{0}" and type_name = "{1}"        
                    """
                sql = sql.format(product_id, type_name)

                drs = db.execute_sql(conn, sql)
                if drs[0][0]:
                    model_name = drs[0][0]
                    file_path1 = os.path.join(path_workstation, product_id)
                    file_path = os.path.join(file_path1, model_name)
                    with open(file_path, "r", encoding="utf-8") as f:
                        str1 = f.read()
                        str2 = str1.replace("{0}", stem1)
                        str2 = str2.replace("{1}", stem2)
                        str2 = str2.replace("{2}", stem3)
                        str2 = str2.replace("{3}", stem4)
                        str2 = str2.replace("{4}", stem5)
                        str2 = str2.replace("{5}", stem6)
                    print_file_path = os.path.join(print_path, model_name)
                    with open(print_file_path, "w+", encoding="utf-8") as k:
                        k.write(str2)
                    data = "print_file_path"

                    result = {"code": 0, "message": "????????????", "data": data}
                    result = json.dumps(result)
                    return HttpResponse(result)
                else:
                    result = {"code": 1, "message": "????????????txt????????????"}
                    result = json.dumps(result)
                    return HttpResponse(result)

            elif type_name == "deliver_goods_code":
                stem1 = request.GET.get("stem1")
                stem2 = request.GET.get("stem2")
                stem3 = request.GET.get("stem3")
                stem4 = request.GET.get("stem4")
                stem5 = request.GET.get("stem5")
                stem6 = request.GET.get("stem6")
                sql = """
                    select model_name from project_model where product_id = "{0}" and type_name = "{1}"        
                    """
                sql = sql.format(product_id, type_name)
                drs = db.execute_sql(conn, sql)

                if drs[0][0]:

                    model_name = drs[0][0]
                    file_path1 = os.path.join(path_workstation, product_id)
                    file_path = os.path.join(file_path1, model_name)

                    with open(file_path, "r", encoding="utf-8") as f:
                        str1 = f.read()
                        str2 = str1.replace("{0}", stem1)
                        str2 = str2.replace("{1}", stem2)
                        str2 = str2.replace("{2}", stem3)
                        str2 = str2.replace("{3}", stem4)
                        str2 = str2.replace("{4}", stem5)
                        str2 = str2.replace("{5}", stem6)
                    print_file_path = os.path.join(print_path, model_name)
                    with open(print_file_path, "w+", encoding="utf-8") as k:
                        k.write(str2)
                    data = "print_file_path"
                    result = {"code": 0, "message": "????????????", "data": data}
                    result = json.dumps(result)
                    db.close_connection(conn)
                    return HttpResponse(result)
                else:
                    result = {"code": 1, "message": "????????????txt????????????"}
                    result = json.dumps(result)
                    db.close_connection(conn)
                    return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????"}
            result = json.dumps(result)
            return HttpResponse(result)


class WorkManage(View):
    """
    ?????? ?????? ??????????????????????????????????????????????????????

    """
    def get(self, request):
        try:
            data_list = []
            matter_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            # conn = db.get_connection("db_common")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            table_name = request.GET.get("table_name")
            work_code = request.GET.get("work_code")

            if table_name:
                table_name_sql = "and table_name = " + "'" + table_name + "'"
            else:
                table_name_sql = ""
            if work_code:
                work_code_sql = "and work_code = " + "'" + work_code + "'"
            else:
                work_code_sql = ""

            sql = """
             select * from work_transit where 2 > 1 {0} {1}
            """

            sql_main = sql.format(table_name_sql, work_code_sql)

            drs = db.execute_sql(conn, sql_main)
            if drs:

                for dr in drs:
                    dict_data = {}
                    dict_data["table_code"] = dr[0]
                    dict_data["table_name"] = dr[1]
                    dict_data["work_code"] = dr[2]
                    dict_data["description"] = dr[3]

                    s = dr[4]
                    if s:
                        s = s.strftime("%Y-%m-%d %H:%M:%S ")
                        dict_data["create_time"] = s
                    else:
                        dict_data["create_time"] = ""

                    sql_matter = """
                    select * from field_list where table_name = '{0}' order by order_number asc;
                    """
                    sql_matter_format = sql_matter.format(dr[1])
                    matter_dfs = db.execute_sql(conn, sql_matter_format)
                    if matter_dfs:
                        respose_list = []
                        for matter_df in matter_dfs:
                            response_data = {}
                            response_data["field_code"] = matter_df[0]
                            response_data["table_name"] = matter_df[1]
                            response_data["field_name"] = matter_df[2]
                            response_data["field_type"] = matter_df[3]
                            response_data["field_PK"] = matter_df[4]
                            response_data["field_NN"] = matter_df[5]
                            response_data["field_AI"] = matter_df[6]
                            # response_data["order_number"] = matter_df[7]
                            respose_list.append(response_data)
                        dict_data["response_datas"] = respose_list
                        data_list.append(dict_data)
                    else:
                        dict_data["response_datas"] = []
                        data_list.append(dict_data)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()

                sql_num_int = int(len(data_list))
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            db.close_connection(conn)

            return HttpResponse(result)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "??????BOM??????????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????")

            return HttpResponse(data)

    def post(self, request):
        """

        ?????????????????????????????????????????????????????????
        :param request:
        :return:
        """
        try:
            product_id = request.session.get("session_projectId")
            json_data = request.body
            str_data = json.loads(json_data)
            product_db = str_data.get("product_id")

            work_code = str_data.get("work_code")
            # table_code = str_data.get("table_code")
            table_name = str_data.get("table_name")
            description = str_data.get("description")
            response_datas = str_data.get("response_datas")
            now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ")

            db = DB()

            conn = db.get_connection(product_id)

            sql_transit = """
            SELECT max(order_number) FROM work_transit;
            """
            matter_type = db.execute_sql(conn, sql_transit)
            if matter_type[0][0] == None:
                transit_num = 1
            else:
                transit_num = matter_type[0][0] + 1

            sql_filed = """
                        SELECT max(order_number) FROM field_list;
                        """
            matter_type = db.execute_sql(conn, sql_filed)
            if matter_type[0][0] == None:
                field_num = 0
            else:
                field_num = matter_type[0][0]

            table_code = str("Im_Work_Tran" + str(transit_num))

            sql_check = """
            select * from work_transit where table_name = '{0}' and work_code = '{1}'
            """
            sql_check = sql_check.format(table_name, work_code)
            drs_checks = db.execute_sql(conn, sql_check)
            if drs_checks:
                data = {"code": 1, "message": "??????????????????????????????????????????????????????"}
                logger.error("??????????????????????????????????????????")
                data = json.dumps(data)
                return HttpResponse(data)
            else:
                sql_insert = """
                            insert into work_transit(table_code, table_name, work_code,
                            description,
                            create_time, order_number)
                            values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                            """
                sql_insert_format = sql_insert.format(table_code, table_name, work_code,
                                                      description, now_time, transit_num)

                drs = db.execute_sql(conn, sql_insert_format)

                for response_data in response_datas:
                    # field_code = response_data.get("field_code")
                    field_name = response_data.get("field_name")
                    field_type = response_data.get("field_type")
                    field_PK = response_data.get("field_PK")
                    field_NN = response_data.get("field_NN")
                    field_AI = response_data.get("field_AI")
                    field_num = field_num + 1
                    field_code = str("Im_Field_" + str(field_num))
                    sql_response = """
                            insert into field_list(field_code,table_name, field_name,
                            field_type, field_PK, field_NN, field_AI, order_number)
                            values('{0}', '{1}', '{2}', '{3}', '{4}','{5}', '{6}', '{7}')
                    """
                    sql_response_format = sql_response.format(field_code, table_name, field_name,
                                                              field_type, field_PK, field_NN, field_AI, field_num)
                    matter_drs = db.execute_sql(conn, sql_response_format)

                sql_create = """
                CREATE TABLE IF NOT EXISTS {0}({1});
                """
                location_1 = ""
                location_2 = ""
                for response_data in response_datas:

                    field_name = response_data.get("field_name")
                    field_type = response_data.get("field_type")
                    field_PK = response_data.get("field_PK")
                    field_NN = response_data.get("field_NN")
                    field_AI = response_data.get("field_AI")

                    if field_AI:
                        field_AI = " AUTO_INCREMENT "
                    else:
                        field_AI = ""
                    #
                    # if field_PK:
                    #     field_PK = " primary key "
                    # else:
                    #     field_PK = ""
                    if field_NN:
                        field_NN = " not null "
                    else:
                        field_NN = ""
                    if field_type == "VARCHAR":
                        field_type = "varchar(128)"

                    # if field_type == "DATETIME":
                    #     field_type = " DATETIME DEFAULT CURRENT_TIMESTAMP "
                    # if field_PK:
                    #     location_1 = field_name + " "+ field_type + field_AI + field_PK+ field_NN
                    # else:
                    location_2 = location_2 + "," + field_name + " " + field_type + field_AI+ field_NN
                location = """
                id int AUTO_INCREMENT primary key not null,finished_product_code varchar(128),
                Result varchar(128),time DATETIME DEFAULT CURRENT_TIMESTAMP not null
                """

                values = location + location_2

                sql_create = sql_create.format(table_name, values)

                db.execute_sql(conn, sql_create)

                data = {"code": 0, "message": "?????????????????????"}
                logger.info("?????????????????????")
                data = json.dumps(data)
                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "?????????????????????"}
            logger.error("????????????????????????")
            data = json.dumps(data)
            return HttpResponse(data)


class PutWorkManage(View):
    """

    ?????????????????????---????????????????????????????????? ????????????????????????????????????
    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            work_code = request.GET.get("work_code")
            table_code = request.GET.get("table_code")
            table_name = request.GET.get("table_name")
            description = request.GET.get("description")
            response_datas = request.GET.get("response_datas")
            create_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ")
            now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S ")
            if response_datas:
                response_datas = eval(response_datas)
            sql = """
                         update work_transit set
                         table_name='{0}',work_code='{1}', description ='{2}',
                         create_time= '{3}' where table_code = '{4}'
                         """
            sql_main = sql.format(table_name, work_code, description, create_time,
                                  table_code)
            drs = db.execute_sql(conn, sql_main)

            sql_matter_num = """
                    SELECT max(order_number) FROM field_list
                   """
            dras = db.execute_sql(conn, sql_matter_num)

            if dras:
                sql_num = dras[0][0]
            else:
                sql_num = 0

            sql_check_name = """
            SELECT * FROM {0}
            """
            sql_check_name = sql_check_name.format(table_name)
            ds_nums = db.execute_sql(conn, sql_check_name)
            # print("sartart===========>", ds_nums)
            # ds_num = ds_nums[0][0]

            if ds_nums:
                data = {"code": 1, "message": "???????????????????????????,????????????"}
                logger.error("???????????????????????????,????????????")
                data = json.dumps(data)
                db.close_connection(conn)

                return HttpResponse(data)

            else:
                sql_delete = """
                            drop table if exists {0}
                            """
                sql_delete = sql_delete.format(table_name)
                db.execute_sql(conn, sql_delete)
                sql_matter = """
                              SELECT field_code FROM field_list where table_name = '{0}'
                              """
                sql_matter_format = sql_matter.format(table_name)
                field_dfs = db.execute_sql(conn, sql_matter_format)
                for field_df in field_dfs:
                    sql_delete = """
                                       delete FROM field_list where field_code = '{0}'
                                       """
                    sql_delete_format = sql_delete.format(field_df[0])
                    db.execute_sql(conn, sql_delete_format)
                sql_delete_transit = """
                delete from work_transit where table_code = '{0}'
                """
                sql_delete_transit = sql_delete_transit.format(table_code)
                db.execute_sql(conn, sql_delete_transit)
                sql_transit = """
                    SELECT max(order_number) FROM work_transit;
                    """
                matter_type = db.execute_sql(conn, sql_transit)
                if matter_type[0][0] == None:
                    transit_num = 1
                else:
                    transit_num = matter_type[0][0] + 1

                sql_filed = """
                                SELECT max(order_number) FROM field_list;
                                """
                matter_type = db.execute_sql(conn, sql_filed)
                if matter_type[0][0] == None:
                    field_num = 0
                else:
                    field_num = matter_type[0][0]

                table_code = str("Im_Work_Tran" + str(transit_num))

                sql_insert = """
                                            insert into work_transit(table_code, table_name, work_code,
                                            description,
                                            create_time, order_number)
                                            values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                                            """
                sql_insert_format = sql_insert.format(table_code, table_name, work_code,
                                                      description, now_time, transit_num)

                drs = db.execute_sql(conn, sql_insert_format)

                for response_data in response_datas:
                    # field_code = response_data.get("field_code")
                    field_name = response_data.get("field_name")
                    field_type = response_data.get("field_type")
                    field_PK = response_data.get("field_PK")
                    field_NN = response_data.get("field_NN")
                    field_AI = response_data.get("field_AI")
                    field_num = field_num + 1
                    field_code = str("Im_Field_" + str(field_num))
                    sql_response = """
                                insert into field_list(field_code,table_name, field_name,
                                field_type, field_PK, field_NN, field_AI, order_number)
                                values('{0}', '{1}', '{2}', '{3}', '{4}','{5}', '{6}', '{7}')
                        """
                    sql_response_format = sql_response.format(field_code, table_name, field_name,
                                                              field_type, field_PK, field_NN, field_AI, field_num)

                    matter_drs = db.execute_sql(conn, sql_response_format)
                sql_create = """
                    CREATE TABLE IF NOT EXISTS {0}({1});
                    """
                location_1 = ""
                location_2 = ""
                for response_data in response_datas:

                    field_name = response_data.get("field_name")
                    field_type = response_data.get("field_type")
                    field_PK = response_data.get("field_PK")
                    field_NN = response_data.get("field_NN")
                    field_AI = response_data.get("field_AI")


                    if field_AI == "True" or field_AI == "true":
                        field_AI = " AUTO_INCREMENT "
                    else:
                        field_AI = ""

                    if field_PK == "True" or field_PK == "true":
                        field_PK = " primary key "
                    else:
                        field_PK = ""
                    if field_NN == "True" or field_NN == "true":
                        field_NN = " not null "
                    else:
                        field_NN = ""
                    if field_type == "VARCHAR":
                        field_type = "varchar(128)"

                    if field_type == "DATETIME":
                        field_type = " DATETIME DEFAULT CURRENT_TIMESTAMP "
                    if field_PK:
                        location_1 = field_name + " " + field_type + field_AI + field_PK  + field_NN
                    else:
                        location_2 = location_2 + "," + field_name + " "  + field_type + field_AI + field_NN

                values = location_1 + location_2

                sql_create = sql_create.format(table_name, values)

                db.execute_sql(conn, sql_create)

                data = {"code": 0, "message": "?????????????????????"}
                logger.info("?????????????????????")

                data = json.dumps(data)
                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            logger.error("????????????---->%s"% e)
            data = json.dumps(data)
            return HttpResponse(data)


class DeleteWorkManage(View):
    """

    ????????????????????????
    """
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            table_code = request.GET.get("table_code")
            # table_name = request.GET.get("table_name")
            sql_ = """
            select table_name from work_transit where table_code = '{0}'
            """
            sql_ = sql_.format(table_code)
            dfs = db.execute_sql(conn, sql_)
            table_name = dfs[0][0]
            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            # materials_production_code = "Im_Product_Pick_Matter_1"
            # response_datas = [{"materials_code": "Im_Materials_Pick_1"}]
            if response_datas:
                for response_data in response_datas:
                    field_code = dict(response_data).get("field_code")
                    field_name = dict(response_data).get("field_name")
                    sql = """
                       delete from field_list where field_code = '{0}'
                       """
                    sql_format = sql.format(field_code)
                    db.execute_sql(conn, sql_format)

                    sql_delete = """
                    ALTER TABLE {0} DROP COLUMN {1}             
                    """
                    sql_delete = sql_delete.format(table_name, field_name)
                    db.execute_sql(conn, sql_delete)
            else:
                sql_table = """
                select * from {0}
                """
                sql_table = sql_table.format(table_name)
                drs_tables = db.execute_sql(conn, sql_table)
                if drs_tables:
                    data = {"code": 1, "message": "???????????????????????????????????????????????????"}
                    logger.error("???????????????????????????,????????????")
                    data = json.dumps(data)
                    db.close_connection(conn)

                    return HttpResponse(data)
                else:
                    sql_matter = """
                                  SELECT field_code FROM field_list where table_name = '{0}'
                                  """
                    sql_matter_format = sql_matter.format(table_name)
                    field_dfs = db.execute_sql(conn, sql_matter_format)
                    for field_df in field_dfs:
                        sql_delete = """
                           delete FROM field_list where field_code = '{0}'
                           """
                        sql_delete_format = sql_delete.format(field_df[0])
                        db.execute_sql(conn, sql_delete_format)
                    sql = """
                           delete from work_transit where table_code = '{0}'
                           """
                    sql_main = sql.format(table_code)
                    des = db.execute_sql(conn, sql_main)

                    sql_drop = """
                    drop table if exists {0}
                    """
                    sql_drop = sql_drop.format(table_name)
                    db.execute_sql(conn, sql_drop)

                    data = {"code": 0, "message": "????????????"}
                    logger.info("?????????????????????????????????")
                    data = json.dumps(data)
                    db.close_connection(conn)
                    return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????,%s" % e)
            return HttpResponse(data)


class NoPageProcessMatterDeal(View):
    """

    ????????????????????????????????? ??????????????????
    """
    def get(self, request):
        try:
            lid = []
            data_list = []
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            matter_code = request.GET.get("matter_code")
            work_id = request.GET.get("work_id")

            if matter_code:
                matter_code_sql = " and matter_code = " + "'" + matter_code + "'"
            else:
                matter_code_sql = ""

            if work_id:
                work_id_sql = " and work_id !=  " + "'" + work_id + "'"
            else:
                work_id_sql = ""

            sql = """
            select * from process_matter_deal where 2 > 1 {0} {1}
            """
            sql = sql.format(matter_code_sql, work_id_sql)
            drs = db.execute_sql(conn, sql)
            for dr in drs:
                dict_data = {}
                # dict_data["process_matter_deal_code"] = dr[0]
                dict_data["work_id"] = dr[1]
                dict_data["matter_code"] = dr[2]
                dict_data["install_number"] = dr[3]
                data_list.append(dict_data)

            for i in data_list:
                if (i['work_id'], i["matter_code"]) not in lid:
                    lid.append((i['work_id'], i["matter_code"]))
            lm = []
            for i in lid:
                lm.append({'work_id': i[0], 'matter_code': i[1], "install_number": 0})

            for i in data_list:
                for o in lm:
                    if i['work_id'] == o['work_id'] and i["matter_code"] == o["matter_code"]:
                        o['install_number'] = o['install_number'] + i['install_number']

            data = {"code": 0, "message": "????????????", "data": lm}
            logger.info("??????????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????---???,%s" % e)
            return HttpResponse(data)


class PackCheckFinishedProduct(View):
    """
    ???????????????????????????
    """
    def get_live_data(self, product_id, finished_product_code):
        func = product_live()
        data_list = func.one_product_live(product_id, finished_product_code)

        return data_list

    def get(self, request):
        try:

            data_list = []
            db = DB()


            product_id = request.session.get("session_projectId")
            current_work_id = request.session.get("session_workId")
            finished_product_code = request.GET.get("finished_product_code")

            # product_id = "co2"
            # current_work_id = "Station_05"
            # finished_product_code = "SPR001300-2A001387-CMSS-A02-083021-0299"

            # request.session['session_workId'] = work_id
            conn = db.get_connection(product_id)

            data_list = self.get_live_data(product_id, finished_product_code)

            sql_storage = """
            select * from enter_storage where finished_product_code = "{0}"
            """
            sql_storage = sql_storage.format(finished_product_code)
            drs_sto = db.execute_sql(conn, sql_storage)
            if drs_sto:
                data = {"code": 1, "message": "???????????????????????????"}
                data = json.dumps(data)
                db.close_connection(conn)
                return HttpResponse(data)
            else:
                data_list.sort(key=lambda x: (x['time']), reverse=False)
                if data_list:
                    enter_work_code = data_list[-1]["work_code"]
                    test_result = data_list[-1]["test_result"]

                    sql_current_id = """
                    select leader_work_id from work_station where work_id = '{0}'
                    """
                    sql_current_id = sql_current_id.format(current_work_id)

                    ds = db.execute_sql(conn, sql_current_id)
                    if ds[0][0]:
                        lead_current_work = ds[0][0]

                        sql_check = """
                        select work_code from work_station where work_id = '{0}'
                        """
                        sql_check = sql_check.format(lead_current_work)
                        work_codes = db.execute_sql(conn, sql_check)
                        work_code = work_codes[0][0]


                        if work_code == enter_work_code:
                            if test_result == "PASS":
                                data = {"code": 0, "message": "??????????????????????????????,????????????"}
                                data = json.dumps(data)
                                db.close_connection(conn)
                                return HttpResponse(data)
                            else:
                                data = {"code": 1, "message": "?????????????????????????????????"}
                                data = json.dumps(data)
                                db.close_connection(conn)
                                return HttpResponse(data)
                        else:
                            data = {"code": 1, "message": "??????????????????????????????????????????????????????"}
                            data = json.dumps(data)

                            return HttpResponse(data)

                    else:
                        data = {"code": 1, "message": "??????????????????????????????????????????"}
                        data = json.dumps(data)
                        return HttpResponse(data)

                else:
                    data = {"code": 1, "message": "????????????????????????"}
                    data = json.dumps(data)
                    return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "?????????????????????????????????"}
            data = json.dumps(data)
            logger.error("?????????????????????????????????->,%s" % e)
            return HttpResponse(data)


class TransitProductQualifiedSearch(View):
    """
    ????????????????????????????????????--- ??????????????????????????????

    """

    def get(self, request):
        try:
            data_add_int = {}
            data_list = []
            db = DB()
            product_id = request.session.get("session_projectId")
            work_id = request.session.get("session_workId")
            select_table = "product_transit_" + str(work_id)
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            start_time = request.GET.get("start_time")
            end_time = request.GET.get("end_time")

            # product_id = "co2"
            # work_id = "Station_05"
            # select_table = "product_transit_" + str(work_id)
            # page = 1
            # page_size = 100
            #
            # start_time = ""
            # end_time = ""
            # product_plan_code = ""

            conn = db.get_connection(product_id)
            product_plan_code = request.GET.get("product_plan_code")
            if product_plan_code:
                product_plan_code_sql = " and product_plan_code = " + "'" + product_plan_code + "'"
            else:
                product_plan_code_sql = ""

            if start_time:
                start_time_sql = " and out_time > " + "'" + start_time + "'"
            else:
                start_time_sql = ""
            if end_time:
                end_time_sql = " and out_time < " + "'" + end_time + "'"
            else:
                end_time_sql = ""

            # print("===->", work_id, product_plan_code_sql, start_time_sql, end_time_sql)

            pass_num = 0
            fail_num = 0
            sql_search = """
                   select * from {0} where 2>1 {1} {2} {3} group by finished_product_code order by out_time desc;
                   """
            sql_search = sql_search.format(select_table, product_plan_code_sql, start_time_sql, end_time_sql)
            # print("=>", sql_search)

            dr_searchs = db.execute_sql(conn, sql_search)

            # print("-=->", dr_searchs)
            if dr_searchs:
                for dr_search in dr_searchs:
                    dict_data = {}
                    dict_data["finished_product_code"] = dr_search[3]
                    dict_data["user_code"] = dr_search[4]
                    dict_data["work_code"] = dr_search[5]
                    dict_data["test_result"] = dr_search[6]
                    if dr_search[6] == "FAIL":
                        fail_num += 1
                    else:
                        pass_num += 1
                    # dict_data["description"] = dr_search[7]
                    if dr_search[8]:
                        dp = dr_search[8]
                        dp = dp.strftime("%Y-%m-%d %H:%M:%S ")
                        dict_data["enter_time"] = dp
                    else:
                        dict_data["enter_time"] = ""
                    ds = dr_search[9]
                    sq = ds.strftime("%Y-%m-%d %H:%M:%S ")
                    dict_data["out_time"] = sq

                    dict_data["product_plan_code"] = dr_search[10]
                    dict_data["product_name"] = product_id
                    data_list.append(dict_data)


                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                dfs = int(len(data_list))
                sql_num_int = dfs
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
                data_add_int["pass_num"] = pass_num
                data_add_int["fail_num"] = fail_num
                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)

                print("=--=>", result)
                logger.info("??????????????????????????????:")
                db.close_connection(conn)
                return HttpResponse(result)

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                data = {"code": 0, "message": "??????????????????", "data":data_add_int}
                data = json.dumps(data)
                db.close_connection(conn)
                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "???????????????????????????????????????"}
            data = json.dumps(data)
            logger.error("???????????????????????????????????????->,%s" % e)
            return HttpResponse(data)


class WorkUnqualifiedResult(View):
    """
    ????????????????????????????????????
    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))
            unqualified_result_code = request.GET.get("unqualified_result_code")
            work_id = request.GET.get("work_id")
            if unqualified_result_code:
                unqualified_result_code_sql = " and unqualified_result_code = " + "'" + unqualified_result_code + "'"
            else:
                unqualified_result_code_sql = ""
            if work_id:
                work_id_sql = " and work_id = " + "'" + work_id + "'"
            else:
                work_id_sql = ""


            sql_search = """
                               select * from work_unqualified_result where 2>1 {0} {1};
                               """
            sql_search = sql_search.format(unqualified_result_code_sql, work_id_sql)
            dr_searchs = db.execute_sql(conn, sql_search)
            if dr_searchs:
                for dr_search in dr_searchs:
                    dict_data = {}
                    dict_data["unqualified_result_code"] = dr_search[0]
                    dict_data["work_id"] = dr_search[1]
                    dict_data["unqualified_result_name"] = dr_search[2]
                    dict_data["content"] = dr_search[3]
                    dict_data["process_mode"] = dr_search[4]
                    dict_data["operate_user"] = dr_search[5]
                    if dr_search[6]:
                        dp = dr_search[6]
                        dp = dp.strftime("%Y-%m-%d %H:%M:%S")
                        dict_data["time"] = dp
                    else:
                        dict_data["time"] = ""
                    data_list.append(dict_data)
                data_list.sort(key=lambda x: (x['unqualified_result_code']), reverse=False)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                dfs = int(len(data_list))
                sql_num_int = dfs
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("????????????????????????????????????:")
                db.close_connection(conn)
                return HttpResponse(result)
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("???????????????????????????????????????????????????:")
                db.close_connection(conn)
                return HttpResponse(result)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("??????????????????????????????????????????->,%s" % e)
            return HttpResponse(data)

    def post(self, request):
        try:
            product_id = request.session.get("session_projectId")
            json_data = request.body
            str_data = json.loads(json_data)
            db = DB()
            conn = db.get_connection(product_id)

            operate_user_code = request.session.get('session_currentId')       # ???????????????
            work_id = str_data.get("work_id")
            unqualified_result_name = str_data.get("unqualified_result_name")
            content = str_data.get("content")
            process_mode = str_data.get("process_mode")

            conn_common = db.get_connection("db_common")

            sql_person = """
                        select user_name from person where user_code = '{0}'
                        """
            sql_person = sql_person.format(operate_user_code)
            dr_persons = db.execute_sql(conn_common, sql_person)
            if dr_persons:
                operate_user = dr_persons[0][0]
            else:
                sql_user = """
                select user_name from person where user_code = '{0}'
                """
                sql_user = sql_user.format(operate_user_code)
                dfs = db.execute_sql(conn, sql_user)
                if dfs:
                    operate_user = dfs[0][0]
                else:
                    operate_user = operate_user_code

            sql_filed = """
                       SELECT max(order_number) FROM work_unqualified_result;
                       """
            matter_type = db.execute_sql(conn, sql_filed)
            if matter_type[0][0] == None:
                field_num = 1
            else:
                field_num = matter_type[0][0] + 1
            unqualified_result_code = "F00000" + str(field_num)

            sql_insert_into = """
            insert into work_unqualified_result(unqualified_result_code, work_id, unqualified_result_name,
                                            content,
                                            process_mode, operate_user, order_number)
                                            values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}')
            """
            sql_insert_into = sql_insert_into.format(unqualified_result_code, work_id, unqualified_result_name,
                                                  content, process_mode, operate_user, field_num)

            drs = db.execute_sql(conn, sql_insert_into)

            data = {"code": 0, "message": "???????????????????????????"}
            logger.info("???????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "???????????????????????????"}
            logger.error("???????????????????????????")
            data = json.dumps(data)

            return HttpResponse(data)


class PutWorkUnqualifiedResult(View):
    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            operate_user = request.GET.get("operate_user")  # ???????????????
            work_id = request.GET.get("work_id")
            unqualified_result_name = request.GET.get("unqualified_result_name")
            content = request.GET.get("content")
            process_mode = request.GET.get("process_mode")
            unqualified_result_code = request.GET.get("unqualified_result_code")
            sql = """
                   update work_unqualified_result set
                  operate_user='{0}',work_id='{1}',unqualified_result_name='{2}',content='{3}',
                  process_mode='{4}' where unqualified_result_code = '{5}'
                   """
            sql = sql.format(operate_user, work_id, unqualified_result_name, content, process_mode,
                             unqualified_result_code)
            drs = db.execute_sql(conn, sql)
            result = {"code": 0, "message": "????????????"}
            result = json.dumps(result)
            logger.info("???????????????????????????")
            db.close_connection(conn)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("???????????????????????????")
            return HttpResponse(result)


class DeleteWorkUnqualifiedResult(View):
    """

    ????????????????????????????????????
    """
    def get(self, request):
        try:
            product_id = request.session.get("session_projectId")
            db = DB()
            conn = db.get_connection(product_id)

            unqualified_result_code = request.GET.get("unqualified_result_code")
            sql = """
            delete FROM work_unqualified_result where unqualified_result_code = '{0}'       
            """
            sql_format = sql.format(unqualified_result_code)
            db.execute_sql(conn, sql_format)
            data = {"code": 0, "message": "????????????"}
            data = json.dumps(data)
            logger.info("????????????????????????")
            db.close_connection(conn)
            return HttpResponse(data)
        except Exception as e:
            data = {"code":1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????")
            return HttpResponse(data)


class OneProductInAllWorkStationInfo(View):
    """
    ????????????????????????????????????  ???????????????  ??????  ???????????????    --> ?????????????????????
    """

    def get_live_data(self, product_id, finished_product_code):
        func = product_live()
        data_list = func.one_product_live(product_id, finished_product_code)


        return data_list

    def get(self, request):
        try:
            # data_list = []
            data_add_int = {}

            product_id = request.session.get("session_projectId")
            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            finished_product_code = request.GET.get("finished_product_code")


            # product_id = "co2"
            # page = 1
            # page_size = 100
            #
            # finished_product_code = "SPR001300-3A001348-CMSS-073019-0128"



            data_list = self.get_live_data(product_id, finished_product_code)


            data_list.sort(key=lambda x: (x['time']), reverse=False)

            page_result = Page(page, page_size, data_list)
            data = page_result.get_str_json()
            dfs = int(len(data_list))
            sql_num_int = dfs
            if dfs:
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            logger.info("??????????????????:")
            return HttpResponse(result)

        except Exception as e:
            data = {"code": 1, "message": "??????????????????%s" % e}
            logger.error("??????????????????%s" % e)
            data = json.dumps(data)
            print(e)
            return HttpResponse(data)


            # conn = db.get_connection(product_id)
            # db = DB()



            # if finished_product_code:
            #     finished_product_code_sql = " and finished_product_code = " + "'" + finished_product_code + "'"
            #     sn_sql = " and SN = " + "'" + finished_product_code + "'"
            # else:
            #     finished_product_code_sql = ""
            #     sn_sql = ""
            #
            # sql = """
            #        select * from work_station where work_type = '??????'
            #        """
            # drs = db.execute_sql(conn, sql)
            # jiagong_list = []
            # if drs:
            #     for dr in drs:
            #         jiagong_list.append(dr[1])
            #     for wor_id in jiagong_list:
            #         one_station = "product_transit_" + str(wor_id)
            #         sql_select = """
            #         select * from {0} where 2 > 1 {1} group by finished_product_code
            #         """
            #         sql_select = sql_select.format(one_station, finished_product_code_sql)
            #         dr_oneworks = db.execute_sql(conn, sql_select)
            #         if dr_oneworks:
            #             for dr_onework in dr_oneworks:
            #                 dict_one_data = {}
            #                 dict_one_data["finished_product_code"] = dr_onework[3]
            #                 dict_one_data["user_code"] = dr_onework[4]
            #                 dict_one_data["work_code"] = dr_onework[5]
            #                 dict_one_data["test_result"] = dr_onework[6]
            #                 if dr_onework[9]:
            #                     dp = dr_onework[9]
            #                     dp = dp.strftime("%Y-%m-%d %H:%M:%S")
            #                     dict_one_data["time"] = dp
            #                 else:
            #                     dict_one_data["time"] = ""
            #                 data_list.append(dict_one_data)
            #         else:
            #             pass
            #
            #     sql_other_table = """
            #     SELECT * FROM work_station left join work_transit on work_station.work_code = work_transit.work_code
            #     ;
            #     """
            #     # sql_other_table = sql_other_table.format(sn_sql)
            #     df_tbales = db.execute_sql(conn, sql_other_table)
            #     if df_tbales:
            #         for df_tab in df_tbales:
            #             if df_tab[7]:
            #                 sql_table = """
            #                 select SN, Result, time from {0} where 2 > 1 {1}
            #                 """
            #                 sql_table = sql_table.format(df_tab[7], sn_sql)
            #                 dffs = db.execute_sql(conn, sql_table)
            #                 if dffs:
            #                     for df in dffs:
            #                         da_dict = {}
            #                         da_dict["finished_product_code"] = df[0]
            #                         da_dict["test_result"] = df[1]
            #                         if df[2]:
            #                             dp = df[2]
            #                             dp = dp.strftime("%Y-%m-%d %H:%M:%S")
            #                             da_dict["time"] = dp
            #                         else:
            #                             da_dict["time"] = ""
            #                         da_dict["work_code"] = df_tab[0]
            #                         da_dict["user_code"] = ""
            #                         data_list.append(da_dict)
            #                 else:
            #                     pass
            #     else:
            #         pass
            #
            #
            #
            #     data_list.sort(key=lambda x: (x['time']), reverse=False)
            #     page_result = Page(page, page_size, data_list)
            #     data = page_result.get_str_json()
            #     dfs = int(len(data_list))
            #     sql_num_int = dfs
            #     data_add_int["data"] = data
            #     data_add_int["total"] = sql_num_int
            #     result = {"code": 0, "message": "????????????", "data": data_add_int}
            #
            #     result = json.dumps(result)
            #     logger.info("???????????????????????????????????????:")
            #     return HttpResponse(result)
            # else:
            #     result = {"code": 0, "message": "????????????????????????"}
            #     result = json.dumps(result)
            #     return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("?????????????????????????????????")
            return HttpResponse(result)


class NopageWorkUnqualifiedResultName(View):
    """
    ???????????????????????? ?????????????????????????????????

    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            # conn = db.get_connection("db_common")
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            work_id = request.GET.get("work_id")

            if work_id:
                work_id_sql = " and work_id = " + "'" + work_id + "'"
            else:
                work_id_sql = ""
            sql = """
            select * from work_unqualified_result where 2>1 {0}
            """
            sql = sql.format(work_id_sql)
            drs = db.execute_sql(conn, sql)
            if drs:
                for dr in drs:
                    dict_data = {}
                    dict_data["unqualified_result_code"] = dr[0]
                    dict_data["unqualified_result_name"] = dr[2]
                    data_list.append(dict_data)

                data_add_int["data"] = data_list

                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("???????????????????????? ???????????????????????????????????????%s")
                db.close_connection(conn)
                return HttpResponse(result)

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                result = json.dumps(result)
                db.close_connection(conn)
                return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("???????????????????????? ???????????????????????????????????????")
            return HttpResponse(result)


class ProductCodeResponseResult(View):
    """
    ??????????????????????????? ???????????????????????????

    """
    def get(self, request):
        try:
            data_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")

            conn = db.get_connection(product_id)
            finished_product_code = request.GET.get("finished_product_code")
            # finished_product_code = "zz"

            if finished_product_code:
                finished_product_code_sql = " and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""

            sql_tables = """
            show tables like "unqualified_product_%" 
            """
            un_tables = db.execute_sql(conn, sql_tables)

            if un_tables:
                for un_table in un_tables:
                    sql_search = """
                    select finished_product_code, description from {0} where description like "%F0000%" {1} group by finished_product_code
                    """
                    sql_search = sql_search.format(un_table[0], finished_product_code_sql)
                    dfs = db.execute_sql(conn, sql_search)
                    if dfs:
                        for df in dfs:
                            dict_data = {}
                            dict_data["finished_product_code"] = df[0]
                            dict_data["description"] = df[1]
                            data_list.append(dict_data)
                    else:
                        pass

                if data_list:

                    result = {"code": 0, "message": "????????????", "data": data_list}
                    result = json.dumps(result)
                    logger.info("??????????????????????????? ?????????????????????????????????%s")
                    db.close_connection(conn)
                    return HttpResponse(result)
                else:
                    result = {"code": 0, "message": "??????????????????????????????", "data": data_list}
                    result = json.dumps(result)
                    logger.info("??????????????????????????? ?????????????????????????????????, ??????????????????????????????%s" )
                    db.close_connection(conn)
                    return HttpResponse(result)
            else:
                result = {"code": 0, "message": "????????????????????????", "data":data_list}
                result = json.dumps(result)
                logger.error("??????????????????????????? ???????????????????????????, ????????????????????????")
                db.close_connection(conn)
                return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("??????????????????????????? ?????????????????????????????????%s"% e)
            return HttpResponse(result)



class ModifyInWorkStation(View):

    def get_live_data(self, product_id, finished_product_code):
        func = product_live()
        data_list = func.one_product_live(product_id, finished_product_code)

        return data_list

    def get(self, request):
        try:
            print("--->", "0o0o0okkkk")
            data_list = []
            data_add_int = {}
            db = DB()
            # conn = db.get_connection("db_common")
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            finished_product_code = request.GET.get("finished_product_code")
            product_plan_code = request.GET.get("product_plan_code")
            status = request.GET.get("status")

            if finished_product_code:
                finished_product_code_sql = " and finished_product_code = " + "'" + finished_product_code + "'"
            else:
                finished_product_code_sql = ""
            if product_plan_code:
                product_plan_code_sql = " and product_plan_code = " + "'" + product_plan_code + "'"
            else:
                product_plan_code_sql = ""
            if status:
                status_sql = " and status = " + "'" + status + "'"
            else:
                status_sql = ""

            sql = """
                   select * from modify_in_workstation where 2>1 {0} {1} {2}
                   """

            sql = sql.format(finished_product_code_sql, product_plan_code_sql, status_sql)
            drs = db.execute_sql(conn, sql)
            print("------>", drs)
            if len(drs) > 0:
                for dr in drs:
                    dict_data = {}
                    dict_data["modify_in_workstation_code"] = dr[0]
                    dict_data["product_name"] = dr[1]
                    dict_data["finished_product_code"] = dr[2]
                    dict_data["work_id"] = dr[3]
                    dict_data["product_plan_code"] = dr[4]
                    dict_data["operate_user"] = dr[5]
                    if dr[6]:
                        dp = dr[6]
                        dp = dp.strftime("%Y-%m-%d %H:%M:%S")
                        dict_data["enter_time"] = dp
                    else:
                        dict_data["enter_time"] = ""
                    dict_data["process_method"] = dr[7]
                    if dr[8]:
                        dsss = dr[8]
                        print("====>", dsss)
                        dsss = dsss.strftime("%Y-%m-%d %H:%M:%S")
                        dict_data["out_time"] = dsss
                    else:
                        dict_data["out_time"] = ""
                    dict_data["status"] = dr[9]
                    dict_data["enter_work_code"] = dr[10]
                    data_list.append(dict_data)
                data_list.sort(key=lambda x: (x['enter_time']), reverse=True)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                dfs = int(len(data_list))
                sql_num_int = dfs
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int

                result = {"code": 0, "message": "????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("?????????????????????????????????%s" )
                db.close_connection(conn)
                return HttpResponse(result)

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0

                result = {"code": 0, "message": "??????????????????", "data": data_add_int}
                result = json.dumps(result)
                db.close_connection(conn)
                return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("?????????????????????????????????%s"% e)
            return HttpResponse(result)

    def post(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            json_data = request.body

            str_data = json.loads(json_data)

            finished_product_code = str_data.get("finished_product_code")
            product_plan_code = str_data.get("product_plan_code")
            # enter_work_code = str_data.get("enter_work_code")  # ZIAAAAAAAAAAAAAAAA
            work_id = request.session.get('session_workId')

            data_list = self.get_live_data(product_id, finished_product_code)
            data_list.sort(key=lambda x: (x['time']), reverse=False)
            if data_list:
                enter_work_code = data_list[-1]["work_code"]
            else:

                data = {"code": 1, "message": "?????????????????????????????????,????????????????????????"}
                logger.info("?????????????????????????????????,????????????????????????")
                data = json.dumps(data)
                return HttpResponse(data)

            data_list = []
            data_add_int = {}

            sql_work_code = """
            select work_code from work_station where work_id = '{0}'
            """
            sql_work_code = sql_work_code.format(work_id)
            drs = db.execute_sql(conn, sql_work_code)
            if drs:
                work_id = drs[0][0]
            else:
                work_id = work_id

            operate_user = request.session.get('session_currentId')  # ???????????????
            status = str_data.get("status")

            sql_filed = """
                       SELECT max(order_number) FROM modify_in_workstation;
                       """
            matter_type = db.execute_sql(conn, sql_filed)
            if matter_type[0][0] == None:
                field_num = 1
            else:
                field_num = matter_type[0][0] + 1
            modify_in_workstation_code = "Im_Modi_In_Sta_" + str(field_num)

            sql_check = """
            select status from modify_in_workstation where finished_product_code = '{0}'
            """
            sql_check = sql_check.format(finished_product_code)
            drs_checks = db.execute_sql(conn, sql_check)


            if drs_checks:
                da_list = []
                for dr_check in drs_checks:
                    da_list.append(dr_check[0])

                if "??????" in da_list:
                    result = {"code": 1, "message": "?????????????????????????????????,??????????????????"}
                    result = json.dumps(result)
                    logger.error("?????????????????????????????????,??????????????????%s")
                    return HttpResponse(result)
                else:
                    sql_insert_into = """
                            insert into modify_in_workstation(modify_in_workstation_code, product_name, 
                                                            finished_product_code,
                                                            work_id,
                                                            product_plan_code, operate_user,
                                                             status, enter_work_code, order_number)
                                                            values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}',
                                                            '{7}', '{8}')
                                                """
                    sql_insert_into = sql_insert_into.format(modify_in_workstation_code, product_id,
                                                             finished_product_code,
                                                             work_id, product_plan_code,
                                                             operate_user, status, enter_work_code, field_num)
                    drs = db.execute_sql(conn, sql_insert_into)

                    data = {"code": 0, "message": "??????????????????"}
                    logger.info("??????????????????")
                    data = json.dumps(data)
                    db.close_connection(conn)

                    return HttpResponse(data)

            else:
                sql_insert_into = """
                            insert into modify_in_workstation(modify_in_workstation_code, product_name, 
                                                            finished_product_code,
                                                            work_id,
                                                            product_plan_code, operate_user,
                                                            status, enter_work_code, order_number)
                                                            values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}',
                                                            '{7}', '{8}')
                            """
                sql_insert_into = sql_insert_into.format(modify_in_workstation_code, product_id, finished_product_code,
                                                         work_id, product_plan_code,
                                                         operate_user,status, enter_work_code, field_num)

                drs = db.execute_sql(conn, sql_insert_into)

                data = {"code": 0, "message": "??????????????????"}
                logger.info("??????????????????")
                data = json.dumps(data)
                db.close_connection(conn)

                return HttpResponse(data)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("??????????????????%s" % e)
            return HttpResponse(result)


class AnalyseWorkStationInfo(View):
    """
    ????????????????????????????????????????????????????????????

    """
    # import datetime
    def getEveryDay(self, begin_date, end_date):
        # ????????????
        date_list = []
        begin_date = datetime.datetime.strptime(begin_date, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        while begin_date <= end_date:
            date_str = begin_date.strftime("%Y-%m-%d")
            date_list.append(date_str)
            begin_date += datetime.timedelta(days=1)
        return date_list


    def get(self, request):
        try:

            data_type_list = []
            data_list = []
            data_add_int = {}
            db = DB()

            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            start_time = request.GET.get("start_time")
            end_time = request.GET.get("end_time")


            # product_id = "co2"
            # conn = db.get_connection(product_id)
            # start_time = "2021-09-03 00:00:00"
            # end_time = "2021-09-03 23:59:59"
            check_start_time = start_time[0:10]
            check_end_time = end_time[0:10]

            sql_type = """
            SELECT work_type FROM work_station group by work_type
            """
            drs_types = db.execute_sql(conn, sql_type)
            if len(drs_types) > 0:
                for drs_type in drs_types:
                    if drs_type[0]:
                        data_type_list.append(drs_type[0])

                for work_type in data_type_list:
                    work_type_sql = " and work_type = " + "'" + work_type + "'"
                    if work_type == "??????":
                        sql = """
                               select * from work_station where 2>1 {0}
                               """
                        sql = sql.format(work_type_sql)
                        drs = db.execute_sql(conn, sql)
                        jiagong_list = []
                        data_type_dict = {}
                        if len(drs) > 0:
                            for dr in drs:
                                num_data = {}
                                dict_data = {}
                                dict_data["work_id"] = dr[1]
                                dict_data["work_name"] = dr[3]
                                dict_data["work_type"] = dr[4]
                                select_table = "product_transit_" + str(dr[1])
                                # if check_start_time != check_end_time:
                                sql_pass = """
                                select count(*) from {0} where test_result = 'PASS' 
                                and out_time >= "{1}" and out_time < "{2}"
                                """
                                sql_pass = sql_pass.format(select_table, start_time, end_time)
                                pass_nums = db.execute_sql(conn, sql_pass)
                                if pass_nums:
                                    pass_num = pass_nums[0][0]
                                else:
                                    pass_num = 0

                                sql_fail = """
                                select count(*) from {0}
                                where test_result = 'FAIL' and out_time >= "{1}" and out_time < "{2}"
                                """
                                sql_fail = sql_fail.format(select_table, start_time, end_time)
                                fail_nums = db.execute_sql(conn, sql_fail)
                                if fail_nums:
                                    fail_num = fail_nums[0][0]
                                else:
                                    fail_num = 0

                                num_data["pass_num"] = pass_num
                                num_data["fail_num"] = fail_num
                                dict_data["data"] = num_data
                                jiagong_list.append(dict_data)
                            data_type_dict["work_type"] = work_type
                            data_type_dict["data"] = jiagong_list
                            data_list.append(data_type_dict)

                        else:
                            data_type_dict["work_type"] = work_type
                            data_type_dict["data"] = jiagong_list
                            data_list.append(data_type_dict)

                    elif work_type == "??????":
                        ceshi_list = []
                        ceshi_dict = {}

                        sql_ce = """
                        SELECT * FROM work_station left join 
                        work_transit on work_station.work_code = work_transit.work_code where work_type = '??????'
                        """
                        drs_tables = db.execute_sql(conn, sql_ce)
                        if len(drs_tables) > 0:
                            for dr_table in drs_tables:
                                num_data = {}
                                dict_data = {}
                                dict_data["work_id"] = dr_table[1]
                                dict_data["work_name"] = dr_table[3]
                                dict_data["work_type"] = dr_table[4]
                                select_table = dr_table[7]
                                sql_pass = """
                                          select count(*) from {0} 
                                          where Result = 'PASS' and time >= "{1}" and time < "{2}"
                                          """
                                sql_pass = sql_pass.format(select_table, start_time, end_time)
                                pass_nums = db.execute_sql(conn, sql_pass)
                                if pass_nums:
                                    pass_num = pass_nums[0][0]
                                else:
                                    pass_num = 0

                                sql_fail = """
                                           select count(*) from {0}
                                            where Result = 'FAIL' and time >= "{1}" and time < "{2}"
                                          """
                                sql_fail = sql_fail.format(select_table, start_time, end_time)
                                fail_nums = db.execute_sql(conn, sql_fail)
                                if fail_nums:
                                    fail_num = fail_nums[0][0]
                                else:
                                    fail_num = 0

                                num_data["pass_num"] = pass_num
                                num_data["fail_num"] = fail_num
                                dict_data["data"] = num_data
                                ceshi_list.append(dict_data)
                            ceshi_dict["work_type"] = work_type
                            ceshi_dict["data"] = ceshi_list
                            data_list.append(ceshi_dict)

                        else:
                            ceshi_dict["work_type"] = work_type
                            ceshi_dict["data"] = ceshi_list
                            data_list.append(ceshi_dict)

                    elif work_type == "??????":
                        sql = """
                               select * from work_station where 2>1 {0}
                               """
                        sql = sql.format(work_type_sql)
                        drs = db.execute_sql(conn, sql)
                        weixiu_list = []
                        weixiu_dict = {}
                        if len(drs) > 0:
                            for dr in drs:
                                num_data = {}
                                dict_data = {}
                                dict_data["work_id"] = dr[1]
                                dict_data["work_name"] = dr[3]
                                dict_data["work_type"] = dr[4]
                                select_table = "modify_in_workstation"
                                sql_pass = """
                                select count(*) from {0} 
                                where process_method = '????????????' and out_time >= "{1}" and out_time <= "{2}"
                                """
                                sql_pass = sql_pass.format(select_table, start_time, end_time)
                                modify_nums = db.execute_sql(conn, sql_pass)
                                if modify_nums:
                                    if modify_nums[0][0]:
                                        maintain_num = modify_nums[0][0]
                                    else:
                                        maintain_num = 0
                                else:
                                    maintain_num = 0

                                sql_disassemble = """
                                select count(*) from {0} 
                                where process_method = '????????????' and out_time >= "{1}" and out_time <= "{2}"
                                """
                                sql_disassemble = sql_disassemble.format(select_table, start_time, end_time)
                                disassemble_nums = db.execute_sql(conn, sql_disassemble)
                                if disassemble_nums:
                                    if disassemble_nums[0][0]:
                                        disassemble_num = disassemble_nums[0][0]
                                    else:
                                        disassemble_num = 0
                                else:
                                    disassemble_num = 0

                                sql_disassemble = """
                                select count(*) from {0} 
                                where 2>1 and out_time >= "{1}" and out_time <= "{2}"
                                """
                                sql_disassemble = sql_disassemble.format(select_table, start_time, end_time)
                                no_status_nums = db.execute_sql(conn, sql_disassemble)
                                if no_status_nums:
                                    if no_status_nums[0][0]:
                                        no_status_num = disassemble_nums[0][0]
                                    else:
                                        no_status_num = 0
                                else:
                                    no_status_num = 0

                                in_num = int(no_status_num) - int(disassemble_num) - int(maintain_num)

                                num_data["maintain_num"] = maintain_num
                                num_data["out_num"] = disassemble_num
                                num_data["in_num"] = in_num
                                dict_data["data"] = num_data
                                weixiu_list.append(dict_data)
                            weixiu_dict["work_type"] = work_type
                            weixiu_dict["data"] = weixiu_list
                            data_list.append(weixiu_dict)

                        else:
                            weixiu_dict["work_type"] = work_type
                            weixiu_dict["data"] = weixiu_list
                            data_list.append(weixiu_dict)

                    elif work_type == "??????":
                        sql = """
                               select * from work_station where 2>1 {0}
                               """
                        sql = sql.format(work_type_sql)
                        drs = db.execute_sql(conn, sql)
                        pakeage_list = []
                        pakeage_dict = {}
                        xAxis_data = []
                        series_data = []
                        if len(drs) > 0:
                            for dr in drs:
                                num_data = {}
                                dict_data = {}
                                dict_data["work_id"] = dr[1]
                                dict_data["work_name"] = dr[3]
                                dict_data["work_type"] = dr[4]
                                select_table = "enter_storage_status"
                                # start_time = start_time[0:10]
                                # end_time = end_time[0:10]
                                xAxis_data = self.getEveryDay(check_start_time, check_end_time)
                                if len(xAxis_data) > 0:
                                    for date_one in xAxis_data:
                                        # sql_pass = """
                                        # select count(*) from {0} where enter_time like "{1} %"
                                        # """
                                        sql_pass = """
                                        SELECT count(*) FROM enter_storage as a 
                                        left join enter_storage_status as b on a.pack_id = b.pack_id 
                                        where enter_time like "{0}%"
                                        """
                                        sql_pass = sql_pass.format(date_one)
                                        count_nums = db.execute_sql(conn, sql_pass)
                                        if len(count_nums) > 0:
                                            if count_nums[0][0]:
                                                count_num = count_nums[0][0]
                                            else:
                                                count_num = 0
                                        else:
                                            count_num = 0
                                        series_data.append(count_num)
                                num_data["xAxis_data"] = xAxis_data
                                num_data["series_data"] = series_data
                                dict_data["data"] = num_data
                                pakeage_list.append(dict_data)

                            pakeage_dict["work_type"] = work_type
                            pakeage_dict["data"] = pakeage_list
                            data_list.append(pakeage_dict)

                        else:
                            pakeage_dict["work_type"] = work_type
                            pakeage_dict["data"] = pakeage_list
                            data_list.append(pakeage_dict)

                result = {"code": 0, "message": "??????????????????", "data" : data_list}
                result = json.dumps(result)
                logger.info("??????????????????%s")
                return HttpResponse(result)

            else:
                result = {"code": 0, "message": "work_station??????????????????????????????????????????", "data": data_list}
                result = json.dumps(result)
                logger.error("work_station??????????????????????????????????????????%s")
                return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("??????????????????????????????????????????????????????????????????%s" % e)
            return HttpResponse(result)


class FromDataModel(View):

    def deletefile(self, request, day_qty):

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))).replace('\\', '/')
        path = os.path.join(BASE_DIR, 'media').replace('\\', '/')


        N = day_qty  # ?????????????????????????????????

        for eachfile in os.listdir(path):
            filename = os.path.join(path, eachfile)
            if os.path.isfile(filename):
                lastmodifytime = os.stat(filename).st_mtime
                endfiletime = time.time() - 3600 * 24 * N # ?????????????????????????????????
                if endfiletime > lastmodifytime:
                    os.remove(filename)
            # elif os.path.isdir(filename):  # ??????????????????????????????????????????
            #     self.deletefile(request, day_qty)

    def co2_report(self, request):
        data_fct = []
        data_leak = []
        data_pack = []
        data_list = []
        data_add_int = {}
        db = DB()
        # conn = db.get_connection("db_common")

        print("==----->")
        product_id = request.session.get("session_projectId")
        conn = db.get_connection(product_id)
        start_time = request.GET.get("start_time")
        end_time = request.GET.get("end_time")
        sql_fct = """
                select ID, finished_product_code,
                 PRV_low, PRV_high, Bleeding_result, Detect_switch,
                  Result, time from fct where time > '{0}' and time < '{1}' group by finished_product_code
                """
        sql_fct = sql_fct.format(start_time, end_time)
        fct_datas = db.execute_sql(conn, sql_fct)
        for fct_dada in fct_datas:
            data_fct.append(list(fct_dada))

        sql_leak = """
                                            select ID, finished_product_code, Test1_level,
                                            Test1_ppm, Test2_level, Test2_ppm, Test3_level,
                                            Test3_ppm, Result, time from leakage_test_new
                                            where time > '{0}' and time < '{1}' group by finished_product_code
                                            """
        sql_leak = sql_leak.format(start_time, end_time)
        leak_datas = db.execute_sql(conn, sql_leak)
        for leak_data in leak_datas:
            data_leak.append(list(leak_data))

        sql_pake = """
              SELECT a.enter_storage_code, a.finished_product_code, a.pack_id, c.Itemcode_C_Shipping,b.status, b.enter_time
              FROM enter_storage a ,enter_storage_status b,operate c 
              where a.pack_id = b.pack_id and b.enter_time > '{0}' and b.enter_time < '{1}' group by finished_product_code
              """
        sql_pake = sql_pake.format(start_time, end_time)

        print(sql_pake)

        # sql_pake = db.execute_sql(conn, sql_pake)

        pake_datas = db.execute_sql(conn, sql_pake)
        print("--->")
        for pake_data in pake_datas:
            data_pack.append(list(pake_data))
        db.close_connection(conn)

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        BASE_LOG_DIR = os.path.join(BASE_DIR, "static\models\CO2.xlsx")
        BASE_LOG_DIR = BASE_LOG_DIR.replace("\\", '/')

        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        pathattr = now[:4] + now[5:7] + now[8:10] + now[11:13] + now[14:16] + now[17:19]
        base_media = os.path.join(BASE_DIR, "media/CO2_" + pathattr + ".xlsx").replace("\\", '/')

        shutil.copy(BASE_LOG_DIR, base_media)

        wb = load_workbook(base_media)

        if len(data_fct) > 0:
            ws = wb["FCT"]
            i = 1
            for data_in in data_fct:
                ws.cell(row=i + 2, column=1).value = str(i)
                ws.cell(row=i + 2, column=2).value = data_in[1]
                ws.cell(row=i + 2, column=3).value = data_in[2]
                ws.cell(row=i + 2, column=4).value = data_in[3]
                ws.cell(row=i + 2, column=5).value = data_in[4]
                ws.cell(row=i + 2, column=6).value = data_in[5]
                ws.cell(row=i + 2, column=7).value = data_in[6]
                ws.cell(row=i + 2, column=8).value = data_in[7].strftime("%Y-%m-%d %H:%M:%S")
                i += 1

        if len(data_leak) > 0:
            ws = wb["Leakage_Test"]
            i = 1
            for data_in in data_leak:
                ws.cell(row=i + 2, column=1).value = str(i)
                ws.cell(row=i + 2, column=2).value = data_in[1]
                ws.cell(row=i + 2, column=3).value = data_in[2]
                ws.cell(row=i + 2, column=4).value = data_in[3]
                ws.cell(row=i + 2, column=5).value = data_in[4]
                ws.cell(row=i + 2, column=6).value = data_in[5]
                ws.cell(row=i + 2, column=7).value = data_in[6]
                ws.cell(row=i + 2, column=8).value = data_in[7]
                ws.cell(row=i + 2, column=9).value = data_in[8]
                ws.cell(row=i + 2, column=10).value = data_in[9].strftime("%Y-%m-%d %H:%M:%S")
                i += 1

        if len(data_pack) > 0:
            ws = wb["Package"]
            i = 1
            for data_in in data_pack:
                ws.cell(row=i + 2, column=1).value = str(i)
                ws.cell(row=i + 2, column=2).value = data_in[1]
                ws.cell(row=i + 2, column=3).value = data_in[2]
                ws.cell(row=i + 2, column=4).value = data_in[3]
                ws.cell(row=i + 2, column=5).value = data_in[4]
                ws.cell(row=i + 2, column=6).value = data_in[5].strftime("%Y-%m-%d %H:%M:%S")
                i += 1

        wb.save(base_media)
        wb.close

        return base_media
    def get(self, request):
        try:

            self.deletefile(request, 1)
            file_path = ""
            product_id = request.session.get("session_projectId")
            if product_id == "co2":
                file_path = self.co2_report(request)

            result = {"code": 0, "message": "??????????????????", "file_path": file_path}
            result = json.dumps(result)
            return HttpResponse(result)

        except Exception as e:
            print(e)
            result = {"code": 1, "message": "????????????????????????"}
            result = json.dumps(result)
            logger.error("????????????????????????")
            return HttpResponse(result)


class downloads(View):

    def download(self, request):  # ????????????
        filespath = request.GET.get('files')
        response = self.get_files(request, filespath)
        return response

    def get_files(self, request, file_path):
        file_name = file_path[file_path.rindex('/') + 1:]

        # ???????????????
        file = open(file_path, 'rb')
        response = FileResponse(file)
        response['content_type'] = "application/octet-stream"

        from django.utils.http import urlquote  # ???????????????????????????
        response['Content-Disposition'] = 'attachment;filename="%s"' % (urlquote(file_name))
        return response

    def get(self, request):
        return self.download(request)


class FinishedProductOut(View):
    def get(self, request):
        try:
            db = DB()
            package_sequence = request.GET.get("package_sequence")
            pack_id = request.GET.get("pack_id")
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            current_person_id = request.session.get('session_currentId')

            print("-->", pack_id)
            print("-->", product_id)
            print("-->", package_sequence)
            print("-->", current_person_id)

            sql_check = """           
            select * from finished_product_out where pack_id = '{0}'           
            """
            sql_check = sql_check.format(pack_id)
            dr_checks = db.execute_sql(conn, sql_check)
            if len(dr_checks) > 0:
                data = {"code": 1, "message": "????????????????????????,???????????????"}
                data = json.dumps(data)

                return HttpResponse(data)
            else:
                sql_num = """
                select * from enter_storage where pack_id = '{0}'
                """
                sql_num = sql_num.format(pack_id)
                sql_nums = db.execute_sql(conn, sql_num)

                if sql_nums:
                    part_qty = int(len(sql_nums))
                else:
                    part_qty = 0

                sql_data = """
                select * from operate 
                """
                # sql_data = sql_data.format(pack_id)
                sql_datas = db.execute_sql(conn, sql_data)

                print("======>", sql_datas)
                db.close_connection(conn)
                for data_pake in sql_datas:
                    flex_part_number = data_pake[3]

                    vendor_code = data_pake[4]
                    part_revision = data_pake[2]
                    manufacturing_date = datetime.datetime.now().strftime("%Y%m%d")

                print(flex_part_number)
                print(vendor_code)
                print(part_revision)
                print(manufacturing_date)

                # re_data = """
                #
                # ^XA
                #
                # ^PW1300
                # ^LL0827
                #
                # ^LH0,60
                #
                # ^FO480,50
                # ^AS,30,30
                # ^FDFlex Part Number^FS
                # ^FO65,90
                # ^BY3
                # ^BCN,100,Y,N,N
                # ^FD{0}^FS
                #
                # ^FO150,270
                # ^AS,30,30
                # ^FDVendor Code^FS
                # ^FO20,310
                # ^BCN,100,Y,N,N
                # ^FD{1}^FS
                #
                # ^FO525,270
                # ^AS,30,30
                # ^FDPart QTY/P^FS
                # ^FO465,310
                # ^BCN,100,Y,N,N
                # ^FD {2}^FS
                #
                # ^FO815,270
                # ^AS,30,30
                # ^FDPackage Sequence^FS
                # ^FO740,310
                # ^BCN,100,Y,N,N
                # ^FD  {3}^FS
                #
                # ^FO275,480
                # ^AS,30,30
                # ^FDPart Revision^FS
                # ^FO200,520
                # ^BCN,100,Y,N,N
                # ^FD{4}^FS
                #
                # ^FO640,480
                # ^AS,30,30
                # ^FDManufacturing Date^FS
                # ^FO583,520
                # ^BCN,100,Y,N,N
                # ^FD{5}^FS
                #
                # ^XZ
                # """

                re_data = """
                ^XA

                ^PW1300
                ^LL0827
                
                ^LH0,60
                
                ^FO480,50
                ^AS,30,30
                ^FDFlex Part Number^FS
                ^FO100,90
                ^BY3
                ^BCN,100,Y,N,N
                ^FD{0}^FS
                
                ^FO190,270
                ^AS,30,30
                ^FDVendor Code^FS
                ^FO60,310
                ^BCN,100,Y,N,N
                ^FD{1}^FS
                
                ^FO545,270
                ^AS,30,30
                ^FDPart QTY/P^FS
                ^FO505,310
                ^BCN,100,Y,N,N
                ^FD  {2}^FS
                
                ^FO825,270
                ^AS,30,30
                ^FDPackage Sequence^FS
                ^FO780,310
                ^BCN,100,Y,N,N
                ^FD  {3}^FS
                
                ^FO315,480
                ^AS,30,30
                ^FDPart Revision^FS
                ^FO240,520
                ^BCN,100,Y,N,N
                ^FD{4}^FS
                
                ^FO680,480
                ^AS,30,30
                ^FDManufacturing Date^FS
                ^FO623,520
                ^BCN,100,Y,N,N
                ^FD{5}^FS
                
                ^XZ 
                """

                re_data = re_data.format(flex_part_number, vendor_code, part_qty,
                                         package_sequence, part_revision, manufacturing_date)

                reponse_data = str(re_data)

                conn = db.get_connection(product_id)

                sql_filed = """
                       SELECT max(order_number) FROM finished_product_out
                       """
                matter_type = db.execute_sql(conn, sql_filed)

                if matter_type[0][0] == None:
                    field_num = 1
                else:
                    field_num = matter_type[0][0] + 1
                product_out_code = "Im_Pro_Out_" + str(field_num)

                sql_insert = """
                insert into finished_product_out(product_out_code, pack_id,
                                                                flex_part_number,
                                                                vendor_code,
                                                                part_qty, package_sequence,
                                                                part_revision, manufacturing_date, operate_user, 
                                                                order_number)
                                                                values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}', '{6}',
                                                                '{7}', '{8}','{9}' )
                """
                sql_insert = sql_insert.format(product_out_code, pack_id, flex_part_number,
                                               vendor_code, part_qty, package_sequence,
                                               part_revision, manufacturing_date, current_person_id, field_num)
                db.execute_sql(conn, sql_insert)

                data = {"code": 0, "message": "????????????", "data": reponse_data}

                data = json.dumps(data)

                return HttpResponse(data)

        except Exception as e:

            result = {"code": 1, "message": "??????????????????"}
            result = json.dumps(result)
            logger.error("??????????????????")
            return HttpResponse(result)


class GetProductOutInfo(View):
    def get(self, request):
        try:
            db = DB()
            data_list = []
            data_add_int = {}
            package_sequence = request.GET.get("package_sequence")
            pack_id = request.GET.get("pack_id")
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            pack_id = request.GET.get("pack_id")
            start_time = request.GET.get("start_time")
            end_time = request.GET.get("end_time")

            if start_time:
                start_time_sql = "and manufacturing_date >= " + "'" + start_time + "'"
            else:
                start_time_sql = ""

            if end_time:
                end_time_sql = "and manufacturing_date < " + "'" + end_time + "'"
            else:
                end_time_sql = ""

            sql = """
                 select * from finished_product_out where 2 > 1 {0} {1}
                """
            sql_main = sql.format(start_time_sql, end_time_sql)

            drs = db.execute_sql(conn, sql_main)

            if len(drs) > 0:
                for dr in drs:
                    dict_data = {}
                    dict_data["product_out_code"] = dr[0]
                    dict_data["pack_id"] = dr[1]
                    dict_data["flex_part_number"] = dr[2]
                    dict_data["vendor_code"] = dr[3]
                    dict_data["part_qty"] = dr[4]
                    dict_data["package_sequence"] = dr[5]
                    dict_data["part_revision"] = dr[6]
                    if dr[7]:
                        dp = dr[7]
                        dp = dp.strftime("%Y-%m-%d")
                    else:
                        dp = ""
                    dict_data["manufacturing_date"] = dp
                    dict_data["operate_user"] = dr[8]
                    dict_data["order_number"] = dr[9]
                    data_list.append(dict_data)
                data_list.sort(key=lambda x: (x['order_number']), reverse=True)
                page_result = Page(page, page_size, data_list)
                data = page_result.get_str_json()
                sql_num_int = int(len(data_list))
                data_add_int["data"] = data
                data_add_int["total"] = sql_num_int
                data_add_int["product_num"] = int(len(drs))
            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                data_add_int["product_num"] = 0

            result = {"code": 0, "message": "????????????", "data": data_add_int}
            result = json.dumps(result)
            db.close_connection(conn)

            return HttpResponse(result)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????")

            return HttpResponse(data)


class OneProductLiveDeal(View):
    def get(self, request):
        try:
            db = DB()
            data_list = []

            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)
            finished_product_code = request.GET.get("finished_product_code")
            work_code = request.GET.get("work_code")

            # product_id = "co2"
            # conn = db.get_connection(product_id)
            # finished_product_code = "SPR001300-2A001387-CMSS-A02-090421-0226"
            # work_code = "Im_WorkStation_6"

            sql_search_id = """
            select work_id, work_type from work_station where work_code = '{0}'
            """
            sql_search_id = sql_search_id.format(work_code)
            dr_types = db.execute_sql(conn, sql_search_id)
            if len(dr_types) > 0:
                if dr_types[0][1] == "??????":
                    table_one = "product_transit_" + str(dr_types[0][0])

                    sql_info = """
                    select * from {0} where finished_product_code = '{1}'
                    """
                    sql_info = sql_info.format(table_one, finished_product_code)
                    dr_infos = db.execute_sql(conn, sql_info)
                    if len(dr_infos) > 0:
                        for dr in dr_infos:
                            dict_data = {}
                            dict_data["product_transit_code"] = dr[0]
                            dict_data["matter_code"] = dr[1]
                            dict_data["matter_id"] = dr[2]
                            dict_data["finished_product_code"] = dr[3]
                            dict_data["user_code"] = dr[4]
                            dict_data["work_code"] = dr[5]
                            dict_data["test_result"] = dr[6]
                            dict_data["description"] = dr[7]

                            if dr[8]:
                                dp = dr[8]
                                dp = dp.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                dp = ""
                            dict_data["enter_time"] = dp
                            if dr[9]:
                                dw = dr[9]
                                dw = dw.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                dw = ""
                            dict_data["out_time"] = dw
                            dict_data["product_plan_code"] = dr[10]
                            dict_data["end_product_code"] = dr[11]
                            data_list.append(dict_data)
                        data_list.sort(key=lambda x: (x['out_time']), reverse=True)
                        data = {"code": 0, "message": "????????????", "data": data_list}
                        data = json.dumps(data)
                        return HttpResponse(data)
                    else:
                        data = {"code": 1, "message": "??????????????????????????????,?????????"}
                        data = json.dumps(data)
                        return HttpResponse(data)

                elif dr_types[0][1] == "??????":
                    sql_table = """
                    select table_name from work_transit where work_code = "{0}"
                    """
                    sql_table = sql_table.format(work_code)
                    dr_tables = db.execute_sql(conn, sql_table)

                    if len(dr_tables)> 0:
                        if dr_tables[0][0]:
                            select_table = dr_tables[0][0]
                            sql_column = """
                            select COLUMN_NAME from information_schema.COLUMNS where table_name = '{0}'
                            """
                            sql_column = sql_column.format(select_table)
                            dr_columns = db.execute_sql(conn, sql_column)

                            col_num = len(dr_columns)

                            sql_info = """
                            select * from {0} where finished_product_code = '{1}'
                            """
                            sql_info = sql_info.format(select_table, finished_product_code)
                            dr_infos = db.execute_sql(conn, sql_info)

                            if len(dr_infos) > 0:
                                for dr in dr_infos:
                                    dict_data = {}
                                    for i in range(int(col_num)):

                                        if dr_columns[i][0] == "time":
                                            s = dr[i]
                                            if s:
                                                s = s.strftime("%Y-%m-%d %H:%M:%S")
                                            else:
                                                s = ""
                                            dict_data["time"] = s
                                        else:
                                            dict_data[dr_columns[i][0]] = dr[i]


                                    data_list.append(dict_data)
                                data_list.sort(key=lambda x: (x['time']), reverse=True)
                                data = {"code": 0, "message": "????????????", "data": data_list}
                                data = json.dumps(data)
                                return HttpResponse(data)
                            else:
                                data = {"code": 1, "message": "???????????????????????????????????????,??????????????????"}
                                data = json.dumps(data)
                                return HttpResponse(data)

                        else:

                            data = {"code": 1, "message": "???????????????????????????????????????,??????????????????"}
                            data = json.dumps(data)
                            return HttpResponse(data)

                elif dr_types[0][1] == "??????":
                    table_one = "modify_in_workstation"

                    sql_info = """
                                        select * from {0} where finished_product_code = '{1}'
                                        """
                    sql_info = sql_info.format(table_one, finished_product_code)
                    dr_infos = db.execute_sql(conn, sql_info)
                    if len(dr_infos) > 0:
                        for dr in dr_infos:
                            dict_data = {}
                            dict_data["modify_in_workstation_code"] = dr[0]
                            dict_data["product_name"] = dr[1]
                            dict_data["finished_product_code"] = dr[2]
                            dict_data["work_id"] = dr[3]
                            dict_data["product_plan_code"] = dr[4]
                            dict_data["operate_user"] = dr[5]
                            if dr[6]:
                                dp = dr[6]
                                dp = dp.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                dp = ""
                            dict_data["enter_time"] = dp
                            dict_data["process_method"] = dr[7]
                            if dr[8]:
                                dw = dr[8]
                                dw = dw.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                dw = ""
                            dict_data["out_time"] = dw
                            dict_data["status"] = dr[9]
                            dict_data["enter_work_code"] = dr[10]
                            data_list.append(dict_data)

                        data_list.sort(key=lambda x: (x['out_time']), reverse=True)

                        data = {"code": 0, "message": "????????????", "data": data_list}
                        data = json.dumps(data)
                        return HttpResponse(data)
                    else:
                        data = {"code": 1, "message": "?????????????????????????????????,?????????"}
                        data = json.dumps(data)
                        return HttpResponse(data)

                elif dr_types[0][1] == "??????":
                    # table_one = "modify_in_workstation"

                    sql_info = """
                                SELECT a.pack_id, product_id,product_name,enter_user, enter_time, 
                                status, product_plan_code,finished_product_code  FROM 
                                enter_storage_status as a left join enter_storage as b on a.pack_id = b.pack_id
                                where finished_product_code = "{0}"
                                """
                    sql_info = sql_info.format(finished_product_code)
                    dr_infos = db.execute_sql(conn, sql_info)
                    if len(dr_infos) > 0:
                        for dr in dr_infos:
                            dict_data = {}
                            dict_data["pack_id"] = dr[0]
                            dict_data["product_id"] = dr[1]
                            dict_data["product_name"] = dr[2]
                            dict_data["enter_user"] = dr[3]
                            if dr[4]:
                                dp = dr[4]
                                dp = dp.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                dp = ""
                            dict_data["enter_time"] = dp

                            dict_data["status"] = dr[5]
                            dict_data["product_plan_code"] = dr[6]
                            dict_data["finished_product_code"] = dr[7]
                            data_list.append(dict_data)
                        data_list.sort(key=lambda x: (x['enter_time']), reverse=True)

                        data = {"code": 0, "message": "????????????", "data":data_list}
                        data = json.dumps(data)
                        return HttpResponse(data)
                    else:
                        data = {"code": 1, "message": "?????????????????????????????????,?????????"}
                        data = json.dumps(data)
                        return HttpResponse(data)
            else:
                data = {"code": 1, "message": "????????????????????????????????????????????????"}
                data = json.dumps(data)
                return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????????????????"}
            data = json.dumps(data)
            logger.error("????????????????????????")

            return HttpResponse(data)


class ManagePickMatter(View):
    """
    ??????????????????
    """

    def get(self, request):
        try:
            data_list = []
            matter_list = []
            data_add_int = {}
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            page = int(request.GET.get("page"))
            page_size = int(request.GET.get("page_size"))

            materials_person = request.GET.get("materials_person")
            product_plan_code = request.GET.get("product_plan_code")

            if materials_person:
                materials_person_sql = "and materials_person = " + "'" + materials_person + "'"
            else:
                materials_person_sql = ""
            if product_plan_code:
                product_plan_code_sql = "and product_plan_code = " + "'" + product_plan_code + "'"
            else:
                product_plan_code_sql = ""

            sql = """
            select * from product_pickmatter where 2 > 1 {0} {1} 
            """
            sql = sql.format(materials_person_sql, product_plan_code_sql)
            drs = db.execute_sql(conn, sql)

            if len(drs) > 0:
                if drs[0][0]:
                    for dr in drs:
                        dict_data = {}
                        dict_data["materials_production_code"] = dr[0]
                        dict_data["materials_person"] = dr[1]
                        dict_data["product_plan_code"] = dr[2]
                        s = dr[3]
                        if s:
                            s1 = s.strftime("%Y-%m-%d %H:%M:%S ")
                        else:
                            s1 = ''
                        dict_data["material_time"] = s1
                        dict_data["description"] = dr[4]
                        sql_matter = """
                        select * from pick_matter where materials_production_code = '{0}'
                        """
                        sql_matter_format = sql_matter.format(dr[0])
                        matter_dfs = db.execute_sql(conn, sql_matter_format)
                        if len(matter_dfs) > 0:
                            respose_list = []
                            if matter_dfs[0][0]:
                                for matter_df in matter_dfs:
                                    response_data = {}
                                    response_data["materials_code"] = matter_df[0]
                                    response_data["materials_production_code"] = matter_df[1]
                                    response_data["matter_code"] = matter_df[2]
                                    response_data["matter_count"] = matter_df[3]
                                    respose_list.append(response_data)
                                dict_data["response_datas"] = respose_list
                                data_list.append(dict_data)
                            else:
                                dict_data["response_datas"] = []
                                data_list.append(dict_data)

                        else:
                            dict_data["response_datas"] = []
                            data_list.append(dict_data)

                    page_result = Page(page, page_size, data_list)
                    data = page_result.get_str_json()

                    sql_num_int = int(len(data_list))
                    data_add_int["data"] = data
                    data_add_int["total"] = sql_num_int
                    result = {"code": 0, "message": "????????????", "data": data_add_int}
                    result = json.dumps(result)
                    logger.info("????????????????????????")
                    db.close_connection(conn)
                    return HttpResponse(result)
                else:
                    data_add_int["data"] = []
                    data_add_int["total"] = 0
                    result = {"code": 0, "message": "????????????,????????????", "data": data_add_int}
                    result = json.dumps(result)
                    logger.info("????????????????????????,??????????????????")
                    return HttpResponse(result)

            else:
                data_add_int["data"] = []
                data_add_int["total"] = 0
                result = {"code": 0, "message": "????????????,????????????", "data": data_add_int}
                result = json.dumps(result)
                logger.info("????????????????????????,??????????????????")
                return HttpResponse(result)

        except Exception as e:

            print(e)
            result = {"code": 1, "message": "????????????", "data": e}
            result = json.dumps(result)
            logger.error("????????????????????????%s" % e)
            return HttpResponse(result)

    def post(self, request):
        try:

            json_data = request.body
            str_data = json.loads(json_data)

            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            sql_matter = """
                   SELECT max(order_number) FROM product_pickmatter;
                   """
            matter_type = db.execute_sql(conn, sql_matter)
            if matter_type[0][0] == None:
                matter_num = 1
            else:
                matter_num = matter_type[0][0] + 1

            sql = """
                  SELECT max(order_number) FROM pick_matter;
                  """
            product_type = db.execute_sql(conn, sql)
            if product_type[0][0] == None:
                id_num = 0
            else:
                id_num = product_type[0][0]

            materials_production_code = str("Im_Product_Pick_Matter_" + str(matter_num))
            materials_person = str_data.get("materials_person")
            product_plan_code = str_data.get("product_plan_code")
            material_time = str_data.get("material_time")
            description = str_data.get("description")
            response_datas = str_data.get("response_datas")

            sql_insert = """
                               insert into product_pickmatter(materials_production_code, materials_person, 
                               product_plan_code,
                               material_time, description, order_number)
                               values('{0}', '{1}', '{2}', '{3}', '{4}', '{5}')
                               """
            sql_insert_format = sql_insert.format(materials_production_code, materials_person,
                                                  product_plan_code, material_time,
                                                  description, matter_num)
            drs = db.execute_sql(conn, sql_insert_format)
            for response_data in response_datas:
                # db = DB()
                # conn = db.get_connection("db_common")
                matter_code = response_data.get("matter_code")
                matter_count = response_data.get("matter_count")
                id_num = id_num + 1
                materials_code = str("Im_Materials_Pick_" + str(id_num))
                sql_response = """
                               insert into pick_matter(materials_code, materials_production_code,matter_code, matter_count,
                               order_number)
                               values('{0}', '{1}', '{2}', '{3}', '{4}')
                       """
                sql_response_format = sql_response.format(materials_code, materials_production_code,
                                                          matter_code, matter_count,
                                                          id_num)
                matter_drs = db.execute_sql(conn, sql_response_format)
            data = {"code": 0, "message": "?????????????????????"}
            logger.info("?????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)

            return HttpResponse(data)
        except Exception as e:
            print(e)
            data = {"code": 1, "message": "?????????????????????"}
            logger.error("?????????????????????")
            data = json.dumps(data)
            return HttpResponse(data)


class PutManagePickMatter(View):
    """
    ???????????????????????????
    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            materials_production_code = request.GET.get("materials_production_code")
            materials_person = request.GET.get("materials_person")
            product_plan_code = request.GET.get("product_plan_code")
            material_time = request.GET.get("material_time")
            description = request.GET.get("description")

            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            else:
                response_datas = response_datas

            sql = """
                      update product_pickmatter set 
                      materials_person='{0}',product_plan_code='{1}', material_time ='{2}',
                      description= '{3}' where materials_production_code = '{4}'
                      """
            sql_main = sql.format(materials_person, product_plan_code, material_time, description,
                                  materials_production_code)
            drs = db.execute_sql(conn, sql_main)

            sql_matter_num = """
             SELECT max(order_number) FROM pick_matter
            """
            dras = db.execute_sql(conn, sql_matter_num)

            if dras[0][0] != None:
                sql_num = dras[0][0]
            else:
                sql_num = 0

            for response_data in response_datas:
                # materials_production_code = response_data.get("materials_production_code")
                materials_code = response_data.get("materials_code")
                # materials_production_code = response_data.get("materials_production_code")
                matter_code = response_data.get("matter_code")
                matter_count = response_data.get("matter_count")
                if materials_code:
                    sql = """
                       update pick_matter set materials_production_code = '{0}', 
                       matter_code = '{1}', matter_count ='{2}' where materials_code = '{3}'
                    """
                    sql = sql.format(materials_production_code, matter_code, matter_count, materials_code)
                    db.execute_sql(conn, sql)
                else:
                    sql_num = sql_num + 1

                    materials_code = str("Im_Materials_Pick_" + str(sql_num))

                    sql_response = """
                                          insert into pick_matter(materials_code,materials_production_code, 
                                          matter_code, matter_count, order_number) values('{0}',
                                           '{1}', '{2}', '{3}', '{4}')
                                  """
                    sql_response_format = sql_response.format(materials_code, materials_production_code,
                                                              matter_code, matter_count, sql_num)
                    db.execute_sql(conn, sql_response_format)

            data = {"code": 0, "message": "????????????"}

            data = json.dumps(data)
            logger.info("???????????????????????????")
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            logger.error("???????????????????????????---->%s"% e)
            data = json.dumps(data)
            return HttpResponse(data)


class DeleteManagePickMatter(View):
    """
    ???????????????????????????????????????

    """

    def get(self, request):
        try:
            db = DB()
            product_id = request.session.get("session_projectId")
            conn = db.get_connection(product_id)

            materials_production_code = request.GET.get("materials_production_code")
            response_datas = request.GET.get("response_datas")
            if response_datas:
                response_datas = eval(response_datas)
            if response_datas:
                for response_data in response_datas:
                    materials_code = dict(response_data).get("materials_code")
                    sql = """
                       delete from pick_matter where materials_code = '{0}'
                       """
                    sql_format = sql.format(materials_code)
                    db.execute_sql(conn, sql_format)
            else:
                sql_matter = """
                                  SELECT * FROM product_pickmatter where materials_production_code = '{0}'
                                  """
                sql_matter_format = sql_matter.format(materials_production_code)
                matter_dfs = db.execute_sql(conn, sql_matter_format)
                for matter_df in matter_dfs:
                    materials_code = matter_df[0]
                    sql_delete = """
                       delete FROM pick_matter where materials_code = '{0}'
                       """
                    sql_delete_format = sql_delete.format(materials_code)
                    db.execute_sql(conn, sql_delete_format)
                sql = """
                       delete from product_pickmatter where materials_production_code = '{0}'
                       """
                sql_main = sql.format(materials_production_code)
                des = db.execute_sql(conn, sql_main)

            data = {"code": 0, "message": "????????????"}
            logger.info("???????????????????????????????????????")
            data = json.dumps(data)
            db.close_connection(conn)
            return HttpResponse(data)

        except Exception as e:
            print(e)
            data = {"code": 1, "message": "????????????"}
            data = json.dumps(data)
            logger.error("???????????????????????????????????????,%s" % e)
            return HttpResponse(data)











































































































































